"""
BAS Reviewer Web App - Xero Integration
"""
import os
import json
import re
import requests
from flask import Flask, redirect, request, session, url_for, render_template, send_file, jsonify
from datetime import datetime, timedelta
import pandas as pd
from io import BytesIO
import secrets
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask_login import LoginManager, login_required, current_user

app = Flask(__name__)

# Security Configuration
app.secret_key = os.environ.get('SECRET_KEY')
if not app.secret_key:
    raise RuntimeError("SECRET_KEY environment variable must be set")

app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file upload

# Enable debug mode only in development
app.config['DEBUG'] = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
app.config['PROPAGATE_EXCEPTIONS'] = app.config['DEBUG']

# Initialize Flask-Limiter for rate limiting
try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    limiter = Limiter(
        app=app,
        key_func=get_remote_address,
        default_limits=["200 per day", "50 per hour"],
        storage_uri="memory://",
    )
except ImportError:
    limiter = None
    print("Warning: flask-limiter not installed, rate limiting disabled")

# Security headers middleware
@app.after_request
def add_security_headers(response):
    # Content Security Policy
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline'; "
        "img-src 'self' data:; "
        "font-src 'self'; "
        "connect-src 'self' https://api.deepseek.com"
    )
    # Other security headers
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    if not app.config['DEBUG']:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response

# Global error handler - sanitize error responses in production
@app.errorhandler(Exception)
def handle_exception(e):
    app.logger.error(f"Unhandled exception: {e}", exc_info=True)
    if app.config['DEBUG']:
        raise e
    return render_template('error.html', error="An unexpected error occurred. Please try again."), 500

# Database Configuration
from config import get_config
app.config.from_object(get_config())

# Initialize database
from models import db, User
db.init_app(app)

# Create tables on first request
with app.app_context():
    db.create_all()

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'auth.login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'info'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(user_id)

@login_manager.unauthorized_handler
def unauthorized():
    """Handle unauthorized access - return JSON for API routes, redirect otherwise"""
    from flask import request, jsonify
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Please log in to access this feature'}), 401
    return redirect(url_for('auth.login'))

# Register blueprints
from blueprints.auth import auth_bp
app.register_blueprint(auth_bp, url_prefix='/auth')

# Xero OAuth2 Configuration
XERO_CLIENT_ID = os.environ.get('XERO_CLIENT_ID')
XERO_CLIENT_SECRET = os.environ.get('XERO_CLIENT_SECRET')
XERO_REDIRECT_URI = os.environ.get('XERO_REDIRECT_URI', 'https://bas-reviewer.up.railway.app/callback')
XERO_SCOPES = 'openid profile email accounting.transactions accounting.transactions.read accounting.journals.read accounting.reports.read accounting.settings.read offline_access'

# DeepSeek API
DEEPSEEK_API_KEY = os.environ.get('DEEPSEEK_API_KEY')

# Cloudflare R2 Configuration (S3-compatible storage for training data)
R2_ACCOUNT_ID = os.environ.get('R2_ACCOUNT_ID')
R2_ACCESS_KEY_ID = os.environ.get('R2_ACCESS_KEY_ID')
R2_SECRET_ACCESS_KEY = os.environ.get('R2_SECRET_ACCESS_KEY')
R2_BUCKET_NAME = os.environ.get('R2_BUCKET_NAME', 'bas-reviewer-uploads')

def upload_to_r2(file_data, filename):
    """Upload file to Cloudflare R2 for training data collection"""
    if not all([R2_ACCOUNT_ID, R2_ACCESS_KEY_ID, R2_SECRET_ACCESS_KEY]):
        print("R2 credentials not configured - skipping file upload")
        return None

    try:
        import boto3
        from botocore.config import Config

        # Create S3 client for R2
        s3_client = boto3.client(
            's3',
            endpoint_url=f'https://{R2_ACCOUNT_ID}.r2.cloudflarestorage.com',
            aws_access_key_id=R2_ACCESS_KEY_ID,
            aws_secret_access_key=R2_SECRET_ACCESS_KEY,
            config=Config(signature_version='s3v4'),
            region_name='auto'
        )

        # Generate unique filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{timestamp}_{filename}"

        # Upload file
        s3_client.upload_fileobj(
            file_data,
            R2_BUCKET_NAME,
            unique_filename,
            ExtraArgs={'ContentType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}
        )

        print(f"File uploaded to R2: {unique_filename}")
        return unique_filename

    except Exception as e:
        print(f"Error uploading to R2: {e}")
        return None

# Xero OAuth URLs
XERO_AUTH_URL = 'https://login.xero.com/identity/connect/authorize'
XERO_TOKEN_URL = 'https://identity.xero.com/connect/token'
XERO_CONNECTIONS_URL = 'https://api.xero.com/connections'
XERO_API_URL = 'https://api.xero.com/api.xro/2.0'


@app.route('/')
def index():
    """Home page - redirect authenticated users to dashboard"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('index.html', logged_in=False, tenant_name='')


@app.route('/dashboard')
@login_required
def dashboard():
    """User dashboard - shows options and review history"""
    # Check if Xero is connected
    xero_connected = False
    tenant_name = session.get('tenant_name', '')
    if 'access_token' in session:
        token_expiry = session.get('token_expiry')
        if token_expiry:
            try:
                expiry_dt = datetime.fromisoformat(token_expiry)
                xero_connected = expiry_dt > datetime.now()
            except:
                xero_connected = True
        else:
            xero_connected = True

    return render_template('dashboard/index.html',
                         user=current_user,
                         xero_connected=xero_connected,
                         tenant_name=tenant_name)


@app.route('/xero/login')
@login_required
def xero_login():
    """Redirect to Xero OAuth - requires user to be logged in first"""
    # Debug: Check if credentials are loaded
    if not XERO_CLIENT_ID:
        return render_template('error.html', error="XERO_CLIENT_ID is not set in environment variables")
    if not XERO_CLIENT_SECRET:
        return render_template('error.html', error="XERO_CLIENT_SECRET is not set in environment variables")

    state = secrets.token_hex(16)
    session['oauth_state'] = state

    auth_url = (
        f"{XERO_AUTH_URL}?"
        f"response_type=code&"
        f"client_id={XERO_CLIENT_ID}&"
        f"redirect_uri={XERO_REDIRECT_URI}&"
        f"scope={XERO_SCOPES}&"
        f"state={state}"
    )
    return redirect(auth_url)


# Debug endpoints removed for security - use environment variables to check config


@app.route('/callback')
def callback():
    """Handle Xero OAuth callback"""
    try:
        error = request.args.get('error')
        error_description = request.args.get('error_description', '')
        if error:
            return render_template('error.html', error=f"Xero authentication failed: {error} - {error_description}")

        code = request.args.get('code')
        state = request.args.get('state')

        if not code:
            return render_template('error.html', error="No authorization code received from Xero")

        # Validate OAuth state to prevent CSRF attacks
        stored_state = session.pop('oauth_state', None)
        if not stored_state or stored_state != state:
            return render_template('error.html', error="Invalid OAuth state. Please try logging in again.")

        # Exchange code for tokens
        token_data = {
            'grant_type': 'authorization_code',
            'code': code,
            'redirect_uri': XERO_REDIRECT_URI
        }

        try:
            response = requests.post(
                XERO_TOKEN_URL,
                data=token_data,
                auth=(XERO_CLIENT_ID, XERO_CLIENT_SECRET),
                headers={'Content-Type': 'application/x-www-form-urlencoded'},
                timeout=30
            )
        except requests.exceptions.RequestException as req_err:
            return render_template('error.html', error=f"Network error during token exchange: {str(req_err)}")

        if response.status_code != 200:
            return render_template('error.html', error=f"Token exchange failed (Status {response.status_code}): {response.text}")

        try:
            tokens = response.json()
        except Exception as json_err:
            return render_template('error.html', error=f"Failed to parse token response: {str(json_err)}")

        session['access_token'] = tokens.get('access_token')
        session['refresh_token'] = tokens.get('refresh_token')

        expires_in = tokens.get('expires_in', 1800)
        session['token_expiry'] = (datetime.now() + timedelta(seconds=expires_in)).isoformat()

        # Get connected tenants
        try:
            connections = get_xero_connections()
            if connections:
                session['tenant_id'] = connections[0]['tenantId']
                session['tenant_name'] = connections[0]['tenantName']
        except Exception as conn_err:
            return render_template('error.html', error=f"Failed to get Xero connections: {str(conn_err)}")

        return redirect(url_for('dashboard'))
    except Exception as e:
        import traceback
        return render_template('error.html', error=f"Callback error: {str(e)}<br><pre>{traceback.format_exc()}</pre>")


@app.route('/xero/disconnect')
@login_required
def xero_disconnect():
    """Disconnect from Xero (clear Xero tokens only)"""
    # Remove Xero-specific session data, keep user login
    session.pop('access_token', None)
    session.pop('refresh_token', None)
    session.pop('token_expiry', None)
    session.pop('tenant_id', None)
    session.pop('tenant_name', None)
    return redirect(url_for('dashboard'))


def parse_xero_date(raw_date):
    """Parse Xero date format to datetime object"""
    if not raw_date:
        return None

    raw_date_str = str(raw_date)

    # Xero returns dates like "/Date(1234567890000)/"
    if '/Date(' in raw_date_str:
        match = re.search(r'/Date\((\d+)', raw_date_str)
        if match:
            timestamp = int(match.group(1)) / 1000
            return datetime.fromtimestamp(timestamp)

    # ISO format or string - try to parse as date
    try:
        date_str = raw_date_str[:10]
        return datetime.strptime(date_str, '%Y-%m-%d')
    except:
        return None


def get_xero_connections():
    """Get list of connected Xero organisations"""
    headers = {
        'Authorization': f"Bearer {session['access_token']}",
        'Content-Type': 'application/json'
    }
    response = requests.get(XERO_CONNECTIONS_URL, headers=headers)
    if response.status_code == 200:
        return response.json()
    return []


def refresh_token_if_needed():
    """Refresh access token if expired"""
    token_expiry = session.get('token_expiry')
    is_expired = True

    if token_expiry:
        try:
            if isinstance(token_expiry, str):
                expiry_dt = datetime.fromisoformat(token_expiry)
            else:
                expiry_dt = token_expiry
            is_expired = expiry_dt < datetime.now()
        except:
            is_expired = False  # If can't parse, assume valid

    if is_expired:
        if 'refresh_token' not in session:
            return False

        token_data = {
            'grant_type': 'refresh_token',
            'refresh_token': session['refresh_token']
        }

        response = requests.post(
            XERO_TOKEN_URL,
            data=token_data,
            auth=(XERO_CLIENT_ID, XERO_CLIENT_SECRET),
            headers={'Content-Type': 'application/x-www-form-urlencoded'}
        )

        if response.status_code == 200:
            tokens = response.json()
            session['access_token'] = tokens['access_token']
            session['refresh_token'] = tokens.get('refresh_token', session['refresh_token'])
            session['token_expiry'] = (datetime.now() + timedelta(seconds=tokens['expires_in'])).isoformat()
            return True
        return False
    return True


def xero_api_request(endpoint, params=None):
    """Make authenticated request to Xero API"""
    if not refresh_token_if_needed():
        print(f"DEBUG xero_api_request: Token refresh failed for {endpoint}")
        return None

    headers = {
        'Authorization': f"Bearer {session['access_token']}",
        'Xero-tenant-id': session['tenant_id'],
        'Content-Type': 'application/json',
        'Accept': 'application/json'
    }

    url = f"{XERO_API_URL}/{endpoint}"
    print(f"DEBUG xero_api_request: Calling {url} with params={params}")
    response = requests.get(url, headers=headers, params=params)
    print(f"DEBUG xero_api_request: Response status={response.status_code}")

    if response.status_code == 200:
        return response.json()
    else:
        print(f"DEBUG xero_api_request: Error response: {response.text[:500]}")
    return None


def push_manual_journal_to_xero(journal_data):
    """Push a manual journal to Xero as DRAFT status"""
    if not refresh_token_if_needed():
        return {'success': False, 'error': 'Token refresh failed'}

    headers = {
        'Authorization': f"Bearer {session['access_token']}",
        'Xero-tenant-id': session['tenant_id'],
        'Content-Type': 'application/json',
        'Accept': 'application/json'
    }

    url = f"{XERO_API_URL}/ManualJournals"

    # Format journal for Xero API
    # Xero ManualJournal format:
    # {
    #   "Narration": "Description",
    #   "Status": "DRAFT",
    #   "JournalLines": [
    #     {"AccountCode": "200", "Description": "...", "LineAmount": 100.00},
    #     {"AccountCode": "800", "Description": "...", "LineAmount": -100.00}
    #   ]
    # }

    payload = {
        "ManualJournals": [{
            "Narration": journal_data.get('narration', 'BAS Review Correcting Entry'),
            "Status": "DRAFT",
            "Date": journal_data.get('date', datetime.now().strftime('%Y-%m-%d')),
            "LineAmountTypes": "Inclusive",
            "JournalLines": journal_data.get('entries', [])
        }]
    }

    try:
        print(f"DEBUG push_manual_journal: Sending payload: {json.dumps(payload, indent=2)}")
        response = requests.post(url, headers=headers, json=payload)
        print(f"DEBUG push_manual_journal: Response status={response.status_code}")
        print(f"DEBUG push_manual_journal: Full response body={response.text}")

        if response.status_code in [200, 201]:
            result = response.json()
            journals = result.get('ManualJournals', [])
            if journals:
                return {
                    'success': True,
                    'journal_id': journals[0].get('ManualJournalID'),
                    'journal_number': journals[0].get('JournalNumber'),
                    'status': journals[0].get('Status')
                }
            return {'success': True, 'message': 'Journal created'}
        else:
            error_msg = f"Status {response.status_code}: "
            try:
                error_json = response.json()
                print(f"DEBUG push_manual_journal: Error JSON: {json.dumps(error_json, indent=2)}")
                # Try to get validation errors from Elements
                if 'Elements' in error_json and error_json['Elements']:
                    for element in error_json['Elements']:
                        validation_errors = element.get('ValidationErrors', [])
                        if validation_errors:
                            error_msg += '; '.join([e.get('Message', str(e)) for e in validation_errors])
                            break
                # Fall back to Message field
                if error_msg == f"Status {response.status_code}: " and 'Message' in error_json:
                    error_msg += error_json['Message']
                # Fall back to raw response
                if error_msg == f"Status {response.status_code}: ":
                    error_msg += response.text[:500]
            except Exception as parse_err:
                print(f"DEBUG push_manual_journal: Error parsing response: {parse_err}")
                error_msg += response.text[:500]
            return {'success': False, 'error': error_msg}
    except Exception as e:
        print(f"DEBUG push_manual_journal: Exception: {str(e)}")
        return {'success': False, 'error': str(e)}


@app.route('/api/push-journal', methods=['POST'])
@login_required
def push_journal():
    """API endpoint to push a correcting journal to Xero as DRAFT"""
    try:
        if 'access_token' not in session:
            return jsonify({'success': False, 'error': 'Not authenticated with Xero'}), 401

        data = request.json
        if not data:
            return jsonify({'success': False, 'error': 'No journal data provided'}), 400

        # Validate journal data
        narration = data.get('narration', '')
        entries = data.get('entries', [])

        if not entries or len(entries) < 2:
            return jsonify({'success': False, 'error': 'Journal must have at least 2 line items'}), 400

        # Check that debits equal credits
        total = sum(float(e.get('LineAmount', 0)) for e in entries)
        if abs(total) > 0.01:  # Allow small rounding differences
            return jsonify({'success': False, 'error': f'Journal does not balance. Difference: {total}'}), 400

        # Push to Xero
        result = push_manual_journal_to_xero({
            'narration': narration,
            'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
            'entries': entries
        })

        if result.get('success'):
            return jsonify(result)
        else:
            return jsonify(result), 400

    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


# ============== Cloudflare D1 API Endpoints ==============

@app.route('/api/init-d1', methods=['POST'])
@login_required
def init_d1_tables():
    """Initialize Cloudflare D1 tables (admin only)"""
    try:
        result = init_cloudflare_d1_tables()
        return jsonify({'success': True, 'result': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/save-review', methods=['POST'])
@login_required
def save_review():
    """Save a BAS review to Cloudflare D1"""
    try:
        data = request.json
        if not data:
            return jsonify({'success': False, 'error': 'No review data provided'}), 400

        # Try to fetch BAS report from Xero if connected
        bas_report_data = data.get('bas_report_data', {})
        if 'access_token' in session and not bas_report_data:
            try:
                bas_report = fetch_xero_bas_report()
                if bas_report:
                    bas_report_data = bas_report
            except Exception as e:
                print(f"Could not fetch BAS report: {e}")

        # Add tenant and user info from session
        review_data = {
            'tenant_id': session.get('tenant_id', ''),
            'tenant_name': session.get('tenant_name', ''),
            'user_email': current_user.email if current_user.is_authenticated else '',
            'period_start': data.get('period_start', ''),
            'period_end': data.get('period_end', ''),
            'total_transactions': data.get('total_transactions', 0),
            'flagged_count': data.get('flagged_count', 0),
            'high_severity_count': data.get('high_severity_count', 0),
            'medium_severity_count': data.get('medium_severity_count', 0),
            'low_severity_count': data.get('low_severity_count', 0),
            'bas_report_data': bas_report_data,
            'review_summary': data.get('review_summary', {}),
            'flagged_items': data.get('flagged_items', []),
            'all_transactions': data.get('all_transactions', [])
        }

        result = save_review_to_d1(review_data)
        return jsonify(result)

    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/reviews')
@login_required
def get_reviews():
    """Get list of saved reviews from Cloudflare D1"""
    try:
        tenant_id = session.get('tenant_id')
        limit = request.args.get('limit', 50, type=int)

        reviews = get_reviews_from_d1(tenant_id=tenant_id, limit=limit)
        return jsonify({'success': True, 'reviews': reviews})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reviews/<statement_id>')
@login_required
def get_review_detail(statement_id):
    """Get detailed review from Cloudflare D1"""
    try:
        review = get_review_details_from_d1(statement_id)
        if not review:
            return jsonify({'success': False, 'error': 'Review not found'}), 404

        return jsonify({'success': True, 'review': review})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/bas-report')
@login_required
def get_bas_report():
    """Get BAS Report from Xero API"""
    try:
        if 'access_token' not in session:
            return jsonify({'success': False, 'error': 'Not authenticated with Xero'}), 401

        report_id = request.args.get('report_id')
        bas_data = fetch_xero_bas_report(report_id)

        if bas_data:
            return jsonify({'success': True, 'bas_report': bas_data})
        else:
            return jsonify({'success': False, 'error': 'No BAS report data available'}), 404

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/bas-reports')
@login_required
def get_bas_reports_list():
    """Get list of available BAS reports from Xero"""
    try:
        if 'access_token' not in session:
            return jsonify({'success': False, 'error': 'Not authenticated with Xero'}), 401

        reports = fetch_xero_bas_report_list()
        return jsonify({'success': True, 'reports': reports})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/review')
@login_required
def review_page():
    """Show review options page (requires Xero connection)"""
    if 'access_token' not in session:
        return redirect(url_for('xero_login'))

    return render_template('review.html', tenant_name=session.get('tenant_name', ''), logged_in=True)


@app.route('/upload')
@login_required
def upload_page():
    """Show upload page for authenticated users"""
    return render_template('review.html', tenant_name='', logged_in=False)


def parse_activity_statement(raw_data):
    """Parse Activity Statement (Transactions by Tax Rate) Excel format"""
    transactions = []

    # Activity Statement format:
    # Row 0: "Transactions by Tax Rate"
    # Row 1: Company name
    # Row 2: Period
    # Row 4: Headers (Date, Account, Reference, Details, Gross, GST, Net)
    # Row 6+: Data grouped by tax type sections

    company_name = str(raw_data.iloc[1, 0]) if len(raw_data) > 1 and not pd.isna(raw_data.iloc[1, 0]) else 'Unknown'
    period = str(raw_data.iloc[2, 0]) if len(raw_data) > 2 and not pd.isna(raw_data.iloc[2, 0]) else ''

    current_tax_type = ''

    def parse_amount(val):
        if pd.isna(val):
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        val_str = str(val).replace('$', '').replace(',', '').replace('(', '-').replace(')', '').strip()
        try:
            return float(val_str)
        except:
            return 0.0

    def parse_account(account_str):
        """Parse account string like 'Sales (200)' into name and code"""
        if pd.isna(account_str) or not account_str:
            return '', ''
        account_str = str(account_str).strip()
        # Look for pattern: "Account Name (CODE)"
        import re
        match = re.match(r'^(.+?)\s*\((\d+)\)$', account_str)
        if match:
            return match.group(1).strip(), match.group(2)
        return account_str, ''

    for idx in range(5, len(raw_data)):
        row = raw_data.iloc[idx]
        first_cell = row.iloc[0] if not pd.isna(row.iloc[0]) else ''
        first_cell_str = str(first_cell).strip()

        # Detect section headers (tax types)
        if first_cell_str in ['GST on Income', 'GST on Expenses', 'GST Free Expenses', 'GST Free Income',
                              'BAS Excluded', 'GST Free', 'Input Taxed', 'Export']:
            current_tax_type = first_cell_str
            continue

        # Skip total rows and empty rows
        if first_cell_str.startswith('Total') or not first_cell_str:
            continue

        # Try to parse as date (DD/MM/YYYY format)
        try:
            date_val = pd.to_datetime(first_cell, dayfirst=True)
            date_str = date_val.strftime('%Y-%m-%d')
        except:
            continue  # Not a transaction row

        # Parse the row: Date, Account, Reference, Details, Gross, GST, Net
        account_str = str(row.iloc[1]) if not pd.isna(row.iloc[1]) else ''
        account_name, account_code = parse_account(account_str)
        reference = str(row.iloc[2]) if len(row) > 2 and not pd.isna(row.iloc[2]) else ''
        details = str(row.iloc[3]) if len(row) > 3 and not pd.isna(row.iloc[3]) else ''
        gross = parse_amount(row.iloc[4] if len(row) > 4 else 0)
        gst = parse_amount(row.iloc[5] if len(row) > 5 else 0)
        net = parse_amount(row.iloc[6] if len(row) > 6 else 0)

        # Skip zero transactions
        if gross == 0 and gst == 0 and net == 0:
            continue

        # Determine transaction type from tax type section
        if 'Income' in current_tax_type:
            tx_type = 'income'
        elif 'Expense' in current_tax_type:
            tx_type = 'expense'
        else:
            tx_type = 'other'

        # Map tax type to GST rate name
        gst_rate_name = ''
        if current_tax_type == 'GST on Income':
            gst_rate_name = 'GST on Income'
        elif current_tax_type == 'GST on Expenses':
            gst_rate_name = 'GST on Expenses'
        elif 'GST Free' in current_tax_type:
            gst_rate_name = 'GST Free'
        elif current_tax_type == 'BAS Excluded':
            gst_rate_name = 'BAS Excluded'

        transaction = {
            'row_number': idx + 1,
            'account_code': account_code,
            'account': account_name,
            'date': date_str,
            'source': '',
            'description': details,
            'invoice_number': '',
            'reference': reference,
            'gross': abs(gross),
            'gst': abs(gst),
            'net': abs(net),
            'gst_rate': 10 if 'GST on' in current_tax_type else 0,
            'gst_rate_name': gst_rate_name,
            'type': tx_type,
            'tax_type_section': current_tax_type
        }

        transactions.append(transaction)

    return transactions, company_name, period


def fetch_xero_bank_transactions(from_date_str, to_date_str):
    """Fetch bank transactions from Xero API with tax information"""
    transactions = []

    # Format dates for Xero Where clause: DateTime(year, month, day)
    from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
    to_date = datetime.strptime(to_date_str, '%Y-%m-%d')

    where_clause = f'Date >= DateTime({from_date.year},{from_date.month},{from_date.day}) AND Date <= DateTime({to_date.year},{to_date.month},{to_date.day})'

    page = 1
    while True:
        data = xero_api_request('BankTransactions', params={
            'where': where_clause,
            'page': page
        })

        if not data or 'BankTransactions' not in data:
            break

        bank_txns = data.get('BankTransactions', [])
        if not bank_txns:
            break

        for txn in bank_txns:
            txn_date = parse_xero_date(txn.get('Date', ''))
            date_str = txn_date.strftime('%Y-%m-%d') if txn_date else 'NO DATE'
            # Check if amounts are tax inclusive or exclusive
            line_amount_types = txn.get('LineAmountTypes', 'Exclusive')
            # Get bank transaction ID for Xero URL
            bank_txn_id = txn.get('BankTransactionID', '')
            xero_url = f"https://go.xero.com/Bank/ViewTransaction.aspx?bankTransactionID={bank_txn_id}" if bank_txn_id else ''

            # Each bank transaction can have multiple line items
            for line in txn.get('LineItems', []):
                account_code = line.get('AccountCode', '')
                # Get account name from chart of accounts if needed
                description = line.get('Description', '') or txn.get('Reference', '') or 'No description'

                # Amounts
                line_amount = float(line.get('LineAmount', 0) or 0)
                tax_amount = float(line.get('TaxAmount', 0) or 0)

                # Calculate gross and net based on LineAmountTypes
                if line_amount_types == 'Inclusive':
                    # LineAmount already includes tax
                    gross = line_amount
                    net = line_amount - tax_amount
                else:
                    # LineAmount is exclusive of tax (default)
                    gross = line_amount + tax_amount
                    net = line_amount

                # Determine transaction type
                is_expense = txn.get('Type', '') == 'SPEND'

                # Get tax type
                tax_type = line.get('TaxType', '')
                gst_rate_name = ''
                if 'OUTPUT' in tax_type.upper():
                    gst_rate_name = 'GST on Income'
                elif 'INPUT' in tax_type.upper():
                    gst_rate_name = 'GST on Expenses'
                elif 'NONE' in tax_type.upper() or 'EXEMPT' in tax_type.upper():
                    gst_rate_name = 'GST Free'
                elif 'BASEXCLUDED' in tax_type.upper():
                    gst_rate_name = 'BAS Excluded'
                else:
                    gst_rate_name = tax_type

                # Determine source type based on transaction type and sign
                txn_type = txn.get('Type', '')
                is_refund = gross < 0
                if txn_type == 'SPEND':
                    source = 'Refund Received' if is_refund else 'Bank Payment'
                elif txn_type == 'RECEIVE':
                    source = 'Refund Given' if is_refund else 'Bank Receipt'
                else:
                    source = txn_type

                transactions.append({
                    'row_number': len(transactions) + 1,
                    'date': date_str,
                    'type': 'expense' if is_expense else 'income',
                    'account_code': account_code,
                    'account': line.get('AccountCode', ''),  # Will be enriched later
                    'description': description,
                    'gross': gross,  # Preserve sign for refunds
                    'gst': tax_amount,  # Preserve sign for refunds
                    'net': net,  # Preserve sign for refunds
                    'gst_rate_name': gst_rate_name,
                    'source': source,
                    'reference': txn.get('Reference', ''),
                    'contact': txn.get('Contact', {}).get('Name', '') if txn.get('Contact') else '',
                    'xero_url': xero_url
                })

        page += 1
        # Xero returns max 100 per page
        if len(bank_txns) < 100:
            break

    return transactions


def fetch_xero_invoices(from_date_str, to_date_str, invoice_type='ACCREC'):
    """Fetch invoices from Xero API (ACCREC for sales, ACCPAY for bills)

    Only includes AUTHORISED and PAID invoices/bills - excludes DRAFT and SUBMITTED
    as these haven't been posted and won't affect the BAS.
    """
    transactions = []

    from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
    to_date = datetime.strptime(to_date_str, '%Y-%m-%d')

    # Only include approved invoices/bills (AUTHORISED or PAID)
    # Exclude DRAFT and SUBMITTED as they haven't been posted yet
    where_clause = f'Date >= DateTime({from_date.year},{from_date.month},{from_date.day}) AND Date <= DateTime({to_date.year},{to_date.month},{to_date.day}) AND Type == "{invoice_type}" AND (Status == "AUTHORISED" OR Status == "PAID")'

    page = 1
    while True:
        # Request invoices with full line item details
        data = xero_api_request('Invoices', params={
            'where': where_clause,
            'page': page,
            'includeArchived': 'false'
        })

        if not data or 'Invoices' not in data:
            break

        invoices = data.get('Invoices', [])
        if not invoices:
            break

        for inv in invoices:
            inv_date = parse_xero_date(inv.get('Date', ''))
            date_str = inv_date.strftime('%Y-%m-%d') if inv_date else 'NO DATE'
            contact_name = inv.get('Contact', {}).get('Name', '') if inv.get('Contact') else ''
            inv_number = inv.get('InvoiceNumber', '')
            inv_id = inv.get('InvoiceID', '')
            # Check if amounts are tax inclusive or exclusive
            line_amount_types = inv.get('LineAmountTypes', 'Exclusive')

            # Generate Xero URL based on invoice type
            if invoice_type == 'ACCPAY':
                xero_url = f"https://go.xero.com/AccountsPayable/Edit.aspx?InvoiceID={inv_id}" if inv_id else ''
            else:
                xero_url = f"https://go.xero.com/AccountsReceivable/Edit.aspx?InvoiceID={inv_id}" if inv_id else ''

            # Each invoice can have multiple line items
            for line in inv.get('LineItems', []):
                account_code = line.get('AccountCode', '')
                description = line.get('Description', '') or f'Invoice {inv_number}'

                line_amount = float(line.get('LineAmount', 0) or 0)
                tax_amount = float(line.get('TaxAmount', 0) or 0)

                # Calculate gross and net based on LineAmountTypes
                if line_amount_types == 'Inclusive':
                    # LineAmount already includes tax
                    gross = line_amount
                    net = line_amount - tax_amount
                else:
                    # LineAmount is exclusive of tax (default)
                    gross = line_amount + tax_amount
                    net = line_amount

                is_expense = invoice_type == 'ACCPAY'

                tax_type = line.get('TaxType', '')
                gst_rate_name = ''
                if 'OUTPUT' in tax_type.upper():
                    gst_rate_name = 'GST on Income'
                elif 'INPUT' in tax_type.upper():
                    gst_rate_name = 'GST on Expenses'
                elif 'NONE' in tax_type.upper() or 'EXEMPT' in tax_type.upper():
                    gst_rate_name = 'GST Free'
                elif 'BASEXCLUDED' in tax_type.upper():
                    gst_rate_name = 'BAS Excluded'
                else:
                    gst_rate_name = tax_type

                # Determine if this is a refund/credit note (negative amount)
                is_refund = gross < 0 or (invoice_type == 'ACCREC' and line_amount < 0)

                transactions.append({
                    'row_number': len(transactions) + 1,
                    'date': date_str,
                    'type': 'expense' if is_expense else 'income',
                    'account_code': account_code,
                    'account': account_code,
                    'description': description,
                    'gross': gross,  # Preserve sign for refunds
                    'gst': tax_amount,  # Preserve sign for refunds
                    'net': net,  # Preserve sign for refunds
                    'gst_rate_name': gst_rate_name,
                    'source': ('Credit Note' if is_refund else 'Invoice') if invoice_type == 'ACCREC' else ('Debit Note' if is_refund else 'Bill'),
                    'reference': inv_number,
                    'contact': contact_name,
                    'xero_url': xero_url
                })

        page += 1
        if len(invoices) < 100:
            break

    return transactions


def fetch_xero_bills(from_date_str, to_date_str):
    """Fetch bills (accounts payable invoices) from Xero API"""
    return fetch_xero_invoices(from_date_str, to_date_str, invoice_type='ACCPAY')


def fetch_xero_bas_report(report_id=None):
    """Fetch BAS (Business Activity Statement) Report from Xero API

    Args:
        report_id: Optional specific BAS report ID. If None, fetches the list of available reports.

    Returns:
        dict: BAS report data including GST amounts, PAYG, etc.
    """
    if not refresh_token_if_needed():
        return None

    # Fetch the BAS report
    if report_id:
        # Get specific BAS report
        data = xero_api_request('Reports/AustralianBASReport', params={'reportID': report_id})
    else:
        # Get list of BAS reports
        data = xero_api_request('Reports/AustralianBASReport')

    if not data or 'Reports' not in data:
        print(f"DEBUG fetch_xero_bas_report: No BAS report data returned")
        return None

    reports = data.get('Reports', [])
    if not reports:
        return None

    report = reports[0]

    # Parse the BAS report into a structured format
    bas_data = {
        'report_id': report.get('ReportID'),
        'report_name': report.get('ReportName'),
        'report_type': report.get('ReportType'),
        'report_date': report.get('ReportDate'),
        'updated_date': report.get('UpdatedDateUTC'),
        'fields': {}
    }

    # Extract BAS fields from rows
    rows = report.get('Rows', [])
    for row in rows:
        row_type = row.get('RowType')
        if row_type == 'Section':
            section_title = row.get('Title', '')
            section_rows = row.get('Rows', [])
            for section_row in section_rows:
                cells = section_row.get('Cells', [])
                if len(cells) >= 2:
                    field_label = cells[0].get('Value', '')
                    field_value = cells[1].get('Value', '')
                    # Extract field code if present (e.g., "G1", "1A", "1B")
                    if field_label:
                        bas_data['fields'][field_label] = field_value

    return bas_data


def fetch_xero_bas_report_list():
    """Fetch list of available BAS reports from Xero"""
    if not refresh_token_if_needed():
        return []

    # The Reports endpoint can return a list when no specific report is requested
    data = xero_api_request('Reports')

    if not data or 'Reports' not in data:
        return []

    # Filter for Australian BAS reports
    bas_reports = []
    for report in data.get('Reports', []):
        if 'BAS' in report.get('ReportName', '') or 'Activity Statement' in report.get('ReportName', ''):
            bas_reports.append({
                'report_id': report.get('ReportID'),
                'report_name': report.get('ReportName'),
                'report_type': report.get('ReportType')
            })

    return bas_reports


# ============== Cloudflare D1 Integration ==============

def get_cloudflare_config():
    """Get Cloudflare D1 configuration from environment variables"""
    return {
        'account_id': os.environ.get('CLOUDFLARE_ACCOUNT_ID', ''),
        'api_token': os.environ.get('CLOUDFLARE_API_TOKEN', ''),
        'database_id': os.environ.get('CLOUDFLARE_D1_DATABASE_ID', '')
    }

def cloudflare_d1_query(sql, params=None):
    """Execute a query on Cloudflare D1 database

    Args:
        sql: SQL query string
        params: Optional list of parameters for the query

    Returns:
        dict: Query results or error
    """
    config = get_cloudflare_config()
    account_id = config['account_id']
    api_token = config['api_token']
    database_id = config['database_id']

    if not all([account_id, api_token, database_id]):
        return {'success': False, 'error': f'Cloudflare D1 not configured. account_id={bool(account_id)}, api_token={bool(api_token)}, database_id={bool(database_id)}'}

    url = f"https://api.cloudflare.com/client/v4/accounts/{account_id}/d1/database/{database_id}/query"

    headers = {
        'Authorization': f'Bearer {api_token}',
        'Content-Type': 'application/json'
    }

    payload = {'sql': sql}
    if params:
        payload['params'] = params

    try:
        response = requests.post(url, headers=headers, json=payload)
        if response.status_code == 200:
            return response.json()
        else:
            return {'success': False, 'error': f'D1 query failed: {response.text}'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def init_cloudflare_d1_tables():
    """Initialize Cloudflare D1 tables for storing reviewed activity statements"""

    # Create reviewed_statements table
    create_statements_sql = """
    CREATE TABLE IF NOT EXISTS reviewed_statements (
        id TEXT PRIMARY KEY,
        tenant_id TEXT NOT NULL,
        tenant_name TEXT,
        user_email TEXT,
        period_start TEXT NOT NULL,
        period_end TEXT NOT NULL,
        review_date TEXT NOT NULL,
        total_transactions INTEGER DEFAULT 0,
        flagged_count INTEGER DEFAULT 0,
        high_severity_count INTEGER DEFAULT 0,
        medium_severity_count INTEGER DEFAULT 0,
        low_severity_count INTEGER DEFAULT 0,
        bas_report_data TEXT,
        review_summary TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )
    """

    # Create flagged_items table
    create_items_sql = """
    CREATE TABLE IF NOT EXISTS flagged_items (
        id TEXT PRIMARY KEY,
        statement_id TEXT NOT NULL,
        row_number INTEGER,
        date TEXT,
        account_code TEXT,
        account_name TEXT,
        description TEXT,
        gross REAL,
        gst REAL,
        net REAL,
        source TEXT,
        severity TEXT,
        comments TEXT,
        correcting_journal TEXT,
        xero_url TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (statement_id) REFERENCES reviewed_statements(id)
    )
    """

    # Create all_transactions table (complete audit trail)
    create_all_transactions_sql = """
    CREATE TABLE IF NOT EXISTS all_transactions (
        id TEXT PRIMARY KEY,
        statement_id TEXT NOT NULL,
        row_number INTEGER,
        date TEXT,
        account_code TEXT,
        account_name TEXT,
        description TEXT,
        gross REAL,
        gst REAL,
        net REAL,
        gst_rate_name TEXT,
        source TEXT,
        reference TEXT,
        contact TEXT,
        xero_url TEXT,
        is_flagged INTEGER DEFAULT 0,
        severity TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (statement_id) REFERENCES reviewed_statements(id)
    )
    """

    # Create index for faster lookups
    create_index_sql = """
    CREATE INDEX IF NOT EXISTS idx_statements_tenant ON reviewed_statements(tenant_id)
    """

    create_index_transactions_sql = """
    CREATE INDEX IF NOT EXISTS idx_transactions_statement ON all_transactions(statement_id)
    """

    result1 = cloudflare_d1_query(create_statements_sql)
    result2 = cloudflare_d1_query(create_items_sql)
    result3 = cloudflare_d1_query(create_all_transactions_sql)
    result4 = cloudflare_d1_query(create_index_sql)
    result5 = cloudflare_d1_query(create_index_transactions_sql)

    return {
        'statements_table': result1,
        'flagged_items_table': result2,
        'all_transactions_table': result3,
        'index_statements': result4,
        'index_transactions': result5
    }


def save_review_to_d1(review_data):
    """Save a BAS review to Cloudflare D1

    Args:
        review_data: dict containing review details and flagged items

    Returns:
        dict: Success status and statement ID
    """
    import uuid
    import json

    statement_id = str(uuid.uuid4())

    # Insert the statement record
    insert_statement_sql = """
    INSERT INTO reviewed_statements (
        id, tenant_id, tenant_name, user_email, period_start, period_end,
        review_date, total_transactions, flagged_count, high_severity_count,
        medium_severity_count, low_severity_count, bas_report_data, review_summary
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """

    params = [
        statement_id,
        review_data.get('tenant_id', ''),
        review_data.get('tenant_name', ''),
        review_data.get('user_email', ''),
        review_data.get('period_start', ''),
        review_data.get('period_end', ''),
        review_data.get('review_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        review_data.get('total_transactions', 0),
        review_data.get('flagged_count', 0),
        review_data.get('high_severity_count', 0),
        review_data.get('medium_severity_count', 0),
        review_data.get('low_severity_count', 0),
        json.dumps(review_data.get('bas_report_data', {})),
        json.dumps(review_data.get('review_summary', {}))
    ]

    result = cloudflare_d1_query(insert_statement_sql, params)

    if not result.get('success', True):
        return {'success': False, 'error': result.get('error')}

    # Insert flagged items
    flagged_items = review_data.get('flagged_items', [])
    for item in flagged_items:
        item_id = str(uuid.uuid4())
        insert_item_sql = """
        INSERT INTO flagged_items (
            id, statement_id, row_number, date, account_code, account_name,
            description, gross, gst, net, source, severity, comments,
            correcting_journal, xero_url
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """

        item_params = [
            item_id,
            statement_id,
            item.get('row_number', 0),
            item.get('date', ''),
            item.get('account_code', ''),
            item.get('account', ''),
            item.get('description', ''),
            item.get('gross', 0),
            item.get('gst', 0),
            item.get('net', 0),
            item.get('source', ''),
            item.get('severity', 'low'),
            item.get('comments', ''),
            json.dumps(item.get('correcting_journal', {})),
            item.get('xero_url', '')
        ]

        cloudflare_d1_query(insert_item_sql, item_params)

    # Insert all transactions (complete audit trail)
    all_transactions = review_data.get('all_transactions', [])
    flagged_row_numbers = set(item.get('row_number', 0) for item in flagged_items)

    for txn in all_transactions:
        txn_id = str(uuid.uuid4())
        row_num = txn.get('row_number', 0)
        is_flagged = 1 if row_num in flagged_row_numbers else 0

        # Find severity if flagged
        severity = ''
        if is_flagged:
            for fi in flagged_items:
                if fi.get('row_number') == row_num:
                    severity = fi.get('severity', '')
                    break

        insert_txn_sql = """
        INSERT INTO all_transactions (
            id, statement_id, row_number, date, account_code, account_name,
            description, gross, gst, net, gst_rate_name, source, reference,
            contact, xero_url, is_flagged, severity
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """

        txn_params = [
            txn_id,
            statement_id,
            row_num,
            txn.get('date', ''),
            txn.get('account_code', ''),
            txn.get('account', ''),
            txn.get('description', ''),
            txn.get('gross', 0),
            txn.get('gst', 0),
            txn.get('net', 0),
            txn.get('gst_rate_name', ''),
            txn.get('source', ''),
            txn.get('reference', ''),
            txn.get('contact', ''),
            txn.get('xero_url', ''),
            is_flagged,
            severity
        ]

        cloudflare_d1_query(insert_txn_sql, txn_params)

    return {'success': True, 'statement_id': statement_id}


def get_reviews_from_d1(tenant_id=None, limit=50):
    """Get list of reviewed statements from Cloudflare D1

    Args:
        tenant_id: Optional filter by tenant
        limit: Max number of records to return

    Returns:
        list: List of reviewed statements
    """
    if tenant_id:
        sql = """
        SELECT * FROM reviewed_statements
        WHERE tenant_id = ?
        ORDER BY review_date DESC
        LIMIT ?
        """
        params = [tenant_id, limit]
    else:
        sql = """
        SELECT * FROM reviewed_statements
        ORDER BY review_date DESC
        LIMIT ?
        """
        params = [limit]

    result = cloudflare_d1_query(sql, params)

    if result.get('success') and result.get('result'):
        return result['result'][0].get('results', [])
    return []


def get_review_details_from_d1(statement_id):
    """Get detailed review including flagged items from Cloudflare D1

    Args:
        statement_id: The statement ID to retrieve

    Returns:
        dict: Statement details with flagged items
    """
    # Get statement
    statement_sql = "SELECT * FROM reviewed_statements WHERE id = ?"
    statement_result = cloudflare_d1_query(statement_sql, [statement_id])

    if not statement_result.get('success') or not statement_result.get('result'):
        return None

    statements = statement_result['result'][0].get('results', [])
    if not statements:
        return None

    statement = statements[0]

    # Get flagged items
    items_sql = "SELECT * FROM flagged_items WHERE statement_id = ? ORDER BY row_number"
    items_result = cloudflare_d1_query(items_sql, [statement_id])

    flagged_items = []
    if items_result.get('success') and items_result.get('result'):
        flagged_items = items_result['result'][0].get('results', [])

    statement['flagged_items'] = flagged_items
    return statement


def fetch_xero_journals_debug(from_date_str, to_date_str):
    """Fetch all journal entries from Xero API with debug info"""
    transactions = []
    debug_info = []

    from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
    to_date = datetime.strptime(to_date_str, '%Y-%m-%d')

    # Xero Journals API uses offset-based pagination
    offset = 0
    last_journal_number = 0
    total_journals_fetched = 0

    while True:
        params = {'offset': offset}
        debug_info.append(f"Calling Journals API with offset={offset}")
        data = xero_api_request('Journals', params=params)
        debug_info.append(f"API returned data={data is not None}")

        if not data or 'Journals' not in data:
            debug_info.append(f"No data or no Journals key. Data keys: {list(data.keys()) if data else 'None'}")
            if data and 'Message' in data:
                debug_info.append(f"API Error Message: {data.get('Message')}")
            break

        journals = data.get('Journals', [])
        debug_info.append(f"Got {len(journals)} journals in this batch")
        if not journals:
            break

        total_journals_fetched += len(journals)
        journals_in_range = 0

        for journal in journals:
            raw_date = journal.get('JournalDate', '')
            journal_date = parse_xero_date(raw_date)

            # Track the highest journal number for proper pagination
            current_journal_number = journal.get('JournalNumber', 0)
            if isinstance(current_journal_number, (int, float)):
                last_journal_number = max(last_journal_number, int(current_journal_number))

            # Skip draft and voided journals - only include posted journals
            journal_status = journal.get('Status', 'POSTED')
            if journal_status in ['DRAFT', 'VOIDED', 'DELETED']:
                continue

            # Filter by date range
            if journal_date:
                if journal_date < from_date or journal_date > to_date:
                    continue
                journals_in_range += 1

            journal_number = journal.get('JournalNumber', '')
            source_type = journal.get('SourceType', '')
            reference = journal.get('Reference', '')
            narration = journal.get('Narration', '')  # Journal-level description (often contains vendor name)

            # Process each journal line (GL entry)
            for line in journal.get('JournalLines', []):
                account_code = line.get('AccountCode', '')
                account_name = line.get('AccountName', '')
                description = line.get('Description', '') or reference or 'No description'

                # Skip lines without account codes
                if not account_code:
                    continue

                # Get amounts
                gross = float(line.get('GrossAmount', 0) or 0)
                net = float(line.get('NetAmount', 0) or 0)
                gst = float(line.get('TaxAmount', 0) or 0)

                # Skip zero-value lines
                if gross == 0 and net == 0 and gst == 0:
                    continue

                # Determine transaction type
                account_type = line.get('AccountType', '')

                # Skip balance sheet accounts - only review P&L accounts for BAS
                # Assets (Bank, Receivables, Fixed Assets) and Liabilities (Payables, GST)
                # are not relevant for BAS GST compliance review
                balance_sheet_types = [
                    'BANK', 'CURRENT', 'FIXED', 'INVENTORY', 'NONCURRENT', 'PREPAYMENT',  # Asset types
                    'CURRLIAB', 'LIABILITY', 'TERMLIAB', 'PAYGLIABILITY', 'SUPERANNUATIONLIABILITY',  # Liability types
                ]
                if account_type in balance_sheet_types:
                    continue

                # Also skip GST control accounts and balance sheet accounts by name
                account_name_lower = account_name.lower()
                skip_keywords = ['gst', 'accounts payable', 'accounts receivable', 'bank', 'petty cash',
                                'clearing', 'suspense', 'control', 'payg', 'superannuation liability',
                                'rounding', 'historical adjustment', 'retained earnings', 'current year earnings']
                if any(x in account_name_lower for x in skip_keywords):
                    continue

                is_expense = gross < 0 or account_type in ['EXPENSE', 'OVERHEADS', 'DIRECTCOSTS']

                date_str = journal_date.strftime('%Y-%m-%d') if journal_date else 'NO DATE'

                # Map tax type to GST rate name
                tax_type = line.get('TaxType', '')
                gst_rate_name = ''
                if 'OUTPUT' in tax_type.upper():
                    gst_rate_name = 'GST on Income'
                elif 'INPUT' in tax_type.upper():
                    gst_rate_name = 'GST on Expenses'
                elif 'NONE' in tax_type.upper() or 'EXEMPT' in tax_type.upper():
                    gst_rate_name = 'GST Free'
                elif 'BASEXCLUDED' in tax_type.upper():
                    gst_rate_name = 'BAS Excluded'
                else:
                    gst_rate_name = tax_type

                transactions.append({
                    'row_number': len(transactions) + 1,
                    'date': date_str,
                    'type': 'expense' if is_expense else 'income',
                    'account_code': account_code,
                    'account': account_name,
                    'description': description,
                    'gross': abs(gross),
                    'gst': abs(gst),
                    'net': abs(net),
                    'gst_rate_name': gst_rate_name,
                    'source': source_type,
                    'reference': reference,
                    'narration': narration,  # Journal-level narration (often has vendor name)
                    'journal_number': journal_number,
                    'account_type': account_type
                })

        # Pagination: if we got fewer than 100 journals, we've reached the end
        if len(journals) < 100:
            debug_info.append(f"Less than 100 journals, stopping pagination")
            break

        # Use the last journal number + 1 as the next offset
        if last_journal_number > 0:
            offset = last_journal_number + 1
        else:
            offset += len(journals)

        debug_info.append(f"{journals_in_range} journals in date range from this batch")

    debug_info.append(f"Total journals fetched: {total_journals_fetched}, Total transactions before dedup: {len(transactions)}")

    # Deduplicate: Keep only the LATEST journal entry for each unique transaction
    # When bills are edited, new journals are created with higher journal numbers
    # We want the latest version (highest journal number) to reflect current state
    #
    # Two-phase deduplication:
    # 1. First by (date, amount, description) - catches re-coded transactions (account changed)
    # 2. Then by (date, account, amount) - catches duplicate entries in same account

    # Phase 1: Group by (date, gross, description) - keep highest journal number
    # This catches when a transaction was re-coded to a different account
    phase1_map = {}
    for t in transactions:
        desc_key = (t.get('description', '') or '')[:30].lower().strip()
        key = (
            t.get('date', ''),
            round(t.get('gross', 0), 2),
            desc_key
        )
        journal_num = t.get('journal_number', 0)
        try:
            journal_num = int(journal_num) if journal_num else 0
        except:
            journal_num = 0

        if key not in phase1_map or journal_num > phase1_map[key]['journal_num']:
            phase1_map[key] = {'transaction': t, 'journal_num': journal_num}

    phase1_transactions = [v['transaction'] for v in phase1_map.values()]

    # Phase 2: Group by (date, account, gross) - remove any remaining duplicates
    phase2_map = {}
    for t in phase1_transactions:
        key = (
            t.get('date', ''),
            t.get('account_code', ''),
            round(t.get('gross', 0), 2)
        )
        journal_num = t.get('journal_number', 0)
        try:
            journal_num = int(journal_num) if journal_num else 0
        except:
            journal_num = 0

        if key not in phase2_map or journal_num > phase2_map[key]['journal_num']:
            phase2_map[key] = {'transaction': t, 'journal_num': journal_num}

    unique_transactions = [v['transaction'] for v in phase2_map.values()]

    debug_info.append(f"After deduplication (keeping latest journals): {len(unique_transactions)} transactions")
    return unique_transactions, debug_info


def enrich_transactions_with_accounts(transactions):
    """Add account names to transactions using Chart of Accounts"""
    # Fetch chart of accounts
    data = xero_api_request('Accounts')
    if not data or 'Accounts' not in data:
        return transactions

    # Build account code to name mapping
    account_map = {}
    for acc in data.get('Accounts', []):
        code = acc.get('Code', '')
        name = acc.get('Name', '')
        acc_type = acc.get('Type', '')
        if code:
            account_map[code] = {'name': name, 'type': acc_type}

    # Enrich transactions
    for txn in transactions:
        code = txn.get('account_code', '')
        if code and code in account_map:
            txn['account'] = f"{account_map[code]['name']} ({code})"
            txn['account_type'] = account_map[code]['type']
        elif code:
            # Code exists but not in chart of accounts
            txn['account'] = f"Unknown Account ({code})"
            txn['account_type'] = ''
        else:
            # No account code - flag this
            txn['account'] = 'No Account Assigned'
            txn['account_code'] = ''
            txn['account_type'] = ''

    return transactions


@app.route('/api/upload-review', methods=['POST'])
@login_required
def upload_review():
    """Process uploaded General Ledger Detail or Activity Statement Excel file"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Please upload an Excel file (.xlsx or .xls)'}), 400

        # Save file content to buffer (so we can both upload to R2 and process)
        file_content = file.read()
        file_buffer = BytesIO(file_content)

        # Upload to R2 for training data collection (async, non-blocking)
        try:
            r2_buffer = BytesIO(file_content)
            upload_to_r2(r2_buffer, file.filename)
        except Exception as r2_err:
            print(f"R2 upload error (non-critical): {r2_err}")

        # Read the Excel file from buffer
        try:
            raw_data = pd.read_excel(file_buffer, header=None)
        except Exception as read_err:
            return jsonify({'error': f'Failed to read Excel file: {str(read_err)}'}), 400

        if len(raw_data) < 5:
            return jsonify({'error': f'File appears too small. Found only {len(raw_data)} rows.'}), 400

        # Detect file format: Activity Statement vs General Ledger Detail
        first_cell = str(raw_data.iloc[0, 0]).strip().lower() if not pd.isna(raw_data.iloc[0, 0]) else ''

        if 'transactions by tax rate' in first_cell or 'activity statement' in first_cell:
            # Parse as Activity Statement format
            transactions, company_name, period = parse_activity_statement(raw_data)
            if not transactions:
                return jsonify({'error': 'No transactions found in Activity Statement file.'}), 404

            # Run BAS review on Activity Statement transactions
            results = run_bas_review(transactions)

            ai_reviewed = 0
            for item in results:
                if item.get('ai_review'):
                    ai_reviewed += 1

            return jsonify({
                'company': company_name,
                'period': period,
                'total': len(transactions),
                'flagged': len(results),
                'ai_reviewed': ai_reviewed,
                'results': results,
                'format': 'activity_statement'
            })

        # Otherwise, parse as General Ledger Detail format
        # Extract metadata
        company_name = str(raw_data.iloc[0, 0]) if len(raw_data) > 0 and not pd.isna(raw_data.iloc[0, 0]) else 'Unknown'
        period = str(raw_data.iloc[1, 0]) if len(raw_data) > 1 and not pd.isna(raw_data.iloc[1, 0]) else ''

        # Find header row - look for various possible header names
        header_row_idx = None
        for idx, row in raw_data.iterrows():
            first_cell = str(row[0]).strip().lower() if not pd.isna(row[0]) else ''
            if first_cell in ['account code', 'account', 'acc code', 'code', 'date']:
                header_row_idx = idx
                break

        if header_row_idx is None:
            # Try to find any row that looks like a header
            for idx, row in raw_data.iterrows():
                if idx > 10:  # Don't look too far
                    break
                first_cell = str(row[0]).strip().lower() if not pd.isna(row[0]) else ''
                if 'account' in first_cell or 'code' in first_cell or 'date' in first_cell:
                    header_row_idx = idx
                    break

        if header_row_idx is None:
            # Show first few rows to help debug
            first_rows = raw_data.head(10).to_string()
            return jsonify({
                'error': 'Could not find header row. Looking for "Account Code" or "Date" column.',
                'hint': 'Make sure this is a General Ledger Detail report or Activity Statement from Xero.',
                'first_rows': first_rows[:500]
            }), 400

        # Dynamic column detection - map header names to column indices
        header_row = raw_data.iloc[header_row_idx]
        col_map = {}

        # Define possible header names for each field (case-insensitive)
        field_mappings = {
            'account_code': ['account code', 'acc code', 'code', 'account'],
            'account_name': ['account name', 'name', 'account', 'description'],
            'date': ['date', 'trans date', 'transaction date', 'posted'],
            'source': ['source', 'type', 'source type'],
            'description': ['description', 'reference', 'memo', 'particulars', 'details'],
            'invoice_number': ['invoice', 'invoice number', 'inv no', 'invoice no', 'inv #'],
            'reference': ['reference', 'ref', 'ref no', 'reference number'],
            'gross': ['gross', 'gross amount', 'total', 'amount', 'debit', 'credit'],
            'gst': ['gst', 'gst amount', 'tax', 'tax amount'],
            'net': ['net', 'net amount', 'ex gst', 'excl gst', 'exclusive'],
            'gst_rate': ['gst rate', 'tax rate', 'rate'],
            'gst_rate_name': ['gst rate name', 'tax type', 'tax code', 'gst code', 'tax name']
        }

        # Map columns based on header names
        for col_idx, cell in enumerate(header_row):
            if pd.isna(cell):
                continue
            cell_lower = str(cell).strip().lower()
            for field, possible_names in field_mappings.items():
                if cell_lower in possible_names or any(name in cell_lower for name in possible_names):
                    if field not in col_map:  # Keep first match
                        col_map[field] = col_idx
                        break

        # Ensure we have at least account_code and some amount column
        if 'account_code' not in col_map:
            col_map['account_code'] = 0  # Default to first column
        if 'gross' not in col_map and 'net' not in col_map:
            # Try to find numeric columns as amount columns
            for col_idx in range(len(header_row)):
                if col_idx not in col_map.values():
                    col_map['gross'] = col_idx
                    break

        # Parse amounts helper
        def parse_amount(val):
            if pd.isna(val):
                return 0.0
            if isinstance(val, (int, float)):
                return float(val)
            val_str = str(val).replace('$', '').replace(',', '').replace('(', '-').replace(')', '').strip()
            try:
                return float(val_str)
            except:
                return 0.0

        # Helper to safely get column value
        def get_col_value(row, field, default=''):
            if field not in col_map:
                return default
            col_idx = col_map[field]
            if col_idx >= len(row):
                return default
            val = row.iloc[col_idx] if hasattr(row, 'iloc') else row[col_idx]
            if pd.isna(val):
                return default
            return val

        # Parse transactions
        transactions = []
        parse_errors = []

        for idx in range(header_row_idx + 1, len(raw_data)):
            try:
                row = raw_data.iloc[idx]

                # Get account code
                account_code_val = get_col_value(row, 'account_code', '')
                if not account_code_val or str(account_code_val).strip() == '':
                    continue

                account_code = str(account_code_val).strip()

                # Skip non-numeric account codes (totals, headers) but allow alphanumeric codes
                # Xero can have codes like "200", "200.0", or even "BANK-001"
                clean_code = account_code.replace('.', '').replace('-', '').replace(' ', '')
                if not any(c.isdigit() for c in clean_code):
                    continue  # Skip if no digits at all (likely a total or header row)

                # Normalize numeric codes
                try:
                    if account_code.replace('.', '').isdigit():
                        account_code = str(int(float(account_code)))
                except:
                    pass  # Keep original if can't normalize

                # Parse date - try to get it but don't skip if missing
                date_value = get_col_value(row, 'date', None)
                date_str = 'NO DATE'
                if date_value:
                    if isinstance(date_value, (datetime, pd.Timestamp)):
                        date_str = date_value.strftime('%Y-%m-%d')
                    else:
                        try:
                            date_str = pd.to_datetime(date_value).strftime('%Y-%m-%d')
                        except:
                            date_str = str(date_value)[:10] if date_value else 'NO DATE'

                # Get amounts
                gross = parse_amount(get_col_value(row, 'gross', 0))
                gst = parse_amount(get_col_value(row, 'gst', 0))
                net = parse_amount(get_col_value(row, 'net', 0))

                # If net is 0 but gross exists, calculate net
                if net == 0 and gross != 0:
                    net = gross - gst

                # Skip truly empty transactions (all amounts zero)
                if gross == 0 and gst == 0 and net == 0:
                    continue

                # Get account name - try account_name column first, then extract from account_code if needed
                account_name = str(get_col_value(row, 'account_name', '')).strip()

                # If account_name is empty, try to extract from account_code value
                # Some Excel exports have "477 Wages & Salaries" in account column
                if not account_name:
                    raw_account = str(get_col_value(row, 'account_code', ''))
                    # Extract name part after first space or dash (e.g., "477 Wages & Salaries" -> "Wages & Salaries")
                    name_match = re.search(r'^\d+[\s\-]+(.+)$', raw_account.strip())
                    if name_match:
                        account_name = name_match.group(1).strip()

                transaction = {
                    'row_number': idx + 1,
                    'account_code': account_code,
                    'account': account_name,
                    'date': date_str,
                    'source': str(get_col_value(row, 'source', '')),
                    'description': str(get_col_value(row, 'description', '')),
                    'invoice_number': str(get_col_value(row, 'invoice_number', '')),
                    'reference': str(get_col_value(row, 'reference', '')),
                    'gross': gross,
                    'gst': gst,
                    'net': net,
                    'gst_rate': parse_amount(get_col_value(row, 'gst_rate', 0)),
                    'gst_rate_name': str(get_col_value(row, 'gst_rate_name', '')),
                }

                # Determine transaction type
                account_lower = transaction['account'].lower()
                if any(word in account_lower for word in ['sales', 'income', 'revenue']):
                    transaction['type'] = 'income'
                elif any(word in account_lower for word in ['expense', 'cost', 'fees']):
                    transaction['type'] = 'expense'
                else:
                    try:
                        code_num = int(account_code)
                        transaction['type'] = 'income' if 200 <= code_num < 400 else 'expense'
                    except:
                        transaction['type'] = 'unknown'

                transactions.append(transaction)

            except Exception as row_err:
                parse_errors.append(f"Row {idx + 1}: {str(row_err)}")
                continue

        if not transactions:
            error_msg = 'No transactions found in the file.'
            if parse_errors:
                error_msg += f' Parse errors: {"; ".join(parse_errors[:5])}'
            return jsonify({
                'error': error_msg,
                'rows_scanned': len(raw_data) - header_row_idx - 1,
                'header_row': header_row_idx
            }), 404

        # Run BAS review - rule-based first, then AI only for flagged items
        flagged_items = []
        rule_flagged = []

        # Infer business context from all transactions to understand industry/income sources
        try:
            business_context = set_business_context(transactions)
            print(f"Inferred business context: {business_context['industry']} (confidence: {business_context['confidence']:.1%})")
        except Exception as e:
            print(f"Error setting business context: {e}")

        # First pass: fast rule-based checks per ATO GST rules
        for transaction in transactions:
            try:
                transaction['account_coding_suspicious'] = check_account_coding(transaction)
            except Exception as e:
                print(f"Error in check_account_coding: {e}")
                transaction['account_coding_suspicious'] = False

            # Check for split allocation patterns (Deep Scan only)
            try:
                split_check = check_split_allocation_pattern(transaction)
                transaction['split_allocation_warning'] = split_check
            except Exception as e:
                print(f"Error in check_split_allocation_pattern: {e}")
                transaction['split_allocation_warning'] = None

            try:
                transaction['alcohol_gst_error'] = check_alcohol_gst(transaction)
            except Exception as e:
                print(f"Error in check_alcohol_gst: {e}")
                transaction['alcohol_gst_error'] = False

            try:
                transaction['input_taxed_gst_error'] = check_input_taxed_gst(transaction)
            except Exception as e:
                print(f"Error in check_input_taxed_gst: {e}")
                transaction['input_taxed_gst_error'] = False

            try:
                transaction['missing_gst_error'] = check_missing_gst(transaction)
            except Exception as e:
                print(f"Error in check_missing_gst: {e}")
                transaction['missing_gst_error'] = False

            try:
                transaction['gst_calculation_correct'] = check_gst_calculation(transaction)
            except Exception as e:
                print(f"Error in check_gst_calculation: {e}")
                transaction['gst_calculation_correct'] = True

            try:
                transaction['drawings_loan_error'] = check_drawings_loan_error(transaction)
            except Exception as e:
                print(f"Error in check_drawings_loan_error: {e}")
                transaction['drawings_loan_error'] = False

            try:
                transaction['personal_in_business_account'] = check_personal_expense_in_business_account(transaction)
            except Exception as e:
                print(f"Error in check_personal_expense_in_business_account: {e}")
                transaction['personal_in_business_account'] = False

            try:
                transaction['asset_capitalization_error'] = check_asset_capitalization(transaction)
            except Exception as e:
                print(f"Error in check_asset_capitalization: {e}")
                transaction['asset_capitalization_error'] = False

            try:
                transaction['computer_equipment_expense'] = check_computer_equipment_expense(transaction)
            except Exception as e:
                print(f"Error in check_computer_equipment_expense: {e}")
                transaction['computer_equipment_expense'] = False

            try:
                transaction['interest_gst_error'] = check_interest_gst_error(transaction)
            except Exception as e:
                print(f"Error in check_interest_gst_error: {e}")
                transaction['interest_gst_error'] = False

            try:
                transaction['other_income_error'] = check_other_income_error(transaction)
            except Exception as e:
                print(f"Error in check_other_income_error: {e}")
                transaction['other_income_error'] = False

            try:
                transaction['sales_gst_error'] = check_sales_gst_error(transaction)
            except Exception as e:
                print(f"Error in check_sales_gst_error: {e}")
                transaction['sales_gst_error'] = False

            try:
                transaction['motor_vehicle_gst_limit'] = check_motor_vehicle_gst_limit(transaction)
            except Exception as e:
                print(f"Error in check_motor_vehicle_gst_limit: {e}")
                transaction['motor_vehicle_gst_limit'] = False

            try:
                transaction['overseas_subscription_gst'] = check_overseas_subscription_gst(transaction)
            except Exception as e:
                print(f"Error in check_overseas_subscription_gst: {e}")
                transaction['overseas_subscription_gst'] = False

            try:
                transaction['government_charges_gst'] = check_government_charges_gst(transaction)
            except Exception as e:
                print(f"Error in check_government_charges_gst: {e}")
                transaction['government_charges_gst'] = False

            try:
                transaction['client_entertainment_gst'] = check_client_entertainment_gst(transaction)
            except Exception as e:
                print(f"Error in check_client_entertainment_gst: {e}")
                transaction['client_entertainment_gst'] = False

            try:
                transaction['staff_entertainment_gst'] = check_staff_entertainment_gst(transaction)
            except Exception as e:
                print(f"Error in check_staff_entertainment_gst: {e}")
                transaction['staff_entertainment_gst'] = False

            try:
                transaction['residential_premises_gst'] = check_residential_premises_gst(transaction)
            except Exception as e:
                print(f"Error in check_residential_premises_gst: {e}")
                transaction['residential_premises_gst'] = False

            try:
                transaction['insurance_gst_error'] = check_insurance_gst_error(transaction)
            except Exception as e:
                print(f"Error in check_insurance_gst_error: {e}")
                transaction['insurance_gst_error'] = False

            try:
                transaction['life_insurance_personal'] = check_life_insurance_personal(transaction)
            except Exception as e:
                print(f"Error in check_life_insurance_personal: {e}")
                transaction['life_insurance_personal'] = False

            try:
                transaction['grants_sponsorship_gst'] = check_grants_sponsorship_gst(transaction)
            except Exception as e:
                print(f"Error in check_grants_sponsorship_gst: {e}")
                transaction['grants_sponsorship_gst'] = False

            try:
                transaction['wages_gst_error'] = check_wages_gst_error(transaction)
            except Exception as e:
                print(f"Error in check_wages_gst_error: {e}")
                transaction['wages_gst_error'] = False

            try:
                transaction['allowance_gst_error'] = check_allowance_gst_error(transaction)
            except Exception as e:
                print(f"Error in check_allowance_gst_error: {e}")
                transaction['allowance_gst_error'] = False

            try:
                transaction['reimbursement_gst'] = check_reimbursement_gst(transaction)
            except Exception as e:
                print(f"Error in check_reimbursement_gst: {e}")
                transaction['reimbursement_gst'] = False

            try:
                transaction['voucher_gst'] = check_voucher_gst(transaction)
            except Exception as e:
                print(f"Error in check_voucher_gst: {e}")
                transaction['voucher_gst'] = False

            try:
                transaction['general_expenses'] = check_general_expenses(transaction)
            except Exception as e:
                print(f"Error in check_general_expenses: {e}")
                transaction['general_expenses'] = False

            try:
                transaction['travel_gst'] = check_travel_gst(transaction)
            except Exception as e:
                print(f"Error in check_travel_gst: {e}")
                transaction['travel_gst'] = False

            try:
                transaction['payment_processor_fees'] = check_payment_processor_fees(transaction)
            except Exception as e:
                print(f"Error in check_payment_processor_fees: {e}")
                transaction['payment_processor_fees'] = False

            try:
                transaction['fines_penalties_gst'] = check_fines_penalties_gst(transaction)
            except Exception as e:
                print(f"Error in check_fines_penalties_gst: {e}")
                transaction['fines_penalties_gst'] = False

            try:
                transaction['donations_gst'] = check_donations_gst(transaction)
            except Exception as e:
                print(f"Error in check_donations_gst: {e}")
                transaction['donations_gst'] = False

            try:
                transaction['property_gst_withholding'] = check_property_gst_withholding(transaction)
            except Exception as e:
                print(f"Error in check_property_gst_withholding: {e}")
                transaction['property_gst_withholding'] = False

            try:
                transaction['livestock_gst'] = check_livestock_gst(transaction)
            except Exception as e:
                print(f"Error in check_livestock_gst: {e}")
                transaction['livestock_gst'] = False

            try:
                transaction['asset_disposal_gst'] = check_asset_disposal_gst(transaction)
            except Exception as e:
                print(f"Error in check_asset_disposal_gst: {e}")
                transaction['asset_disposal_gst'] = False

            try:
                transaction['export_gst_error'] = check_export_gst_error(transaction)
            except Exception as e:
                print(f"Error in check_export_gst_error: {e}")
                transaction['export_gst_error'] = False

            try:
                transaction['borrowing_expenses_error'] = check_borrowing_expenses_error(transaction)
            except Exception as e:
                print(f"Error in check_borrowing_expenses_error: {e}")
                transaction['borrowing_expenses_error'] = False

            # Flag if any rule triggered (including new checks)
            # Note: split_allocation_warning removed from flagging - personal_in_business_account handles actual errors
            has_rule_issues = (
                transaction['account_coding_suspicious'] or
                transaction['alcohol_gst_error'] or
                transaction['input_taxed_gst_error'] or
                transaction['missing_gst_error'] or
                not transaction['gst_calculation_correct'] or
                transaction['drawings_loan_error'] or
                transaction.get('personal_in_business_account') or
                transaction['asset_capitalization_error'] or
                transaction['computer_equipment_expense'] or
                transaction['interest_gst_error'] or
                transaction['other_income_error'] or
                transaction['sales_gst_error'] or
                transaction['motor_vehicle_gst_limit'] or
                transaction['overseas_subscription_gst'] or
                transaction['government_charges_gst'] or
                transaction['client_entertainment_gst'] or
                transaction['staff_entertainment_gst'] or
                transaction['residential_premises_gst'] or
                transaction['insurance_gst_error'] or
                transaction['life_insurance_personal'] or
                transaction['grants_sponsorship_gst'] or
                transaction['wages_gst_error'] or
                transaction['allowance_gst_error'] or
                transaction['reimbursement_gst'] or
                transaction['voucher_gst'] or
                transaction['general_expenses'] or
                transaction['travel_gst'] or
                transaction['payment_processor_fees'] or
                transaction['fines_penalties_gst'] or
                transaction['donations_gst'] or
                transaction['property_gst_withholding'] or
                transaction['livestock_gst'] or
                transaction['asset_disposal_gst'] or
                transaction['export_gst_error'] or
                transaction['borrowing_expenses_error']
            )

            # Skip flagging for correctly coded Telstra transactions
            # (Telstra business in Telephone, Telstra personal in Drawings)
            if has_rule_issues and is_correctly_coded_telstra(transaction):
                has_rule_issues = False

            if has_rule_issues:
                rule_flagged.append(transaction)

        # Second pass: AI review for flagged items (batch processing)
        # Limit AI review to 200 items to prevent timeout on large datasets
        MAX_AI_REVIEW = 200
        ai_review_limit = min(len(rule_flagged), MAX_AI_REVIEW)

        if len(rule_flagged) > MAX_AI_REVIEW:
            print(f"Large dataset: {len(rule_flagged)} flagged items, limiting AI review to {MAX_AI_REVIEW}")

        # Use batch AI review for faster processing (5 transactions per API call)
        try:
            ai_results = review_batch_with_ai(rule_flagged[:ai_review_limit], batch_size=5)
        except Exception as e:
            print(f"Error in batch AI review: {e}")
            ai_results = [{'has_issues': True, 'severity': 'high', 'comments': '', 'issues': []} for _ in range(ai_review_limit)]

        for i, transaction in enumerate(rule_flagged[:ai_review_limit]):
            ai_result = ai_results[i] if i < len(ai_results) else {'has_issues': True, 'severity': 'high', 'comments': '', 'issues': []}

            # Use simplified AI-generated comment, with rule-based fallback if AI doesn't provide useful comment
            comments = []
            ai_comment = ai_result.get('comments', '').strip()

            # Check if AI comment is useful (not empty or too generic)
            # Include "correctly applied/coded/recorded" as generic since AI may miss the actual issue
            generic_phrases = ['requires review', 'please review', 'review required', 'ok -', 'appears correct',
                              'correctly applied', 'correctly coded', 'correctly recorded', 'no issues',
                              'looks correct', 'is correct', 'recorded correctly', 'coded correctly']
            is_useful_ai_comment = ai_comment and len(ai_comment) > 20 and not any(phrase in ai_comment.lower() for phrase in generic_phrases)

            # Override AI comment if it uses incorrect terminology
            if is_useful_ai_comment and ai_comment:
                ai_lower = ai_comment.lower()
                # Wages are BAS Excluded, NOT input-taxed - force rule-based if AI uses wrong term
                if transaction.get('wages_gst_error') and ('input-taxed' in ai_lower or 'input taxed' in ai_lower):
                    is_useful_ai_comment = False
                # Government charges have NO GST (not input-taxed, not taxable) - force rule-based if wrong
                if transaction.get('government_charges_gst') and ('input-taxed' in ai_lower or 'input taxed' in ai_lower or 'taxable supply' in ai_lower or 'taxable supplies' in ai_lower):
                    is_useful_ai_comment = False
                # Council/water rates should not mention "residential" or "taxable"
                if ('council' in ai_lower or 'water' in ai_lower) and 'rates' in ai_lower and ('residential' in ai_lower or 'taxable supply' in ai_lower or 'taxable supplies' in ai_lower):
                    is_useful_ai_comment = False
                # Body corporate assumed residential without evidence
                if 'body corporate' in ai_lower and 'residential' in ai_lower and 'residential' not in transaction.get('description', '').lower():
                    is_useful_ai_comment = False
                # International travel flagged but AI says taxable - international flights are GST-FREE
                if transaction.get('travel_gst') == 'international_with_gst' and 'taxable' in ai_lower:
                    is_useful_ai_comment = False
                # Life/income protection insurance - always use rule-based for consistent Owner Drawings advice
                if transaction.get('life_insurance_personal') or transaction.get('insurance_gst_error'):
                    # Force rule-based comment which includes Owner Drawings recode advice
                    is_useful_ai_comment = False
                # Payment processor fees - always use rule-based for specific PayPal/Stripe guidance
                if transaction.get('payment_processor_fees'):
                    is_useful_ai_comment = False
                # Government charges - always use rule-based (AI incorrectly says "input-taxed" or "residential")
                if transaction.get('government_charges_gst'):
                    is_useful_ai_comment = False
                # Export GST errors - always use rule-based for consistent GSTR 2002/6 reference
                if transaction.get('export_gst_error'):
                    is_useful_ai_comment = False

            # Always prioritize AI comments, use rule-based as fallback only
            if is_useful_ai_comment:
                comments.append(ai_comment)
            else:
                # Fallback to rule-based comments when AI doesn't provide useful info
                # Check if it's a personal expense - if so, skip business-specific rules like capitalization
                is_personal = transaction.get('life_insurance_personal') or transaction.get('personal_in_business_account')

                if transaction.get('life_insurance_personal'):
                    ato_comment = generate_ato_comment('life_insurance_personal')
                    comments.append(ato_comment or 'Life/income protection insurance - NOT a deductible business expense (ATO). Recode to Owner Drawings. Owner may claim income protection on personal tax return.')
                if transaction.get('personal_in_business_account'):
                    comments.append('Personal expense in business account - NOT deductible. Recode to Owner Drawings (personal expenses cannot be claimed as business deductions).')
                # Skip asset/equipment capitalization rules for personal expenses (not relevant)
                if transaction.get('asset_capitalization_error') and not is_personal:
                    comments.append('Asset over $20,000 - should be capitalized per ATO instant asset write-off rules, not expensed')
                if transaction.get('computer_equipment_expense') and not is_personal:
                    comments.append('Computer equipment over $300 - should be capitalized as asset, not expensed to Office Supplies')
                if transaction.get('insurance_gst_error'):
                    ato_comment = generate_ato_comment('life_insurance_personal')
                    comments.append(ato_comment or 'Life/income protection insurance - NOT a deductible business expense. Recode to Owner Drawings. No GST credit claimable (input-taxed). Owner may claim on personal tax return.')
                if transaction.get('wages_gst_error'):
                    ato_comment = generate_ato_comment('wages_gst_error')
                    comments.append(ato_comment or 'Wages/salaries/super - should be BAS Excluded (no GST)')
                if transaction.get('alcohol_gst_error') or transaction.get('client_entertainment_gst') or transaction.get('staff_entertainment_gst'):
                    ato_comment = generate_ato_comment('entertainment')
                    comments.append(ato_comment or 'Entertainment expense - NO GST credit claimable. Entertainment is non-deductible and GST credits blocked unless FBT is paid.')
                if transaction.get('missing_gst_error'):
                    comments.append('Should include GST (10%) - currently coded as GST Free')
                if transaction.get('input_taxed_gst_error'):
                    ato_comment = generate_ato_comment('input_taxed_gst_error')
                    comments.append(ato_comment or 'Input-taxed supply - GST incorrectly claimed (no GST credit on financial supplies)')
                if transaction.get('general_expenses'):
                    comments.append('General/Sundry Expenses - recode to specific category to reduce audit risk')
                if transaction.get('drawings_loan_error'):
                    comments.append('Drawings/Loan account - should be BAS Excluded')
                if transaction.get('fines_penalties_gst'):
                    ato_comment = generate_ato_comment('fines_penalties_gst')
                    comments.append(ato_comment or 'Fine/penalty - BAS Excluded (non-reportable, no GST)')
                if transaction.get('government_charges_gst'):
                    ato_comment = generate_ato_comment('government_charges_gst')
                    comments.append(ato_comment or 'Government charge - NO GST applies (not a taxable supply)')
                if transaction.get('donations_gst'):
                    ato_comment = generate_ato_comment('donations_gst')
                    comments.append(ato_comment or 'Donation - NO GST applies. Use GST Free Expenses for P&L accounts.')
                if transaction.get('travel_gst') == 'international_with_gst':
                    ato_comment = generate_ato_comment('travel_gst_international')
                    comments.append(ato_comment or 'International travel - GST FREE. Cannot claim GST credits on international travel expenses.')
                elif transaction.get('travel_gst') == 'domestic_no_gst':
                    ato_comment = generate_ato_comment('travel_gst_domestic')
                    comments.append(ato_comment or 'Domestic travel (within Australia) - TAXABLE. Should include GST (10%). Domestic flights, hotels, taxis are GST taxable.')
                if transaction.get('grants_sponsorship_gst') == 'grant_with_gst':
                    ato_comment = generate_ato_comment('grants_sponsorship_gst')
                    comments.append(ato_comment or 'Grant income with GST charged - grants are typically GST-FREE unless binding supply obligation exists.')
                elif transaction.get('grants_sponsorship_gst') == 'sponsorship_no_gst':
                    comments.append('Sponsorship income without GST - sponsorship is TAXABLE (GST applies) as it involves promotional services in return.')
                if transaction.get('export_gst_error'):
                    ato_comment = generate_ato_comment('export_gst_error')
                    comments.append(ato_comment or 'Export sale with GST charged - exports should be GST-FREE. No GST should be charged on exported goods/services.')
                if transaction.get('residential_premises_gst'):
                    ato_comment = generate_ato_comment('residential_premises_gst')
                    comments.append(ato_comment or 'Residential property expense - Input Taxed (no GST credit claimable)')

            # Generate correcting journal entry
            try:
                correcting_journal = generate_correcting_journal(transaction)
            except Exception as e:
                print(f"Error in generate_correcting_journal: {e}")
                correcting_journal = {'narration': 'Error generating journal', 'entries': []}

            flagged_items.append({
                **transaction,
                'severity': ai_result.get('severity', 'high'),
                'comments': ' | '.join(comments) if comments else 'Requires review',
                'issues': [],  # AI issues removed as they duplicate rule-based comments
                'correcting_journal': correcting_journal
            })

        # Add remaining flagged items without AI review if over limit
        for transaction in rule_flagged[ai_review_limit:]:
            comments = []
            # Check if it's a personal expense - if so, skip business-specific rules like capitalization
            is_personal = transaction.get('life_insurance_personal') or transaction.get('personal_in_business_account')

            if transaction['account_coding_suspicious']:
                comments.append('Account coding may be incorrect')
            if transaction['alcohol_gst_error']:
                comments.append('Entertainment expense - should be GST Free Expenses')
            if transaction['input_taxed_gst_error']:
                comments.append('Input-taxed supply - GST incorrectly claimed (ATO: no GST on financial supplies)')
            if transaction['missing_gst_error']:
                comments.append('Should include GST (10%) - currently coded as GST Free')
            if not transaction['gst_calculation_correct']:
                comments.append('GST calculation error')
            if transaction.get('drawings_loan_error'):
                comments.append('Drawings/Loan account - should be BAS Excluded')
            # Skip asset/equipment capitalization rules for personal expenses (not relevant)
            if transaction.get('asset_capitalization_error') and not is_personal:
                comments.append('Asset over $20,000 - should be capitalized per ATO rules')
            if transaction.get('computer_equipment_expense') and not is_personal:
                comments.append('Computer equipment over $300 - should be capitalized as asset')
            if transaction.get('interest_gst_error'):
                comments.append('Interest should be GST Free Income or Input Taxed only')
            if transaction.get('other_income_error'):
                comments.append('Other Income coded as BAS Excluded - INCORRECT for business income. Commission, rebates, insurance payouts, hire/rental income, service income, fees should be GST on Income (taxable) or GST Free. BAS Excluded is only for private income, gifts, loans, capital contributions. Source: ATO BAS reporting rules')
            if transaction.get('sales_gst_error'):
                # Check if it's a commercial service coded as GST Free (specific error)
                desc = transaction.get('description', '').lower()
                gst_rate = transaction.get('gst_rate_name', '').lower()
                commercial_keywords = ['project management', 'consulting', 'advisory', 'training', 'software', 'it support', 'professional', 'service']
                is_commercial = any(kw in desc for kw in commercial_keywords)
                is_bas_excluded = 'bas excluded' in gst_rate
                if is_bas_excluded:
                    comments.append('Sales coded as BAS Excluded - INCORRECT. Sales must be GST on Income (10%) or GST Free Income. BAS Excluded is NEVER valid for sales.')
                elif is_commercial:
                    comments.append('Commercial/professional service coded as GST Free - INCORRECT. Services like project management, consulting, IT support, training should have GST (10%). Only medical, accredited education, childcare, or exports can be GST-free.')
                else:
                    comments.append('Sales coded as GST Free - verify this is a valid GST-free category (medical, accredited education, childcare, exports). If not, should be GST on Income (10%).')
            if transaction.get('motor_vehicle_gst_limit'):
                comments.append('Motor vehicle GST exceeds ATO car limit - max GST credit $6,334')
            if transaction.get('overseas_subscription_gst'):
                comments.append('Overseas subscription - GST credit INVALID (provide ABN for refund, reverse charge applies)')
            if transaction.get('government_charges_gst'):
                ato_comment = generate_ato_comment('government_charges_gst')
                comments.append(ato_comment or 'Government charge (council rates, stamp duty, land tax, rego fees) - NO GST. These are government levies, not taxable supplies.')
            if transaction.get('client_entertainment_gst'):
                ato_comment = generate_ato_comment('entertainment')
                comments.append(ato_comment or 'Client entertainment - NO GST credit claimable. Entertainment is non-deductible.')
            if transaction.get('staff_entertainment_gst'):
                ato_comment = generate_ato_comment('entertainment')
                comments.append(ato_comment or 'Staff entertainment - NO GST credit unless FBT is paid on the benefit.')
            if transaction.get('residential_premises_gst'):
                ato_comment = generate_ato_comment('residential_premises_gst')
                comments.append(ato_comment or 'Residential property expense - NO GST credit claimable (input-taxed)')
            if transaction.get('insurance_gst_error'):
                ato_comment = generate_ato_comment('life_insurance_personal')
                comments.append(ato_comment or 'Life/income protection insurance - NOT a deductible business expense. Recode to Owner Drawings. No GST credit claimable (input-taxed). Owner may claim on personal tax return.')
            if transaction.get('life_insurance_personal'):
                ato_comment = generate_ato_comment('life_insurance_personal')
                comments.append(ato_comment or 'Life/income protection insurance - NOT a deductible business expense (ATO). Personal insurance for owner should be coded to Owner Drawings. Owner may claim income protection on their personal tax return.')
            if transaction.get('personal_in_business_account'):
                comments.append('Personal expense in business account - NOT deductible. Recode to Owner Drawings (personal expenses cannot be claimed as business deductions).')
            if transaction.get('grants_sponsorship_gst') == 'sponsorship_no_gst':
                comments.append('Sponsorship income without GST - sponsorship is TAXABLE (GST applies) as it involves promotional services in return.')
            if transaction.get('grants_sponsorship_gst') == 'grant_with_gst':
                ato_comment = generate_ato_comment('grants_sponsorship_gst')
                comments.append(ato_comment or 'Grant income with GST charged - grants are typically GST-FREE unless binding supply obligation exists.')
            if transaction.get('wages_gst_error'):
                ato_comment = generate_ato_comment('wages_gst_error')
                comments.append(ato_comment or 'Wages/salaries/super - should be BAS Excluded (no GST)')
            if transaction.get('allowance_gst_error'):
                comments.append('Allowance - NO GST credit (not a purchase from supplier)')
            if transaction.get('reimbursement_gst'):
                comments.append('Reimbursement > $82.50 - verify tax invoice exists for GST credit')
            if transaction.get('voucher_gst'):
                comments.append('Voucher/gift card - check face value (no GST) vs non-face value (GST at sale)')
            if transaction.get('general_expenses'):
                comments.append('General/Sundry Expenses - recode to specific category (audit risk)')
            if transaction.get('travel_gst') == 'international_with_gst':
                ato_comment = generate_ato_comment('travel_gst_international')
                comments.append(ato_comment or 'International travel - GST FREE. Cannot claim GST credits on international travel.')
            elif transaction.get('travel_gst') == 'domestic_no_gst':
                ato_comment = generate_ato_comment('travel_gst_domestic')
                comments.append(ato_comment or 'Domestic travel (within Australia) - TAXABLE. Should include GST (10%).')
            if transaction.get('payment_processor_fees') == 'paypal_with_gst':
                ato_comment = generate_ato_comment('paypal_fees')
                comments.append(ato_comment or 'PayPal fees - NO GST. PayPal (Singapore) does not charge GST on transaction fees. Recode to Input Taxed. GST should be $0.')
            elif transaction.get('payment_processor_fees') in ['stripe_no_gst', 'merchant_no_gst']:
                ato_comment = generate_ato_comment('merchant_fees')
                comments.append(ato_comment or 'Merchant/Stripe/eBay fees - GST INCLUDED. These fees include GST and credits can be claimed. Recode to GST on Expenses.')
            elif transaction.get('payment_processor_fees'):
                comments.append('Payment processor fee GST issue - PayPal (no GST), Stripe/eBay/bank (GST included)')
            if transaction.get('fines_penalties_gst'):
                ato_comment = generate_ato_comment('fines_penalties_gst')
                comments.append(ato_comment or 'Fine/penalty - BAS Excluded (non-reportable, no GST)')
            if transaction.get('donations_gst'):
                ato_comment = generate_ato_comment('donations_gst')
                comments.append(ato_comment or 'Donation - NO GST applies.')
            if transaction.get('property_gst_withholding'):
                comments.append('Property purchase - check GST withholding obligations')
            if transaction.get('livestock_gst'):
                comments.append('Livestock sale - live animals are TAXABLE (not GST-free)')
            if transaction.get('asset_disposal_gst'):
                comments.append('Asset disposal - business asset sales are TAXABLE')
            if transaction.get('export_gst_error'):
                ato_comment = generate_ato_comment('export_gst_error')
                comments.append(ato_comment or 'Export sale - should be GST-FREE (no GST charged)')
            if transaction.get('borrowing_expenses_error'):
                comments.append('Borrowing expenses > $100 - must be spread over 5 years')

            # Generate correcting journal entry
            try:
                correcting_journal = generate_correcting_journal(transaction)
            except Exception as e:
                print(f"Error in generate_correcting_journal: {e}")
                correcting_journal = {'narration': 'Error generating journal', 'entries': []}

            flagged_items.append({
                **transaction,
                'severity': 'high',
                'comments': ' | '.join(comments) if comments else 'Requires review',
                'issues': [],
                'correcting_journal': correcting_journal
            })

        # Store results for download (only flagged items to save memory on large datasets)
        session['review_results'] = {
            'transactions': [],  # Don't store all transactions - too large for session
            'flagged_items': flagged_items,
            'from_date': '',
            'to_date': '',
            'tenant_name': company_name,
            'review_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'total_count': len(transactions)
        }

        return jsonify({
            'total_transactions': len(transactions),
            'flagged_count': len(flagged_items),
            'flagged_items': flagged_items,
            'company_name': company_name,
            'period': period
        })

    except Exception as e:
        import traceback
        error_msg = str(e) if str(e) else f"Unexpected error: {type(e).__name__}"
        print(f"Upload review error: {error_msg}")
        print(f"Traceback: {traceback.format_exc()}")
        return jsonify({'error': error_msg, 'trace': traceback.format_exc()}), 500


@app.route('/api/accounts')
@login_required
def get_accounts():
    """Get chart of accounts from Xero"""
    if 'access_token' not in session:
        return jsonify({'error': 'Not authenticated'}), 401

    data = xero_api_request('Accounts')
    if data:
        accounts = [{'code': a.get('Code', ''), 'name': a.get('Name', ''), 'type': a.get('Type', '')}
                   for a in data.get('Accounts', [])]
        return jsonify({'accounts': accounts})
    return jsonify({'error': 'Failed to fetch accounts'}), 500


@app.route('/api/health')
def health_check():
    """Simple health check - no auth required"""
    return jsonify({
        'status': 'ok',
        'version': '2.0',
        'scopes': XERO_SCOPES
    })


@app.route('/api/test-journals')
@login_required
def test_journals():
    """Test endpoint to debug Journals API"""
    try:
        if 'access_token' not in session:
            return jsonify({'error': 'Not authenticated - no access_token in session'}), 401

        if 'tenant_id' not in session:
            return jsonify({'error': 'Not authenticated - no tenant_id in session'}), 401

        result = {
            'tenant_id': session.get('tenant_id'),
            'tenant_name': session.get('tenant_name'),
            'has_access_token': 'access_token' in session,
        }

        # Test Journals API directly
        headers = {
            'Authorization': f"Bearer {session['access_token']}",
            'Xero-tenant-id': session['tenant_id'],
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        }

        # Try Journals endpoint
        try:
            url = f"{XERO_API_URL}/Journals"
            response = requests.get(url, headers=headers, params={'offset': 0})
            result['journals_status'] = response.status_code
            result['journals_response'] = response.text[:1000] if response.text else 'Empty'
        except Exception as e:
            result['journals_error'] = str(e)

        # Try BankTransactions endpoint
        try:
            url2 = f"{XERO_API_URL}/BankTransactions"
            response2 = requests.get(url2, headers=headers)
            result['bank_txn_status'] = response2.status_code
            if response2.status_code == 200:
                result['bank_txn_count'] = len(response2.json().get('BankTransactions', []))
            else:
                result['bank_txn_response'] = response2.text[:500]
        except Exception as e:
            result['bank_txn_error'] = str(e)

        # Try Invoices endpoint
        try:
            url3 = f"{XERO_API_URL}/Invoices"
            response3 = requests.get(url3, headers=headers)
            result['invoices_status'] = response3.status_code
            if response3.status_code == 200:
                result['invoices_count'] = len(response3.json().get('Invoices', []))
            else:
                result['invoices_response'] = response3.text[:500]
        except Exception as e:
            result['invoices_error'] = str(e)

        return jsonify(result)
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/api/run-review', methods=['POST'])
@login_required
def run_review():
    """Run BAS review on Activity Statement transactions from Xero"""
    try:
        if 'access_token' not in session:
            return jsonify({'error': 'Not authenticated'}), 401

        # Get parameters - use provided dates or default to current quarter
        data = request.json or {}
        from_date_str = data.get('start_date') or data.get('from_date')
        to_date_str = data.get('end_date') or data.get('to_date')
        review_mode = data.get('review_mode', 'quick')  # 'quick' or 'deep'

        # Default to current Australian financial year if no dates provided
        # Australian FY runs July 1 to June 30
        # FY2026 = July 1, 2025 to June 30, 2026
        if not from_date_str or not to_date_str:
            today = datetime.now()
            # Determine current financial year
            if today.month >= 7:
                # July onwards = current year's FY
                fy_start_year = today.year
            else:
                # Jan-June = previous year's FY
                fy_start_year = today.year - 1

            from_date = datetime(fy_start_year, 7, 1)  # July 1
            to_date = datetime(fy_start_year + 1, 6, 30)  # June 30

            # If we're in the current FY, only go up to today
            if to_date > today:
                to_date = today

            from_date_str = from_date.strftime('%Y-%m-%d')
            to_date_str = to_date.strftime('%Y-%m-%d')

        # Convert date strings to datetime objects
        try:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d')
        except ValueError:
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

        transactions = []
        debug_info = []

        # Deep Scan Mode: Fetch 12 months of history first to detect patterns
        if review_mode == 'deep':
            print("Deep Scan mode: Analyzing 12 months of history for patterns...")
            debug_info.append("Deep Scan: Fetching 12 months of history for pattern detection")

            # Calculate 12 months back from review end date
            history_end = to_date
            history_start = history_end - timedelta(days=365)
            history_start_str = history_start.strftime('%Y-%m-%d')
            history_end_str = history_end.strftime('%Y-%m-%d')

            debug_info.append(f"Scanning history from {history_start_str} to {history_end_str}")

            # Fetch historical transactions for pattern detection
            # Need BOTH bank transactions AND bills (invoices) to capture all vendor data
            history_transactions = []

            # 1. Bank transactions (direct spend/receive money)
            bank_txns = fetch_xero_bank_transactions(history_start_str, history_end_str)
            if bank_txns:
                history_transactions.extend(bank_txns)
            debug_info.append(f"Got {len(bank_txns) if bank_txns else 0} bank transactions")

            # 2. Bills (Accounts Payable) - this is where Telstra etc would be
            bills = fetch_xero_invoices(history_start_str, history_end_str, 'ACCPAY')
            if bills:
                history_transactions.extend(bills)
            debug_info.append(f"Got {len(bills) if bills else 0} bills (ACCPAY)")

            debug_info.append(f"Total: {len(history_transactions)} transactions for pattern detection")

            if history_transactions:
                # Enrich with account names
                history_transactions = enrich_transactions_with_accounts(history_transactions)

                # Detect allocation patterns
                try:
                    patterns = detect_allocation_patterns(history_transactions)
                    set_allocation_patterns(patterns)
                    pattern_count = len([p for p in patterns.values() if p.get('is_split_allocation')])
                    debug_info.append(f"Detected {pattern_count} split allocation patterns from {len(history_transactions)} history transactions")
                    print(f"Deep Scan: Detected {pattern_count} split allocation patterns")
                    for vendor, pattern in patterns.items():
                        if pattern.get('is_split_allocation'):
                            acct_str = ', '.join([f"{a}: {p:.0%}" for a, p in pattern['accounts'].items()])
                            print(f"  - {vendor}: {acct_str} ({pattern['count']} transactions)")
                except Exception as e:
                    print(f"Error detecting patterns: {e}")
                    import traceback
                    traceback.print_exc()
                    debug_info.append(f"Pattern detection error: {e}")
            else:
                debug_info.append("No historical transactions found for pattern detection")
                set_allocation_patterns({})
        else:
            # Quick mode - clear any previous patterns
            set_allocation_patterns({})

        # Fetch transactions using Bills and Bank Transactions APIs
        # These return the CURRENT state of transactions, not audit trail like Journals
        # This avoids phantom entries from edited bills

        debug_info.append(f"Fetching bills from {from_date_str} to {to_date_str}")
        bills = fetch_xero_invoices(from_date_str, to_date_str, 'ACCPAY')
        debug_info.append(f"Got {len(bills) if bills else 0} bill line items")
        if bills:
            transactions.extend(bills)

        debug_info.append(f"Fetching sales invoices from {from_date_str} to {to_date_str}")
        sales = fetch_xero_invoices(from_date_str, to_date_str, 'ACCREC')
        debug_info.append(f"Got {len(sales) if sales else 0} sales invoice line items")
        if sales:
            transactions.extend(sales)

        debug_info.append(f"Fetching bank transactions from {from_date_str} to {to_date_str}")
        bank_txns = fetch_xero_bank_transactions(from_date_str, to_date_str)
        debug_info.append(f"Got {len(bank_txns) if bank_txns else 0} bank transactions")
        if bank_txns:
            transactions.extend(bank_txns)

        # Enrich transactions with account names from Chart of Accounts
        if transactions:
            transactions = enrich_transactions_with_accounts(transactions)

        if not transactions:
            # Return more detailed error with debug info
            return jsonify({
                'error': 'No transactions found for the selected period',
                'from_date': from_date_str,
                'to_date': to_date_str,
                'debug': debug_info
            }), 404

        # Run BAS review on transactions - rule-based first, then AI only for flagged items
        flagged_items = []
        rule_flagged = []

        # Infer business context from all transactions to understand industry/income sources
        try:
            business_context = set_business_context(transactions)
            print(f"Inferred business context: {business_context['industry']} (confidence: {business_context['confidence']:.1%})")
        except Exception as e:
            print(f"Error setting business context: {e}")

        # First pass: fast rule-based checks per ATO GST rules
        for transaction in transactions:
            try:
                transaction['account_coding_suspicious'] = check_account_coding(transaction)
            except Exception as e:
                print(f"Error in check_account_coding: {e}")
                transaction['account_coding_suspicious'] = False

            # Check for split allocation patterns (Deep Scan only)
            try:
                split_check = check_split_allocation_pattern(transaction)
                transaction['split_allocation_warning'] = split_check
            except Exception as e:
                print(f"Error in check_split_allocation_pattern: {e}")
                transaction['split_allocation_warning'] = None

            try:
                transaction['alcohol_gst_error'] = check_alcohol_gst(transaction)
            except Exception as e:
                print(f"Error in check_alcohol_gst: {e}")
                transaction['alcohol_gst_error'] = False

            try:
                transaction['input_taxed_gst_error'] = check_input_taxed_gst(transaction)
            except Exception as e:
                print(f"Error in check_input_taxed_gst: {e}")
                transaction['input_taxed_gst_error'] = False

            try:
                transaction['missing_gst_error'] = check_missing_gst(transaction)
            except Exception as e:
                print(f"Error in check_missing_gst: {e}")
                transaction['missing_gst_error'] = False

            try:
                transaction['gst_calculation_correct'] = check_gst_calculation(transaction)
            except Exception as e:
                print(f"Error in check_gst_calculation: {e}")
                transaction['gst_calculation_correct'] = True

            try:
                transaction['drawings_loan_error'] = check_drawings_loan_error(transaction)
            except Exception as e:
                print(f"Error in check_drawings_loan_error: {e}")
                transaction['drawings_loan_error'] = False

            try:
                transaction['personal_in_business_account'] = check_personal_expense_in_business_account(transaction)
            except Exception as e:
                print(f"Error in check_personal_expense_in_business_account: {e}")
                transaction['personal_in_business_account'] = False

            try:
                transaction['asset_capitalization_error'] = check_asset_capitalization(transaction)
            except Exception as e:
                print(f"Error in check_asset_capitalization: {e}")
                transaction['asset_capitalization_error'] = False

            try:
                transaction['computer_equipment_expense'] = check_computer_equipment_expense(transaction)
            except Exception as e:
                print(f"Error in check_computer_equipment_expense: {e}")
                transaction['computer_equipment_expense'] = False

            try:
                transaction['interest_gst_error'] = check_interest_gst_error(transaction)
            except Exception as e:
                print(f"Error in check_interest_gst_error: {e}")
                transaction['interest_gst_error'] = False

            try:
                transaction['other_income_error'] = check_other_income_error(transaction)
            except Exception as e:
                print(f"Error in check_other_income_error: {e}")
                transaction['other_income_error'] = False

            try:
                transaction['sales_gst_error'] = check_sales_gst_error(transaction)
            except Exception as e:
                print(f"Error in check_sales_gst_error: {e}")
                transaction['sales_gst_error'] = False

            try:
                transaction['motor_vehicle_gst_limit'] = check_motor_vehicle_gst_limit(transaction)
            except Exception as e:
                print(f"Error in check_motor_vehicle_gst_limit: {e}")
                transaction['motor_vehicle_gst_limit'] = False

            try:
                transaction['overseas_subscription_gst'] = check_overseas_subscription_gst(transaction)
            except Exception as e:
                print(f"Error in check_overseas_subscription_gst: {e}")
                transaction['overseas_subscription_gst'] = False

            try:
                transaction['government_charges_gst'] = check_government_charges_gst(transaction)
            except Exception as e:
                print(f"Error in check_government_charges_gst: {e}")
                transaction['government_charges_gst'] = False

            try:
                transaction['client_entertainment_gst'] = check_client_entertainment_gst(transaction)
            except Exception as e:
                print(f"Error in check_client_entertainment_gst: {e}")
                transaction['client_entertainment_gst'] = False

            try:
                transaction['staff_entertainment_gst'] = check_staff_entertainment_gst(transaction)
            except Exception as e:
                print(f"Error in check_staff_entertainment_gst: {e}")
                transaction['staff_entertainment_gst'] = False

            try:
                transaction['residential_premises_gst'] = check_residential_premises_gst(transaction)
            except Exception as e:
                print(f"Error in check_residential_premises_gst: {e}")
                transaction['residential_premises_gst'] = False

            try:
                transaction['insurance_gst_error'] = check_insurance_gst_error(transaction)
            except Exception as e:
                print(f"Error in check_insurance_gst_error: {e}")
                transaction['insurance_gst_error'] = False

            try:
                transaction['life_insurance_personal'] = check_life_insurance_personal(transaction)
            except Exception as e:
                print(f"Error in check_life_insurance_personal: {e}")
                transaction['life_insurance_personal'] = False

            try:
                transaction['grants_sponsorship_gst'] = check_grants_sponsorship_gst(transaction)
            except Exception as e:
                print(f"Error in check_grants_sponsorship_gst: {e}")
                transaction['grants_sponsorship_gst'] = False

            try:
                transaction['wages_gst_error'] = check_wages_gst_error(transaction)
            except Exception as e:
                print(f"Error in check_wages_gst_error: {e}")
                transaction['wages_gst_error'] = False

            try:
                transaction['allowance_gst_error'] = check_allowance_gst_error(transaction)
            except Exception as e:
                print(f"Error in check_allowance_gst_error: {e}")
                transaction['allowance_gst_error'] = False

            try:
                transaction['reimbursement_gst'] = check_reimbursement_gst(transaction)
            except Exception as e:
                print(f"Error in check_reimbursement_gst: {e}")
                transaction['reimbursement_gst'] = False

            try:
                transaction['voucher_gst'] = check_voucher_gst(transaction)
            except Exception as e:
                print(f"Error in check_voucher_gst: {e}")
                transaction['voucher_gst'] = False

            try:
                transaction['general_expenses'] = check_general_expenses(transaction)
            except Exception as e:
                print(f"Error in check_general_expenses: {e}")
                transaction['general_expenses'] = False

            try:
                transaction['travel_gst'] = check_travel_gst(transaction)
            except Exception as e:
                print(f"Error in check_travel_gst: {e}")
                transaction['travel_gst'] = False

            try:
                transaction['payment_processor_fees'] = check_payment_processor_fees(transaction)
            except Exception as e:
                print(f"Error in check_payment_processor_fees: {e}")
                transaction['payment_processor_fees'] = False

            try:
                transaction['fines_penalties_gst'] = check_fines_penalties_gst(transaction)
            except Exception as e:
                print(f"Error in check_fines_penalties_gst: {e}")
                transaction['fines_penalties_gst'] = False

            try:
                transaction['donations_gst'] = check_donations_gst(transaction)
            except Exception as e:
                print(f"Error in check_donations_gst: {e}")
                transaction['donations_gst'] = False

            try:
                transaction['property_gst_withholding'] = check_property_gst_withholding(transaction)
            except Exception as e:
                print(f"Error in check_property_gst_withholding: {e}")
                transaction['property_gst_withholding'] = False

            try:
                transaction['livestock_gst'] = check_livestock_gst(transaction)
            except Exception as e:
                print(f"Error in check_livestock_gst: {e}")
                transaction['livestock_gst'] = False

            try:
                transaction['asset_disposal_gst'] = check_asset_disposal_gst(transaction)
            except Exception as e:
                print(f"Error in check_asset_disposal_gst: {e}")
                transaction['asset_disposal_gst'] = False

            try:
                transaction['export_gst_error'] = check_export_gst_error(transaction)
            except Exception as e:
                print(f"Error in check_export_gst_error: {e}")
                transaction['export_gst_error'] = False

            try:
                transaction['borrowing_expenses_error'] = check_borrowing_expenses_error(transaction)
            except Exception as e:
                print(f"Error in check_borrowing_expenses_error: {e}")
                transaction['borrowing_expenses_error'] = False

            # Flag if any rule triggered (including new checks)
            # Note: split_allocation_warning removed from flagging - personal_in_business_account handles actual errors
            has_rule_issues = (
                transaction['account_coding_suspicious'] or
                transaction['alcohol_gst_error'] or
                transaction['input_taxed_gst_error'] or
                transaction['missing_gst_error'] or
                not transaction['gst_calculation_correct'] or
                transaction['drawings_loan_error'] or
                transaction.get('personal_in_business_account') or
                transaction['asset_capitalization_error'] or
                transaction['computer_equipment_expense'] or
                transaction['interest_gst_error'] or
                transaction['other_income_error'] or
                transaction['sales_gst_error'] or
                transaction['motor_vehicle_gst_limit'] or
                transaction['overseas_subscription_gst'] or
                transaction['government_charges_gst'] or
                transaction['client_entertainment_gst'] or
                transaction['staff_entertainment_gst'] or
                transaction['residential_premises_gst'] or
                transaction['insurance_gst_error'] or
                transaction['life_insurance_personal'] or
                transaction['grants_sponsorship_gst'] or
                transaction['wages_gst_error'] or
                transaction['allowance_gst_error'] or
                transaction['reimbursement_gst'] or
                transaction['voucher_gst'] or
                transaction['general_expenses'] or
                transaction['travel_gst'] or
                transaction['payment_processor_fees'] or
                transaction['fines_penalties_gst'] or
                transaction['donations_gst'] or
                transaction['property_gst_withholding'] or
                transaction['livestock_gst'] or
                transaction['asset_disposal_gst'] or
                transaction['export_gst_error'] or
                transaction['borrowing_expenses_error']
            )

            # Skip flagging for correctly coded Telstra transactions
            # (Telstra business in Telephone, Telstra personal in Drawings)
            if has_rule_issues and is_correctly_coded_telstra(transaction):
                has_rule_issues = False

            if has_rule_issues:
                rule_flagged.append(transaction)

        # Second pass: AI review for flagged items (batch processing)
        # Limit AI review to 200 items to prevent timeout on large datasets
        MAX_AI_REVIEW = 200
        ai_review_limit = min(len(rule_flagged), MAX_AI_REVIEW)

        if len(rule_flagged) > MAX_AI_REVIEW:
            print(f"Large dataset: {len(rule_flagged)} flagged items, limiting AI review to {MAX_AI_REVIEW}")

        # Use batch AI review for faster processing (5 transactions per API call)
        try:
            ai_results = review_batch_with_ai(rule_flagged[:ai_review_limit], batch_size=5)
        except Exception as e:
            print(f"Error in batch AI review: {e}")
            ai_results = [{'has_issues': True, 'severity': 'high', 'comments': '', 'issues': []} for _ in range(ai_review_limit)]

        for i, transaction in enumerate(rule_flagged[:ai_review_limit]):
            ai_result = ai_results[i] if i < len(ai_results) else {'has_issues': True, 'severity': 'high', 'comments': '', 'issues': []}

            # Use simplified AI-generated comment, with rule-based fallback if AI doesn't provide useful comment
            comments = []
            ai_comment = ai_result.get('comments', '').strip()

            # Check if AI comment is useful (not empty or too generic)
            # Include "correctly applied/coded/recorded" as generic since AI may miss the actual issue
            generic_phrases = ['requires review', 'please review', 'review required', 'ok -', 'appears correct',
                              'correctly applied', 'correctly coded', 'correctly recorded', 'no issues',
                              'looks correct', 'is correct', 'recorded correctly', 'coded correctly']
            is_useful_ai_comment = ai_comment and len(ai_comment) > 20 and not any(phrase in ai_comment.lower() for phrase in generic_phrases)

            # Override AI comment if it uses incorrect terminology
            if is_useful_ai_comment and ai_comment:
                ai_lower = ai_comment.lower()
                # Wages are BAS Excluded, NOT input-taxed - force rule-based if AI uses wrong term
                if transaction.get('wages_gst_error') and ('input-taxed' in ai_lower or 'input taxed' in ai_lower):
                    is_useful_ai_comment = False
                # Government charges have NO GST (not input-taxed, not taxable) - force rule-based if wrong
                if transaction.get('government_charges_gst') and ('input-taxed' in ai_lower or 'input taxed' in ai_lower or 'taxable supply' in ai_lower or 'taxable supplies' in ai_lower):
                    is_useful_ai_comment = False
                # Council/water rates should not mention "residential" or "taxable"
                if ('council' in ai_lower or 'water' in ai_lower) and 'rates' in ai_lower and ('residential' in ai_lower or 'taxable supply' in ai_lower or 'taxable supplies' in ai_lower):
                    is_useful_ai_comment = False
                # Body corporate assumed residential without evidence
                if 'body corporate' in ai_lower and 'residential' in ai_lower and 'residential' not in transaction.get('description', '').lower():
                    is_useful_ai_comment = False
                # International travel flagged but AI says taxable - international flights are GST-FREE
                if transaction.get('travel_gst') == 'international_with_gst' and 'taxable' in ai_lower:
                    is_useful_ai_comment = False
                # Life/income protection insurance - always use rule-based for consistent Owner Drawings advice
                if transaction.get('life_insurance_personal') or transaction.get('insurance_gst_error'):
                    # Force rule-based comment which includes Owner Drawings recode advice
                    is_useful_ai_comment = False
                # Payment processor fees - always use rule-based for specific PayPal/Stripe guidance
                if transaction.get('payment_processor_fees'):
                    is_useful_ai_comment = False
                # Government charges - always use rule-based (AI incorrectly says "input-taxed" or "residential")
                if transaction.get('government_charges_gst'):
                    is_useful_ai_comment = False
                # Export GST errors - always use rule-based for consistent GSTR 2002/6 reference
                if transaction.get('export_gst_error'):
                    is_useful_ai_comment = False

            # Always prioritize AI comments, use rule-based as fallback only
            if is_useful_ai_comment:
                comments.append(ai_comment)
            else:
                # Fallback to rule-based comments when AI doesn't provide useful info
                # Check if it's a personal expense - if so, skip business-specific rules like capitalization
                is_personal = transaction.get('life_insurance_personal') or transaction.get('personal_in_business_account')

                if transaction.get('life_insurance_personal'):
                    ato_comment = generate_ato_comment('life_insurance_personal')
                    comments.append(ato_comment or 'Life/income protection insurance - NOT a deductible business expense (ATO). Recode to Owner Drawings. Owner may claim income protection on personal tax return.')
                if transaction.get('personal_in_business_account'):
                    comments.append('Personal expense in business account - NOT deductible. Recode to Owner Drawings (personal expenses cannot be claimed as business deductions).')
                # Skip asset/equipment capitalization rules for personal expenses (not relevant)
                if transaction.get('asset_capitalization_error') and not is_personal:
                    comments.append('Asset over $20,000 - should be capitalized per ATO instant asset write-off rules, not expensed')
                if transaction.get('computer_equipment_expense') and not is_personal:
                    comments.append('Computer equipment over $300 - should be capitalized as asset, not expensed to Office Supplies')
                if transaction.get('insurance_gst_error'):
                    ato_comment = generate_ato_comment('life_insurance_personal')
                    comments.append(ato_comment or 'Life/income protection insurance - NOT a deductible business expense. Recode to Owner Drawings. No GST credit claimable (input-taxed). Owner may claim on personal tax return.')
                if transaction.get('wages_gst_error'):
                    ato_comment = generate_ato_comment('wages_gst_error')
                    comments.append(ato_comment or 'Wages/salaries/super - should be BAS Excluded (no GST)')
                if transaction.get('alcohol_gst_error') or transaction.get('client_entertainment_gst') or transaction.get('staff_entertainment_gst'):
                    ato_comment = generate_ato_comment('entertainment')
                    comments.append(ato_comment or 'Entertainment expense - NO GST credit claimable. Entertainment is non-deductible and GST credits blocked unless FBT is paid.')
                if transaction.get('missing_gst_error'):
                    comments.append('Should include GST (10%) - currently coded as GST Free')
                if transaction.get('input_taxed_gst_error'):
                    ato_comment = generate_ato_comment('input_taxed_gst_error')
                    comments.append(ato_comment or 'Input-taxed supply - GST incorrectly claimed (no GST credit on financial supplies)')
                if transaction.get('general_expenses'):
                    comments.append('General/Sundry Expenses - recode to specific category to reduce audit risk')
                if transaction.get('drawings_loan_error'):
                    comments.append('Drawings/Loan account - should be BAS Excluded')
                if transaction.get('fines_penalties_gst'):
                    ato_comment = generate_ato_comment('fines_penalties_gst')
                    comments.append(ato_comment or 'Fine/penalty - BAS Excluded (non-reportable, no GST)')
                if transaction.get('government_charges_gst'):
                    ato_comment = generate_ato_comment('government_charges_gst')
                    comments.append(ato_comment or 'Government charge - NO GST applies (not a taxable supply)')
                if transaction.get('donations_gst'):
                    ato_comment = generate_ato_comment('donations_gst')
                    comments.append(ato_comment or 'Donation - NO GST applies. Use GST Free Expenses for P&L accounts.')
                if transaction.get('travel_gst') == 'international_with_gst':
                    ato_comment = generate_ato_comment('travel_gst_international')
                    comments.append(ato_comment or 'International travel - GST FREE. Cannot claim GST credits on international travel expenses.')
                elif transaction.get('travel_gst') == 'domestic_no_gst':
                    ato_comment = generate_ato_comment('travel_gst_domestic')
                    comments.append(ato_comment or 'Domestic travel (within Australia) - TAXABLE. Should include GST (10%). Domestic flights, hotels, taxis are GST taxable.')
                if transaction.get('grants_sponsorship_gst') == 'grant_with_gst':
                    ato_comment = generate_ato_comment('grants_sponsorship_gst')
                    comments.append(ato_comment or 'Grant income with GST charged - grants are typically GST-FREE unless binding supply obligation exists.')
                elif transaction.get('grants_sponsorship_gst') == 'sponsorship_no_gst':
                    comments.append('Sponsorship income without GST - sponsorship is TAXABLE (GST applies) as it involves promotional services in return.')
                if transaction.get('export_gst_error'):
                    ato_comment = generate_ato_comment('export_gst_error')
                    comments.append(ato_comment or 'Export sale with GST charged - exports should be GST-FREE. No GST should be charged on exported goods/services.')
                if transaction.get('residential_premises_gst'):
                    ato_comment = generate_ato_comment('residential_premises_gst')
                    comments.append(ato_comment or 'Residential property expense - Input Taxed (no GST credit claimable)')

            # Generate correcting journal entry
            try:
                correcting_journal = generate_correcting_journal(transaction)
            except Exception as e:
                print(f"Error in generate_correcting_journal: {e}")
                correcting_journal = {'narration': 'Error generating journal', 'entries': []}

            flagged_items.append({
                **transaction,
                'severity': ai_result.get('severity', 'high'),
                'comments': ' | '.join(comments) if comments else 'Requires review',
                'issues': [],  # AI issues removed as they duplicate rule-based comments
                'correcting_journal': correcting_journal
            })

        # Add remaining flagged items without AI review if over limit
        for transaction in rule_flagged[ai_review_limit:]:
            comments = []
            # Check if it's a personal expense - if so, skip business-specific rules like capitalization
            is_personal = transaction.get('life_insurance_personal') or transaction.get('personal_in_business_account')

            if transaction['account_coding_suspicious']:
                comments.append('Account coding may be incorrect')
            if transaction['alcohol_gst_error']:
                comments.append('Entertainment expense - should be GST Free Expenses')
            if transaction['input_taxed_gst_error']:
                comments.append('Input-taxed supply - GST incorrectly claimed (ATO: no GST on financial supplies)')
            if transaction['missing_gst_error']:
                comments.append('Should include GST (10%) - currently coded as GST Free')
            if not transaction['gst_calculation_correct']:
                comments.append('GST calculation error')
            if transaction.get('drawings_loan_error'):
                comments.append('Drawings/Loan account - should be BAS Excluded')
            # Skip asset/equipment capitalization rules for personal expenses (not relevant)
            if transaction.get('asset_capitalization_error') and not is_personal:
                comments.append('Asset over $20,000 - should be capitalized per ATO rules')
            if transaction.get('computer_equipment_expense') and not is_personal:
                comments.append('Computer equipment over $300 - should be capitalized as asset')
            if transaction.get('interest_gst_error'):
                comments.append('Interest should be GST Free Income or Input Taxed only')
            if transaction.get('other_income_error'):
                comments.append('Other Income coded as BAS Excluded - INCORRECT for business income. Commission, rebates, insurance payouts, hire/rental income, service income, fees should be GST on Income (taxable) or GST Free. BAS Excluded is only for private income, gifts, loans, capital contributions. Source: ATO BAS reporting rules')
            if transaction.get('sales_gst_error'):
                # Check if it's a commercial service coded as GST Free (specific error)
                desc = transaction.get('description', '').lower()
                gst_rate = transaction.get('gst_rate_name', '').lower()
                commercial_keywords = ['project management', 'consulting', 'advisory', 'training', 'software', 'it support', 'professional', 'service']
                is_commercial = any(kw in desc for kw in commercial_keywords)
                is_bas_excluded = 'bas excluded' in gst_rate
                if is_bas_excluded:
                    comments.append('Sales coded as BAS Excluded - INCORRECT. Sales must be GST on Income (10%) or GST Free Income. BAS Excluded is NEVER valid for sales.')
                elif is_commercial:
                    comments.append('Commercial/professional service coded as GST Free - INCORRECT. Services like project management, consulting, IT support, training should have GST (10%). Only medical, accredited education, childcare, or exports can be GST-free.')
                else:
                    comments.append('Sales coded as GST Free - verify this is a valid GST-free category (medical, accredited education, childcare, exports). If not, should be GST on Income (10%).')
            if transaction.get('motor_vehicle_gst_limit'):
                comments.append('Motor vehicle GST exceeds ATO car limit - max GST credit $6,334')
            if transaction.get('overseas_subscription_gst'):
                comments.append('Overseas subscription - GST credit INVALID (provide ABN for refund, reverse charge applies)')
            if transaction.get('government_charges_gst'):
                ato_comment = generate_ato_comment('government_charges_gst')
                comments.append(ato_comment or 'Government charge (council rates, stamp duty, land tax, rego fees) - NO GST. These are government levies, not taxable supplies.')
            if transaction.get('client_entertainment_gst'):
                ato_comment = generate_ato_comment('entertainment')
                comments.append(ato_comment or 'Client entertainment - NO GST credit claimable. Entertainment is non-deductible.')
            if transaction.get('staff_entertainment_gst'):
                ato_comment = generate_ato_comment('entertainment')
                comments.append(ato_comment or 'Staff entertainment - NO GST credit unless FBT is paid on the benefit.')
            if transaction.get('residential_premises_gst'):
                ato_comment = generate_ato_comment('residential_premises_gst')
                comments.append(ato_comment or 'Residential property expense - NO GST credit claimable (input-taxed)')
            if transaction.get('insurance_gst_error'):
                ato_comment = generate_ato_comment('life_insurance_personal')
                comments.append(ato_comment or 'Life/income protection insurance - NOT a deductible business expense. Recode to Owner Drawings. No GST credit claimable (input-taxed). Owner may claim on personal tax return.')
            if transaction.get('life_insurance_personal'):
                ato_comment = generate_ato_comment('life_insurance_personal')
                comments.append(ato_comment or 'Life/income protection insurance - NOT a deductible business expense (ATO). Personal insurance for owner should be coded to Owner Drawings. Owner may claim income protection on their personal tax return.')
            if transaction.get('personal_in_business_account'):
                comments.append('Personal expense in business account - NOT deductible. Recode to Owner Drawings (personal expenses cannot be claimed as business deductions).')
            if transaction.get('grants_sponsorship_gst') == 'sponsorship_no_gst':
                comments.append('Sponsorship income without GST - sponsorship is TAXABLE (GST applies) as it involves promotional services in return.')
            if transaction.get('grants_sponsorship_gst') == 'grant_with_gst':
                ato_comment = generate_ato_comment('grants_sponsorship_gst')
                comments.append(ato_comment or 'Grant income with GST charged - grants are typically GST-FREE unless binding supply obligation exists.')
            if transaction.get('wages_gst_error'):
                ato_comment = generate_ato_comment('wages_gst_error')
                comments.append(ato_comment or 'Wages/salaries/super - should be BAS Excluded (no GST)')
            if transaction.get('allowance_gst_error'):
                comments.append('Allowance - NO GST credit (not a purchase from supplier)')
            if transaction.get('reimbursement_gst'):
                comments.append('Reimbursement > $82.50 - verify tax invoice exists for GST credit')
            if transaction.get('voucher_gst'):
                comments.append('Voucher/gift card - check face value (no GST) vs non-face value (GST at sale)')
            if transaction.get('general_expenses'):
                comments.append('General/Sundry Expenses - recode to specific category (audit risk)')
            if transaction.get('travel_gst') == 'international_with_gst':
                ato_comment = generate_ato_comment('travel_gst_international')
                comments.append(ato_comment or 'International travel - GST FREE. Cannot claim GST credits on international travel.')
            elif transaction.get('travel_gst') == 'domestic_no_gst':
                ato_comment = generate_ato_comment('travel_gst_domestic')
                comments.append(ato_comment or 'Domestic travel (within Australia) - TAXABLE. Should include GST (10%).')
            if transaction.get('payment_processor_fees') == 'paypal_with_gst':
                ato_comment = generate_ato_comment('paypal_fees')
                comments.append(ato_comment or 'PayPal fees - NO GST. PayPal (Singapore) does not charge GST on transaction fees. Recode to Input Taxed. GST should be $0.')
            elif transaction.get('payment_processor_fees') in ['stripe_no_gst', 'merchant_no_gst']:
                ato_comment = generate_ato_comment('merchant_fees')
                comments.append(ato_comment or 'Merchant/Stripe/eBay fees - GST INCLUDED. These fees include GST and credits can be claimed. Recode to GST on Expenses.')
            elif transaction.get('payment_processor_fees'):
                comments.append('Payment processor fee GST issue - PayPal (no GST), Stripe/eBay/bank (GST included)')
            if transaction.get('fines_penalties_gst'):
                ato_comment = generate_ato_comment('fines_penalties_gst')
                comments.append(ato_comment or 'Fine/penalty - BAS Excluded (non-reportable, no GST)')
            if transaction.get('donations_gst'):
                ato_comment = generate_ato_comment('donations_gst')
                comments.append(ato_comment or 'Donation - NO GST applies.')
            if transaction.get('property_gst_withholding'):
                comments.append('Property purchase - check GST withholding obligations')
            if transaction.get('livestock_gst'):
                comments.append('Livestock sale - live animals are TAXABLE (not GST-free)')
            if transaction.get('asset_disposal_gst'):
                comments.append('Asset disposal - business asset sales are TAXABLE')
            if transaction.get('export_gst_error'):
                ato_comment = generate_ato_comment('export_gst_error')
                comments.append(ato_comment or 'Export sale - should be GST-FREE (no GST charged)')
            if transaction.get('borrowing_expenses_error'):
                comments.append('Borrowing expenses > $100 - must be spread over 5 years')

            correcting_journal = generate_correcting_journal(transaction)

            flagged_items.append({
                **transaction,
                'severity': 'high',
                'comments': ' | '.join(comments) if comments else 'Requires review',
                'issues': [],
                'correcting_journal': correcting_journal
            })

        # Store results in session for download (only flagged items to save memory)
        session['review_results'] = {
            'transactions': [],  # Don't store all transactions - too large for session
            'flagged_items': flagged_items,
            'from_date': from_date_str,
            'to_date': to_date_str,
            'tenant_name': session.get('tenant_name', ''),
            'review_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'total_count': len(transactions)
        }

        # Include pattern debug info (ALL patterns, not just split ones)
        patterns = get_allocation_patterns()
        pattern_debug = {}
        for vendor, pattern in patterns.items():
            pattern_debug[vendor] = {
                'accounts': pattern.get('accounts', {}),
                'count': pattern.get('count', 0),
                'is_split': pattern.get('is_split_allocation', False)
            }

        # Prepare all transactions for complete audit trail
        all_transactions_data = []
        for i, txn in enumerate(transactions):
            all_transactions_data.append({
                'row_number': i + 1,
                'date': txn.get('date'),
                'account_code': txn.get('account_code'),
                'account_name': txn.get('account', ''),
                'description': txn.get('description', ''),
                'gross': txn.get('gross', 0),
                'gst': txn.get('gst', 0),
                'net': txn.get('net', 0),
                'gst_rate_name': txn.get('gst_rate_name', ''),
                'source': txn.get('source', ''),
                'reference': txn.get('reference', ''),
                'contact': txn.get('contact', ''),
                'xero_url': txn.get('xero_url', '')
            })

        # Auto-save review to Cloudflare D1 (if configured)
        saved_to_cloud = False
        statement_id = None
        try:
            config = get_cloudflare_config()
            if config['account_id'] and config['api_token'] and config['database_id']:
                # Fetch BAS report data if available
                bas_report_data = {}
                try:
                    bas_report = fetch_xero_bas_report(from_date_str, to_date_str)
                    if bas_report:
                        bas_report_data = bas_report
                except Exception as e:
                    print(f"Could not fetch BAS report for auto-save: {e}")

                # Prepare review data for saving
                review_data = {
                    'tenant_id': session.get('tenant_id', ''),
                    'tenant_name': session.get('tenant_name', ''),
                    'user_email': current_user.email if current_user else '',
                    'period_start': from_date_str,
                    'period_end': to_date_str,
                    'total_transactions': len(transactions),
                    'flagged_count': len(flagged_items),
                    'high_severity_count': len([f for f in flagged_items if f.get('severity') == 'high']),
                    'medium_severity_count': len([f for f in flagged_items if f.get('severity') == 'medium']),
                    'low_severity_count': len([f for f in flagged_items if f.get('severity') == 'low']),
                    'bas_report_data': bas_report_data,
                    'review_summary': {
                        'review_mode': review_mode,
                        'patterns_detected': len(pattern_debug)
                    },
                    'flagged_items': flagged_items,
                    'all_transactions': all_transactions_data
                }

                save_result = save_review_to_d1(review_data)
                if save_result.get('success'):
                    saved_to_cloud = True
                    statement_id = save_result.get('statement_id')
                    print(f"Auto-saved review to D1: {statement_id}")
        except Exception as e:
            print(f"Auto-save to D1 failed (non-blocking): {e}")

        return jsonify({
            'total_transactions': len(transactions),
            'flagged_count': len(flagged_items),
            'flagged_items': flagged_items,
            'all_transactions': all_transactions_data,
            'patterns_detected': pattern_debug,
            'company_name': session.get('tenant_name', ''),
            'saved_to_cloud': saved_to_cloud,
            'statement_id': statement_id
        })
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


def generate_correcting_journal(transaction):
    """Generate suggested correcting journal entry for flagged transaction"""
    description = transaction.get('description', '').lower()
    account_code = str(transaction.get('account_code', '') or '').strip()
    account_name = str(transaction.get('account', '') or '').strip()
    gross = abs(float(transaction.get('gross', 0) or 0))
    gst = abs(float(transaction.get('gst', 0) or 0))
    net = abs(float(transaction.get('net', 0) or 0))

    # Debug: Log if account info is missing
    if not account_code:
        print(f"WARNING: Empty account_code for transaction: {transaction.get('description', 'Unknown')}")
        print(f"  Full transaction: account_code={transaction.get('account_code')}, account={transaction.get('account')}")
        # Use placeholder to prevent blank entries in Xero
        account_code = '999'  # Suspense/Unknown account
        account_name = account_name or 'Unknown Account'

    journal_entries = []
    narration = f"Reallocate: {transaction.get('description', '')}"

    # Track what corrections have been made to avoid duplicates
    gst_correction_done = False
    recode_done = False

    # Determine the CORRECT tax code based on all flagged errors (priority order)
    # This ensures we use the right tax code when recoding AND avoid duplicate GST fix journals
    correct_tax_code = 'GST on Expenses' if gst > 0 else 'GST Free'  # default

    # Priority 1: BAS Excluded items (fines, wages, allowances) - NOT reportable on BAS
    if (transaction.get('fines_penalties_gst') or
        transaction.get('wages_gst_error') or
        transaction.get('allowance_gst_error')):
        correct_tax_code = 'BAS Excluded'
    # Priority 2: Input Taxed items (residential, insurance, input_taxed)
    elif (transaction.get('residential_premises_gst') or
          transaction.get('insurance_gst_error') or
          transaction.get('input_taxed_gst_error')):
        correct_tax_code = 'Input Taxed'
    # Priority 3: GST Free items (entertainment, government charges, international travel, donations)
    elif (transaction.get('alcohol_gst_error') or
          transaction.get('client_entertainment_gst') or
          transaction.get('staff_entertainment_gst') or
          transaction.get('government_charges_gst') or
          transaction.get('donations_gst') or
          transaction.get('travel_gst') == 'international_with_gst'):
        correct_tax_code = 'GST Free Expenses'

    # Check what type of error this is
    if transaction.get('account_coding_suspicious'):
        # Skip recoding suggestion if already in a valid travel account
        # Employee reimbursements for travel are normal and shouldn't be recoded
        travel_accounts = ['travel', 'travel national', 'travel - national', 'domestic travel',
                          'travel international', 'accommodation', 'airfare', 'motor vehicle']
        is_travel_account = any(keyword in account_name.lower() for keyword in travel_accounts)

        # Skip if this is a PERSONAL item - let personal_in_business_account handle it
        # Personal items should go to Owner Drawings, not suggested business account
        is_personal = transaction.get('personal_in_business_account')

        # Skip if this is ENTERTAINMENT - don't recode to different account, just fix tax code
        is_entertainment = (transaction.get('alcohol_gst_error') or
                           transaction.get('client_entertainment_gst') or
                           transaction.get('staff_entertainment_gst'))
        # Also skip if account is already Entertainment (regardless of flags)
        is_entertainment_account = 'entertainment' in account_name.lower()
        # Skip if account is Fines & Penalties - already correct account for fines
        is_fines_account = 'fines' in account_name.lower() or 'penalties' in account_name.lower()

        if is_travel_account or is_personal or is_entertainment or is_entertainment_account or is_fines_account:
            # Don't suggest recoding - travel accounts are fine, personal items have dedicated handler
            pass
        else:
            # Determine suggested correct account based on description
            suggested_account = suggest_correct_account(description)

            # Only suggest recoding if we have a specific account (not generic "General Expenses")
            if suggested_account and suggested_account['name'] != 'General Expenses':
                # Standardized description format: "[Wrong] to [Correct] - [Details]"
                trans_desc = transaction.get('description', '')[:50] or 'No description'
                std_desc = f"{account_name} to {suggested_account['name']} - {trans_desc}"

                # Debit to correct account first, then credit to reverse wrong account
                # Use the CORRECT tax code determined above
                journal_entries.append({
                    'line': 1,
                    'account_code': suggested_account['code'],
                    'account_name': suggested_account['name'],
                    'debit': gross if gross > 0 else 0,
                    'credit': 0 if gross > 0 else gross,
                    'tax_code': correct_tax_code,
                    'description': std_desc
                })
                journal_entries.append({
                    'line': 2,
                    'account_code': account_code,
                    'account_name': account_name,
                    'debit': 0 if gross > 0 else gross,
                    'credit': gross if gross > 0 else 0,
                    'tax_code': 'GST on Expenses' if gst > 0 else 'GST Free',  # Original (being reversed)
                    'description': std_desc
                })
                recode_done = True
                # If we recoded with the correct tax code, GST is also fixed
                if correct_tax_code != ('GST on Expenses' if gst > 0 else 'GST Free'):
                    gst_correction_done = True

    # Group ALL entertainment errors together - only ONE journal needed
    # (alcohol_gst_error, client_entertainment_gst, staff_entertainment_gst all have same correction)
    is_entertainment_error = (
        transaction.get('alcohol_gst_error') or
        transaction.get('client_entertainment_gst') or
        transaction.get('staff_entertainment_gst')
    )
    if is_entertainment_error and not gst_correction_done:
        # Entertainment - reverse original GST on Expenses coding to GST Free Expenses
        # Same account, just changing the tax code (no GST credit claimable on entertainment)
        trans_desc = transaction.get('description', '')[:50] or 'No description'
        std_desc = f"GST on Expenses to GST Free - {trans_desc}"

        if gross > 0:
            # Debit first: re-enter with GST Free Expenses (correct treatment for entertainment)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gross,
                'credit': 0,
                'tax_code': 'GST Free Expenses',
                'description': std_desc
            })
            # Credit: reverse the original GST on Expenses entry
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'GST on Expenses',
                'description': std_desc
            })
            gst_correction_done = True

    if not transaction.get('gst_calculation_correct', True) and not is_entertainment_error:
        # GST calculation error - need to adjust using same account with different tax codes
        expected_gst = round(net * 0.10, 2)
        gst_diff = round(expected_gst - gst, 2)
        trans_desc = transaction.get('description', '')[:50] or 'No description'

        if abs(gst_diff) > 0.02:
            std_desc = f"GST adjustment - {trans_desc}"
            if gst_diff > 0:
                # Under-claimed GST - need to increase GST claimed
                journal_entries.append({
                    'line': len(journal_entries) + 1,
                    'account_code': account_code,
                    'account_name': account_name,
                    'debit': gst_diff,
                    'credit': 0,
                    'tax_code': 'GST on Expenses',
                    'description': std_desc
                })
                journal_entries.append({
                    'line': len(journal_entries) + 1,
                    'account_code': account_code,
                    'account_name': account_name,
                    'debit': 0,
                    'credit': gst_diff,
                    'tax_code': 'BAS Excluded',
                    'description': std_desc
                })
            else:
                # Over-claimed GST - need to reduce GST claimed
                journal_entries.append({
                    'line': len(journal_entries) + 1,
                    'account_code': account_code,
                    'account_name': account_name,
                    'debit': abs(gst_diff),
                    'credit': 0,
                    'tax_code': 'BAS Excluded',
                    'description': std_desc
                })
                journal_entries.append({
                    'line': len(journal_entries) + 1,
                    'account_code': account_code,
                    'account_name': account_name,
                    'debit': 0,
                    'credit': abs(gst_diff),
                    'tax_code': 'GST on Expenses',
                    'description': std_desc
                })

    if transaction.get('missing_gst_error'):
        # Item coded as GST Free but should include GST (e.g., toner, stationery)
        # Correcting journal: reverse GST Free entry and re-enter with GST on Expenses
        # Same account, just changing the tax code
        original_tax_code = transaction.get('gst_rate_name', 'GST Free') or 'GST Free'
        trans_desc = transaction.get('description', '')[:50] or 'No description'
        std_desc = f"GST adjustment - {trans_desc}"

        if gross > 0:
            # Debit first: re-enter with correct GST on Expenses tax code
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gross,
                'credit': 0,
                'tax_code': 'GST on Expenses',
                'description': std_desc
            })
            # Credit: reverse the original GST Free entry
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': original_tax_code,
                'description': std_desc
            })

    # Skip input_taxed_gst_error if life_insurance_personal (whole expense moving to Drawings)
    if transaction.get('input_taxed_gst_error') and not transaction.get('life_insurance_personal'):
        # GST incorrectly claimed on input-taxed supply (e.g., bank fees, interest)
        # Per ATO: Input-taxed supplies have NO GST and you CANNOT claim GST credits
        trans_desc = transaction.get('description', '')[:50] or 'No description'
        std_desc = f"GST adjustment - {trans_desc}"

        if gst > 0:
            # Debit: Same account with Input Taxed (correct tax code for financial supplies)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gst,
                'credit': 0,
                'tax_code': 'Input Taxed',
                'description': std_desc
            })
            # Credit: Same account with GST on Expenses (reverses GST incorrectly claimed)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gst,
                'tax_code': 'GST on Expenses',
                'description': std_desc
            })

    if transaction.get('drawings_loan_error'):
        # Personal expense coded to Drawings/Loan with GST claimed
        # Drawings should be BAS Excluded - reverse the incorrect GST claim
        trans_desc = transaction.get('description', '')[:50] or 'No description'
        std_desc = f"GST adjustment - {trans_desc}"

        if gross > 0:
            # Debit first: re-enter as BAS Excluded (correct treatment for drawings)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gross,
                'credit': 0,
                'tax_code': 'BAS Excluded',
                'description': std_desc
            })
            # Credit: reverse the original GST on Expenses entry
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'GST on Expenses',
                'description': std_desc
            })

    if transaction.get('wages_gst_error') or transaction.get('allowance_gst_error'):
        # Wages/Salaries/Superannuation/Allowances - NO GST applies
        # These are NOT supplies and should be BAS Excluded (not reportable on BAS)
        trans_desc = transaction.get('description', '')[:50] or 'No description'
        std_desc = f"GST adjustment - {trans_desc}"

        # Use GST amount if available, otherwise calculate from gross (10% GST rate)
        gst_to_reverse = gst if gst > 0 else round(gross * 10 / 110, 2)

        if gst_to_reverse > 0:
            # Debit: Same account with BAS Excluded (correct - wages/super not reportable)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gst_to_reverse,
                'credit': 0,
                'tax_code': 'BAS Excluded',
                'description': std_desc
            })
            # Credit: Reverse the incorrect GST on Expenses entry
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gst_to_reverse,
                'tax_code': 'GST on Expenses',
                'description': std_desc
            })

    # Skip interest_gst_error journal if input_taxed_gst_error already handled it (avoid duplicates)
    if transaction.get('interest_gst_error') and not transaction.get('input_taxed_gst_error'):
        # Interest incorrectly coded - should be Input Taxed (no GST credit)
        original_tax_code = transaction.get('gst_rate_name', '') or 'Unknown'
        trans_desc = transaction.get('description', '')[:50] or 'No description'
        std_desc = f"GST adjustment - {trans_desc}"

        if gst > 0:
            # GST was claimed - need to reverse it using Input Taxed
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gst,
                'credit': 0,
                'tax_code': 'Input Taxed',
                'description': std_desc
            })
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gst,
                'tax_code': 'GST on Expenses',
                'description': std_desc
            })
        elif gross > 0:
            # BAS Excluded or wrong coding - reverse and re-enter as GST Free Income
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gross,
                'credit': 0,
                'tax_code': original_tax_code,
                'description': std_desc
            })
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'GST Free Income',
                'description': std_desc
            })

    if transaction.get('sales_gst_error'):
        # Sales incorrectly coded - should be GST on Income or valid GST Free Income
        original_tax_code = transaction.get('gst_rate_name', '') or 'Unknown'
        trans_desc = transaction.get('description', '')[:50] or 'No description'
        std_desc = f"GST adjustment - {trans_desc}"

        if gross > 0:
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gross,
                'credit': 0,
                'tax_code': original_tax_code,
                'description': std_desc
            })
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'GST on Income',
                'description': std_desc
            })

    if transaction.get('other_income_error'):
        # Other Income incorrectly coded as BAS Excluded
        original_tax_code = transaction.get('gst_rate_name', '') or 'BAS Excluded'
        trans_desc = transaction.get('description', '')[:50] or 'No description'
        std_desc = f"GST adjustment - {trans_desc}"

        if gross > 0:
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gross,
                'credit': 0,
                'tax_code': original_tax_code,
                'description': std_desc
            })
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'GST on Income',
                'description': std_desc
            })

    if transaction.get('export_gst_error'):
        # Export sale incorrectly charged GST - exports should be GST-FREE
        # Per ATO: Exports are GST-free (no GST charged, but CAN claim input credits)
        original_tax_code = transaction.get('gst_rate_name', '') or 'GST on Income'
        trans_desc = transaction.get('description', '')[:50] or 'No description'
        std_desc = f"GST adjustment - {trans_desc}"

        if gross > 0:
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gross,
                'credit': 0,
                'tax_code': original_tax_code,
                'description': std_desc
            })
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'GST Free Income',
                'description': std_desc
            })

    if transaction.get('grants_sponsorship_gst') == 'grant_with_gst':
        # Grant income incorrectly charged GST - grants typically GST-FREE per GSTR 2012/2
        # Unless there's a binding obligation to provide specific services/goods in return
        original_tax_code = transaction.get('gst_rate_name', '') or 'GST on Income'
        trans_desc = transaction.get('description', '')[:50] or 'No description'
        std_desc = f"GST adjustment (grant) - {trans_desc}"

        if gross > 0:
            # Reverse original GST on Income entry and re-enter as GST Free Income
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gross,
                'credit': 0,
                'tax_code': original_tax_code,
                'description': std_desc
            })
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'GST Free Income',
                'description': std_desc
            })

    if transaction.get('motor_vehicle_gst_limit'):
        # Motor vehicle GST exceeds ATO car limit - need to reverse excess GST claimed
        car_limit = 69674
        amount_over_limit = gross - car_limit if gross > car_limit else 0
        trans_desc = transaction.get('description', '')[:50] or 'No description'
        std_desc = f"GST adjustment - {trans_desc}"

        if amount_over_limit > 0:
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': amount_over_limit,
                'credit': 0,
                'tax_code': 'BAS Excluded',
                'description': std_desc
            })
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': amount_over_limit,
                'tax_code': 'GST on Expenses',
                'description': std_desc
            })

    # REMOVED: overseas_subscription_gst check
    # Both GST and GST-Free are valid for overseas digital services (Adobe, Slack, Zoom, etc.)
    # - With GST: Netflix tax or reverse charge applied = valid
    # - Without GST: GST Free = valid

    # Group government charges and fines/penalties together (fines are a type of government charge)
    # Only generate if not already corrected
    is_gov_charge_or_fine = (
        transaction.get('government_charges_gst') or
        transaction.get('fines_penalties_gst')
    )
    if is_gov_charge_or_fine and not gst_correction_done:
        # Government charges/fines - NO GST applies
        # Fines use BAS Excluded, other gov charges use GST Free
        trans_desc = transaction.get('description', '')[:50] or 'No description'

        if transaction.get('fines_penalties_gst'):
            # Fines/penalties are BAS Excluded (non-reportable)
            target_tax_code = 'BAS Excluded'
            std_desc = f"GST adjustment (fine/penalty) - {trans_desc}"
            # For fines, only adjust the GST component
            if gst > 0:
                journal_entries.append({
                    'line': len(journal_entries) + 1,
                    'account_code': account_code,
                    'account_name': account_name,
                    'debit': gst,
                    'credit': 0,
                    'tax_code': 'BAS Excluded',
                    'description': std_desc
                })
                journal_entries.append({
                    'line': len(journal_entries) + 1,
                    'account_code': account_code,
                    'account_name': account_name,
                    'debit': 0,
                    'credit': gst,
                    'tax_code': 'GST on Expenses',
                    'description': std_desc
                })
                gst_correction_done = True
        else:
            # Other government charges (stamp duty, rates, etc.) are GST Free
            std_desc = f"GST adjustment (gov charge) - {trans_desc}"
            if gross > 0:
                journal_entries.append({
                    'line': len(journal_entries) + 1,
                    'account_code': account_code,
                    'account_name': account_name,
                    'debit': gross,
                    'credit': 0,
                    'tax_code': 'GST Free Expenses',
                    'description': std_desc
                })
                journal_entries.append({
                    'line': len(journal_entries) + 1,
                    'account_code': account_code,
                    'account_name': account_name,
                    'debit': 0,
                    'credit': gross,
                    'tax_code': 'GST on Expenses',
                    'description': std_desc
                })
                gst_correction_done = True

    # NOTE: client_entertainment_gst and staff_entertainment_gst are now handled
    # in the combined is_entertainment_error block above (around line 3295)

    if transaction.get('residential_premises_gst') and not gst_correction_done:
        # Residential property expense - GST not claimable (input-taxed supply)
        trans_desc = transaction.get('description', '')[:50] or 'No description'
        std_desc = f"GST adjustment - {trans_desc}"
        if gst > 0:
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gst,
                'credit': 0,
                'tax_code': 'Input Taxed',
                'description': std_desc
            })
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gst,
                'tax_code': 'GST on Expenses',
                'description': std_desc
            })
            gst_correction_done = True

    if transaction.get('insurance_gst_error') and not gst_correction_done and not transaction.get('life_insurance_personal'):
        # Life/income protection insurance - GST not claimable (input-taxed)
        # Skip if life_insurance_personal is flagged (whole expense being moved to Owner Drawings)
        trans_desc = transaction.get('description', '')[:50] or 'No description'
        std_desc = f"GST adjustment - {trans_desc}"
        if gst > 0:
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gst,
                'credit': 0,
                'tax_code': 'Input Taxed',
                'description': std_desc
            })
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gst,
                'tax_code': 'GST on Expenses',
                'description': std_desc
            })
            gst_correction_done = True

    if transaction.get('life_insurance_personal'):
        # Life/income protection insurance - NOT a business expense
        # Should be recoded to Owner Drawings (owner can claim on personal tax return)
        trans_desc = transaction.get('description', '')[:50] or 'No description'
        std_desc = f"Recode to Owner Drawings - {trans_desc}"
        if gross > 0:
            # Debit: Owner A Drawings (880) with BAS Excluded
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': '880',
                'account_name': 'Owner A Drawings',
                'debit': gross,
                'credit': 0,
                'tax_code': 'BAS Excluded',
                'description': std_desc
            })
            # Credit: Reverse the original expense account
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'GST on Expenses' if gst > 0 else 'GST Free Expenses',
                'description': std_desc
            })

    if transaction.get('allowance_gst_error') and not gst_correction_done:
        # Employee allowance - GST not claimable (not a purchase from supplier)
        trans_desc = transaction.get('description', '')[:50] or 'No description'
        std_desc = f"GST adjustment - {trans_desc}"
        if gst > 0:
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gst,
                'credit': 0,
                'tax_code': 'BAS Excluded',
                'description': std_desc
            })
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gst,
                'tax_code': 'GST on Expenses',
                'description': std_desc
            })
            gst_correction_done = True

    # NOTE: fines_penalties_gst is now handled in the combined is_gov_charge_or_fine block above

    if transaction.get('donations_gst') and not gst_correction_done:
        # Donations - NO GST applies
        # For P&L expense accounts, use GST Free Expenses (not BAS Excluded)
        # GST Free Expenses is correct for donations as they are GST-free, not out of scope
        if gst > 0:
            # Debit: Same account with GST Free Expenses (donations have no GST)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gst,
                'credit': 0,
                'tax_code': 'GST Free Expenses',
                'description': f"Adjust donation - no GST applies"
            })
            # Credit: Same account with GST on Expenses (reverse GST claimed via tax code)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gst,
                'tax_code': 'GST on Expenses',
                'description': f"Reverse GST claimed on donation"
            })
            gst_correction_done = True

    if transaction.get('voucher_gst') == 'face_value_with_gst' and not gst_correction_done:
        # Face value voucher sale with GST - should be NO GST at sale
        # GST only applies when voucher is redeemed
        # Reverse the GST charged at sale
        if gst > 0:
            # For face value vouchers sold, we need to reverse the GST charged
            # Debit: reverse GST on Income, Credit: re-enter as BAS Excluded
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gross,
                'credit': 0,
                'tax_code': 'GST on Income',
                'description': f"Reverse GST charged on face value voucher sale"
            })
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'BAS Excluded',
                'description': f"Re-enter voucher sale - no GST at time of sale"
            })
            gst_correction_done = True

    if transaction.get('travel_gst') == 'international_with_gst' and not gst_correction_done:
        # International travel with GST claimed - should be GST-free
        # Reverse the GST claimed on international travel expenses
        if gst > 0:
            # Debit: Same account with GST Free Expenses (correct for international)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gross,
                'credit': 0,
                'tax_code': 'GST Free Expenses',
                'description': f"Re-enter as GST Free - international travel"
            })
            # Credit: Same account with GST on Expenses (reverse incorrect GST claim)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'GST on Expenses',
                'description': f"Reverse GST claimed on international travel"
            })
            gst_correction_done = True

    if transaction.get('travel_gst') == 'domestic_no_gst' and not gst_correction_done:
        # Domestic travel without GST - should be taxable
        # Add GST to domestic travel expenses
        if gross > 0:
            # Debit: Same account with GST on Expenses (correct for domestic)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gross,
                'credit': 0,
                'tax_code': 'GST on Expenses',
                'description': f"Re-enter with GST - domestic travel is taxable"
            })
            # Credit: Same account with GST Free Expenses (reverse incorrect coding)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'GST Free Expenses',
                'description': f"Reverse GST Free coding on domestic travel"
            })

    if transaction.get('payment_processor_fees') == 'paypal_with_gst':
        # PayPal fees with GST claimed - should be Input Taxed (no GST)
        # PayPal fees are GST exempt financial supplies
        if gst > 0:
            # Debit: Same account with Input Taxed (PayPal is financial supply)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gross,
                'credit': 0,
                'tax_code': 'Input Taxed',
                'description': f"Re-enter as Input Taxed - PayPal fees are GST exempt"
            })
            # Credit: Same account with GST on Expenses (reverse incorrect GST claim)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'GST on Expenses',
                'description': f"Reverse GST claimed on PayPal fees"
            })

    if transaction.get('payment_processor_fees') in ['stripe_no_gst', 'merchant_no_gst']:
        # Stripe/Bank merchant fees without GST - should include GST
        if gross > 0:
            # Debit: Same account with GST on Expenses (correct - includes GST)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gross,
                'credit': 0,
                'tax_code': 'GST on Expenses',
                'description': f"Re-enter with GST - Stripe/bank fees include GST"
            })
            # Credit: Same account with GST Free (reverse incorrect coding)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'GST Free Expenses',
                'description': f"Reverse GST Free coding on merchant fees"
            })

    if transaction.get('livestock_gst'):
        # Livestock sale incorrectly coded as GST-free - should be taxable
        # Live animals are subject to GST (meat only becomes GST-free after inspection)
        original_tax_code = transaction.get('gst_rate_name', '') or 'GST Free Income'

        if gross > 0:
            # For INCOME accounts: Debit reverses (reduces income), Credit re-enters correctly
            # Debit first: reverse original GST-free entry
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gross,
                'credit': 0,
                'tax_code': original_tax_code,
                'description': f"Reverse original entry coded as {original_tax_code}"
            })
            # Credit: re-enter with GST on Income (livestock sales are taxable)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'GST on Income',
                'description': f"Re-enter with GST on Income - livestock sales are taxable"
            })

    if transaction.get('asset_disposal_gst'):
        # Asset disposal incorrectly coded as BAS Excluded - should be taxable
        # Business asset sales must include GST (report at G1, remit GST at 1A)
        original_tax_code = transaction.get('gst_rate_name', '') or 'BAS Excluded'

        if gross > 0:
            # For INCOME accounts: Debit reverses (reduces income), Credit re-enters correctly
            # Debit first: reverse original BAS Excluded entry
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': gross,
                'credit': 0,
                'tax_code': original_tax_code,
                'description': f"Reverse original entry coded as {original_tax_code}"
            })
            # Credit: re-enter with GST on Income (asset sales are taxable)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'GST on Income',
                'description': f"Re-enter with GST on Income - asset disposal is taxable"
            })

    if transaction.get('asset_capitalization_error'):
        # Asset should be capitalized instead of expensed (over $20,000 threshold)
        # This is an account coding error - need to move from expense to asset account
        suggested_asset = suggest_asset_account(description)

        if suggested_asset and gross > 0:
            # Debit: Asset account (capitalize the expense)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': suggested_asset['code'],
                'account_name': suggested_asset['name'],
                'debit': gross,
                'credit': 0,
                'tax_code': 'GST on Capital' if gst > 0 else 'BAS Excluded',
                'description': f"Capitalize asset - over $20,000 threshold"
            })
            # Credit: Original expense account (reverse the expense)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'GST on Expenses' if gst > 0 else 'GST Free',
                'description': f"Reverse expense - should be capitalized as asset"
            })

    # Skip computer equipment capitalization if it's a PERSONAL item (personal takes priority)
    # Personal items go to Owner Drawings, not Computer Equipment asset
    is_personal_item = transaction.get('personal_in_business_account')

    if transaction.get('computer_equipment_expense') and not recode_done and not is_personal_item:
        # Computer equipment coded to expense account should be capitalized as asset
        # Even with instant asset write-off, should go through asset account for tracking
        # ATO effective life: Laptops 2 years, Computers 4 years
        if gross > 0:
            # Debit: Computer Equipment - At Cost (asset account)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': '160',
                'account_name': 'Computer Equipment - At Cost',
                'debit': gross,
                'credit': 0,
                'tax_code': 'GST on Capital' if gst > 0 else 'BAS Excluded',
                'description': f"Capitalize computer equipment - ATO depreciation applies"
            })
            # Credit: Original expense account (reverse the expense)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'GST on Expenses' if gst > 0 else 'GST Free',
                'description': f"Reverse expense - should be capitalized as computer equipment"
            })
            recode_done = True

    if transaction.get('borrowing_expenses_error') and not recode_done:
        # Borrowing expenses > $100 should be capitalized and spread over 5 years
        # This is an account coding error - move from expense to prepaid/deferred asset
        if gross > 0:
            # Debit: Prepaid Borrowing Costs (asset account)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': '180',
                'account_name': 'Prepaid Borrowing Costs',
                'debit': gross,
                'credit': 0,
                'tax_code': 'BAS Excluded',
                'description': f"Capitalize borrowing expenses - spread over 5 years per ATO s25.25"
            })
            # Credit: Original expense account (reverse the expense)
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': 'GST Free',
                'description': f"Reverse expense - borrowing costs > $100 must be capitalized"
            })
            recode_done = True

    if transaction.get('personal_in_business_account') and not recode_done:
        # Personal expense incorrectly coded to business expense account
        # Need to move from business account (e.g., Telephone) to Owner Drawings
        # AND reverse any GST claimed (personal expenses can't claim GST)
        trans_desc = transaction.get('description', '')[:50] or 'No description'
        std_desc = f"Recode personal expense - {trans_desc}"
        original_tax_code = transaction.get('gst_rate_name', 'GST on Expenses') or 'GST on Expenses'

        if gross > 0:
            # Debit: Owner A Drawings (880) with BAS Excluded
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': '880',
                'account_name': 'Owner A Drawings',
                'debit': gross,
                'credit': 0,
                'tax_code': 'BAS Excluded',
                'description': std_desc
            })
            # Credit: Original business expense account to reverse it
            journal_entries.append({
                'line': len(journal_entries) + 1,
                'account_code': account_code,
                'account_name': account_name,
                'debit': 0,
                'credit': gross,
                'tax_code': original_tax_code,
                'description': std_desc
            })
            recode_done = True

    return {
        'narration': narration,
        'entries': journal_entries
    }


def suggest_asset_account(description):
    """Suggest the correct asset account based on transaction description for capitalization"""
    description = description.lower()

    # Asset account mappings based on keywords
    asset_mappings = [
        # Motor Vehicles
        (['car ', 'vehicle', 'motor vehicle', 'toyota', 'ford', 'holden', 'mazda', 'hyundai',
          'kia', 'nissan', 'honda', 'bmw', 'mercedes', 'audi', 'volkswagen', 'subaru',
          'mitsubishi', 'lexus', 'tesla', 'ute ', 'suv ', 'sedan', 'hatchback', 'wagon'],
         {'code': '150', 'name': 'Motor Vehicles - At Cost'}),

        # Computer Equipment
        (['laptop', 'computer', 'macbook', 'imac', 'pc ', 'desktop', 'server', 'monitor',
          'apple mac', 'dell ', 'hp ', 'lenovo', 'surface'],
         {'code': '160', 'name': 'Computer Equipment - At Cost'}),

        # Office Equipment & Furniture
        (['furniture', 'desk', 'chair', 'cabinet', 'shelving', 'table', 'workstation',
          'reception', 'lounge', 'sofa', 'couch'],
         {'code': '161', 'name': 'Office Equipment - At Cost'}),

        # Plant & Equipment / Machinery
        (['machinery', 'equipment', 'plant', 'forklift', 'crane', 'excavator', 'tools',
          'industrial', 'manufacturing', 'production'],
         {'code': '162', 'name': 'Plant & Equipment - At Cost'}),

        # Leasehold Improvements
        (['fit out', 'fitout', 'renovation', 'refurbishment', 'improvements', 'shopfit',
          'partition', 'flooring', 'ceiling', 'lighting installation'],
         {'code': '170', 'name': 'Leasehold Improvements - At Cost'}),

        # Printers & Office Machines
        (['printer', 'scanner', 'copier', 'photocopier', 'fax', 'shredder', 'projector'],
         {'code': '161', 'name': 'Office Equipment - At Cost'}),

        # HVAC & Building
        (['air conditioner', 'hvac', 'heating', 'cooling', 'split system', 'ducted'],
         {'code': '170', 'name': 'Leasehold Improvements - At Cost'}),

        # Signage
        (['signage', 'sign', 'billboard', 'display'],
         {'code': '170', 'name': 'Leasehold Improvements - At Cost'}),

        # Phones & Communication
        (['phone system', 'pbx', 'telephone system', 'voip'],
         {'code': '161', 'name': 'Office Equipment - At Cost'}),
    ]

    for keywords, asset_account in asset_mappings:
        if any(keyword in description for keyword in keywords):
            return asset_account

    # Default to Plant & Equipment if no specific match
    return {'code': '162', 'name': 'Plant & Equipment - At Cost'}


def suggest_correct_account(description):
    """Suggest the correct account based on transaction description"""
    description = description.lower()

    # Mapping of keywords to suggested accounts
    account_mappings = [
        # Travel
        (['flight', 'qantas', 'virgin', 'jetstar', 'airline', 'airfare'], {'code': '420', 'name': 'Travel - National'}),
        (['hotel', 'accommodation', 'motel'], {'code': '420', 'name': 'Travel - National'}),
        (['taxi', 'uber', 'didi', 'ola'], {'code': '420', 'name': 'Travel - National'}),
        (['parking', 'car park', 'wilson'], {'code': '449', 'name': 'Motor Vehicle Expenses'}),

        # Office & Admin
        (['stationery', 'officeworks', 'office supplies'], {'code': '453', 'name': 'Printing & Stationery'}),
        (['toner', 'ink', 'cartridge', 'printer'], {'code': '453', 'name': 'Printing & Stationery'}),
        (['postage', 'stamps', 'auspost'], {'code': '458', 'name': 'Postage'}),

        # Meals & Entertainment
        (['restaurant', 'cafe', 'lunch', 'dinner', 'catering'], {'code': '424', 'name': 'Entertainment'}),
        (['dan murphy', 'bws', 'liquorland', 'wine', 'alcohol'], {'code': '424', 'name': 'Entertainment'}),

        # Professional Services
        (['legal', 'lawyer', 'solicitor'], {'code': '440', 'name': 'Legal Expenses'}),
        (['accounting', 'accountant', 'bookkeep'], {'code': '404', 'name': 'Accounting & Audit'}),
        (['consulting', 'consultant'], {'code': '460', 'name': 'Consulting & Professional Fees'}),

        # Subscriptions & Software
        (['software', 'subscription', 'license', 'saas'], {'code': '463', 'name': 'Subscriptions'}),
        (['microsoft', 'adobe', 'xero', 'myob'], {'code': '463', 'name': 'Subscriptions'}),

        # Utilities & Communications
        (['phone', 'mobile', 'telstra', 'optus', 'vodafone'], {'code': '469', 'name': 'Telephone & Internet'}),
        (['internet', 'nbn', 'broadband'], {'code': '469', 'name': 'Telephone & Internet'}),
        (['electricity', 'gas', 'power', 'energy'], {'code': '445', 'name': 'Light, Power, Heating'}),

        # Bank & Finance
        (['bank fee', 'account fee', 'transaction fee'], {'code': '404', 'name': 'Bank Fees'}),
        (['merchant fee', 'credit card fee', 'eftpos'], {'code': '404', 'name': 'Bank Fees'}),

        # Insurance
        (['insurance', 'premium'], {'code': '441', 'name': 'Insurance'}),

        # Repairs & Maintenance
        (['repairs', 'maintenance', 'fix'], {'code': '461', 'name': 'Repairs & Maintenance'}),

        # Freight & Courier
        (['freight', 'courier', 'delivery', 'shipping'], {'code': '425', 'name': 'Freight & Courier'}),
    ]

    for keywords, account in account_mappings:
        if any(keyword in description for keyword in keywords):
            return account

    # Default expense account if no match
    return {'code': '400', 'name': 'General Expenses'}


def check_account_coding(transaction):
    """Check if account coding seems suspicious"""
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    account_code = str(transaction.get('account_code', ''))
    account_type = transaction.get('account_type', '').upper()
    transaction_type = transaction.get('type', '').lower()

    # Get business context to avoid false positives for industry-specific expenses
    business_context = get_business_context()
    industry = business_context.get('industry', '')

    # For automotive businesses, car-related expenses are legitimate stock/COGS
    # Don't flag parts from Repco, Supercheap, etc. as needing motor vehicle account
    if industry == 'automotive':
        automotive_expense_keywords = [
            'repco', 'supercheap', 'burson', 'autobarn', 'bursons', 'car parts', 'auto parts',
            'spare parts', 'oil', 'filter', 'brake', 'tyre', 'tire', 'battery', 'engine',
            'transmission', 'exhaust', 'suspension', 'radiator', 'windscreen', 'wiper',
            'alternator', 'starter motor', 'panel', 'spray paint', 'car wash', 'detailing'
        ]
        if any(keyword in description for keyword in automotive_expense_keywords):
            return False  # Legitimate automotive business expense

    # Check if this matches a known allocation pattern from Deep Scan
    # e.g., Telstra split 70/30 between Telephone and Drawings
    if is_known_allocation_pattern(transaction):
        return False  # This is a known allocation pattern - don't flag

    # Check for expenses coded to Sales/Revenue accounts
    is_revenue_account = (
        'sales' in account or
        'revenue' in account or
        account_type == 'REVENUE' or
        account_code in ['200', '201', '202', '210', '215', '260', '270']
    )

    if is_revenue_account:
        # Don't flag legitimate service revenue
        service_revenue_keywords = ['fixed fee', 'monthly fee', 'service fee', 'consulting fee',
                                    'professional fee', 'retainer', 'project fee', 'hourly rate',
                                    'rate as agreed', 'as agreed', 'as per agreement', 'per hour',
                                    'management fee', 'advisory fee', 'invoice', 'services rendered',
                                    'billable hours', 'professional services', 'fee for service',
                                    # Training/education as a service
                                    'training', 'workshop', 'course', 'session', 'coaching',
                                    'microsoft', 'ms office', 'excel', 'word', 'powerpoint',
                                    'instruction', 'tutoring', 'lesson', 'seminar', 'webinar']
        is_service_revenue = any(keyword in description for keyword in service_revenue_keywords)
        if is_service_revenue:
            return False  # This is legitimate service income

        # Also check if this matches expected income for the business type
        likely_income_sources = business_context.get('likely_income_sources', [])
        is_expected_income = any(source in description for source in likely_income_sources)
        if is_expected_income:
            return False  # Matches expected income for this business type

        is_refund = any(word in description for word in ['refund', 'credit note', 'reversal', 'cancelled', 'returned'])
        expense_keywords = [
            'flight', 'qantas', 'virgin', 'jetstar', 'airline', 'airfare',
            'hotel', 'accommodation', 'parking', 'taxi', 'uber', 'car park',
            'office', 'stationery', 'supplies', 'toner', 'printer',
            'software', 'license',
            'meal', 'lunch', 'dinner', 'restaurant', 'cafe', 'catering',
            'insurance', 'premium', 'rent', 'lease',
            'phone', 'mobile', 'internet', 'electricity', 'utilities',
            'bank fee', 'merchant fee', 'freight', 'courier', 'postage',
            'repairs', 'maintenance', 'conference',
            'legal', 'accounting'
        ]
        if any(keyword in description for keyword in expense_keywords) and not is_refund:
            return True

    # Check for alcohol in wrong accounts - should ONLY be in Entertainment or Gift accounts
    alcohol_keywords = ['dan murphy', 'bws', 'liquorland', 'wine', 'beer', 'spirits', 'alcohol', 'champagne', 'liquor',
                        'vintage cellar', 'cellarbrations', 'bottlemart', 'liquor barn', 'first choice']
    valid_alcohol_accounts = ['entertainment', 'gift', 'staff amenities']

    if any(keyword in description for keyword in alcohol_keywords):
        if not any(valid_acct in account for valid_acct in valid_alcohol_accounts):
            return True

    # Check for parking expenses coded to wrong accounts (e.g., Legal)
    parking_keywords = ['parking', 'car park', 'wilson parking', 'secure parking', 'care park']
    valid_parking_accounts = ['motor vehicle', 'travel', 'parking', 'transport', 'vehicle']
    if any(keyword in description for keyword in parking_keywords):
        if not any(valid_acct in account for valid_acct in valid_parking_accounts):
            return True

    # Check for software subscriptions coded to wrong accounts (e.g., Consulting)
    subscription_keywords = ['xero', 'myob', 'quickbooks', 'microsoft 365', 'office 365', 'adobe',
                            'dropbox', 'google workspace', 'slack', 'zoom', 'canva', 'mailchimp',
                            'hubspot', 'salesforce', 'asana', 'trello', 'notion', 'figma',
                            'github', 'atlassian', 'jira', 'confluence', 'shopify', 'squarespace',
                            'wix', 'wordpress', 'aws', 'azure', 'gcp', 'heroku', 'netlify',
                            'subscription', 'monthly fee', 'annual fee', 'saas', 'software']
    valid_subscription_accounts = ['subscription', 'software', 'computer', 'it expense', 'cloud', 'license']
    wrong_subscription_accounts = ['consulting', 'professional', 'legal', 'accounting']

    if any(keyword in description for keyword in subscription_keywords):
        # Flag if coded to consulting/professional/legal instead of subscriptions
        if any(wrong_acct in account for wrong_acct in wrong_subscription_accounts):
            return True
        # Also flag if not in any valid subscription account
        if not any(valid_acct in account for valid_acct in valid_subscription_accounts):
            # Only flag if it's clearly a subscription service
            clear_subscription = any(kw in description for kw in ['xero', 'myob', 'quickbooks', 'microsoft 365',
                                                                   'office 365', 'adobe', 'dropbox', 'slack', 'zoom'])
            if clear_subscription:
                return True

    # Check for office supplies/stationery coded to wrong accounts
    stationery_keywords = ['toner', 'cartridge', 'ink', 'paper', 'stationery', 'officeworks',
                          'staples', 'pens', 'folders', 'envelopes', 'printer supplies']
    valid_stationery_accounts = ['stationery', 'printing', 'office', 'supplies', 'consumables']
    wrong_stationery_accounts = ['legal', 'consulting', 'professional', 'travel', 'entertainment']

    if any(keyword in description for keyword in stationery_keywords):
        if any(wrong_acct in account for wrong_acct in wrong_stationery_accounts):
            return True

    return False


def check_alcohol_gst(transaction):
    """
    Check if alcohol/entertainment purchases have incorrect GST coding.
    Entertainment alcohol should be GST Free Expenses.
    BWS, Dan Murphy, wine etc. are ONLY acceptable in Entertainment or Gift accounts.
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_amount = abs(transaction.get('gst', 0))
    gst_rate_name = transaction.get('gst_rate_name', '').lower()

    # Alcohol retailer keywords
    alcohol_retailers = ['dan murphy', 'bws', 'liquorland', 'first choice', 'aldi liquor',
                        'vintage cellar', 'cellarbrations', 'bottlemart', 'liquor barn']

    # Alcohol product keywords
    alcohol_products = ['wine', 'beer', 'spirits', 'alcohol', 'champagne', 'liquor',
                       'vodka', 'whisky', 'whiskey', 'gin', 'rum', 'tequila', 'brandy',
                       'prosecco', 'sparkling wine', 'cider', 'ale', 'lager']

    is_alcohol_purchase = (
        any(keyword in description for keyword in alcohol_retailers) or
        any(keyword in description for keyword in alcohol_products)
    )

    # Valid accounts for alcohol purchases (Entertainment or Gift only)
    is_valid_alcohol_account = (
        'entertainment' in account or
        'gift' in account or
        'staff amenities' in account
    )

    if is_alcohol_purchase:
        # If in valid account (Entertainment/Gift), check if GST is incorrectly claimed
        # Entertainment should be GST Free Expenses, not GST on Expenses
        if is_valid_alcohol_account:
            has_gst_claimed = gst_amount > 0 or ('gst on' in gst_rate_name and 'free' not in gst_rate_name)
            if has_gst_claimed:
                return True  # Flag - should be GST Free Expenses
        # If NOT in valid account, this is handled by check_account_coding
        # Don't double-flag here

    return False


def check_input_taxed_gst(transaction):
    """
    Check if GST is incorrectly claimed on input-taxed supplies.
    Per ATO rules: Input-taxed supplies (financial supplies) have no GST
    and you CANNOT claim GST credits.

    Financial supplies include:
    - Lending money or provision of credit
    - Bank account fees, account keeping fees
    - Interest (paid or received)
    - Share transfers, securities trading
    - Foreign currency transactions
    - Loan fees, loan establishment
    - Life insurance, income protection

    Note: Credit card fees/merchant fees/EFTPOS fees to BUSINESSES are TAXABLE (have GST)
          Only fees FROM banks for operating accounts are input-taxed

    Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/when-to-charge-gst-and-when-not-to/input-taxed-sales/financial-supplies
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_amount = abs(transaction.get('gst', 0))
    gst_rate_name = transaction.get('gst_rate_name', '').lower()

    # Skip wages/salary/super accounts - these are handled by check_wages_gst_error
    # Wages/super are BAS Excluded, not Input Taxed (different treatment)
    wage_account_keywords = ['wage', 'salary', 'payroll', 'superannuation', 'super ']
    is_wage_account = any(keyword in account for keyword in wage_account_keywords)
    if is_wage_account:
        return False

    # Skip wage/super descriptions - handled by check_wages_gst_error
    wage_description_keywords = ['wage', 'salary', 'payroll', 'super contribution', 'super guarantee', 'sgc', 'superannuation contribution']
    is_wage_description = any(keyword in description for keyword in wage_description_keywords)
    if is_wage_description:
        return False

    # Credit card fees/merchant fees to BUSINESSES CAN have GST - exclude from input-taxed check
    # These are fees charged by payment processors (Square, Stripe, etc.) - TAXABLE
    # Per ATO: Merchant fees include GST and businesses can claim input tax credits
    # Source: https://www.ato.gov.au/law/view/view.htm?docid=GIR/Financial-services-ch3
    cc_fee_keywords = ['cc fee', 'credit card fee', 'surcharge', 'credit card surcharge',
                       'card fee', 'merchant fee', 'merchant facility', 'merchant service',
                       'eftpos fee', 'eftpos charge', 'payment processing',
                       'square fee', 'stripe fee', 'paypal fee', 'afterpay fee']
    is_cc_fee = any(keyword in description for keyword in cc_fee_keywords)

    # Also check for "merchant" in description (catches "merchant fees", "merchant facility fees", etc.)
    if 'merchant' in description and ('fee' in description or 'charge' in description):
        is_cc_fee = True

    # If it's a credit card/merchant fee, it's NOT input-taxed (GST applies)
    if is_cc_fee:
        return False

    # Input-taxed financial supplies - these should NOT have GST
    # Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/when-to-charge-gst-and-when-not-to/input-taxed-sales/financial-supplies
    financial_supply_keywords = [
        # Bank fees and charges (from operating bank accounts)
        'bank fee', 'bank charge', 'account keeping fee', 'monthly account fee',
        'overdraft fee', 'dishonour fee', 'overdrawn fee', 'maintenance fee',
        'bank maintenance', 'account maintenance', 'transaction fee bank',
        # Interest (input-taxed - both paid and received)
        'interest charge', 'interest expense', 'loan interest', 'interest income',
        'interest received', 'term deposit interest',
        # Loan fees (input-taxed)
        'loan fee', 'loan establishment', 'facility fee', 'line fee',
        'commitment fee', 'drawdown fee',
        # Investment/securities related (input-taxed)
        'brokerage fee', 'brokerage', 'share trading', 'share purchase', 'share sale',
        'investment management fee', 'fund management fee', 'custody fee',
        'securities transfer', 'stock transfer',
        # Foreign currency (input-taxed)
        'forex fee', 'foreign exchange', 'currency conversion', 'fx fee',
        # Life insurance & super (input-taxed)
        'life insurance premium', 'income protection premium',
        'superannuation admin', 'super admin fee', 'super contribution',
        'tpd insurance', 'trauma insurance',
        # Dividends/distributions (input-taxed)
        'dividend', 'distribution', 'franking credit',
    ]

    # Check if it's an input-taxed supply
    is_input_taxed = any(keyword in description for keyword in financial_supply_keywords)

    # Flag if GST is claimed on an input-taxed supply (this is incorrect)
    if is_input_taxed and gst_amount > 0:
        return True

    return False


def check_missing_gst(transaction):
    """
    Check if transaction should have GST but is coded as GST Free.

    Per ATO rules, these are TAXABLE supplies (10% GST applies):
    - Most goods and services sold by GST-registered businesses
    - Business assets (office equipment, motor vehicles)
    - Processed foods (biscuits, soft drinks, restaurant meals, takeaway)
    - New residential and commercial properties
    - Real estate agent fees
    - Digital products (software, eBooks)
    - Professional services (accounting, legal, consulting)
    - Utilities (electricity, gas, phone, internet)
    - Commercial rent

    Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/when-to-charge-gst-and-when-not-to/taxable-sales
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_amount = abs(transaction.get('gst', 0))
    gross_amount = abs(transaction.get('gross', 0))
    gst_rate_name = transaction.get('gst_rate_name', '').lower()

    # Skip if already has GST
    if gst_amount > 0:
        return False

    # Skip Drawings/Loan accounts - these are personal expenses and should be BAS Excluded
    # Personal expenses cannot claim GST credits, so no GST is correct
    drawings_loan_keywords = ['drawing', 'drawings', 'loan', 'director loan', 'shareholder loan', 'private']
    is_drawings_loan = any(keyword in account for keyword in drawings_loan_keywords)
    if is_drawings_loan:
        return False

    # Skip personal expenses - these should be BAS Excluded with no GST
    personal_keywords = ['personal', 'private', 'private use', 'personal use']
    is_personal = any(keyword in description for keyword in personal_keywords)
    if is_personal:
        return False

    # Skip entertainment expenses - GST credits CANNOT be claimed on entertainment
    # Per ATO: Entertainment is non-deductible and GST credits are blocked
    # So entertainment coded as GST-Free is CORRECT, not an error
    entertainment_accounts = ['entertainment', 'staff amenities', 'gift']
    entertainment_keywords = ['dan murphy', 'bws', 'liquorland', 'wine', 'beer', 'alcohol',
                              'client dinner', 'client lunch', 'staff party', 'christmas party',
                              'team building', 'celebration', 'farewell']
    is_entertainment = (
        any(keyword in account for keyword in entertainment_accounts) or
        any(keyword in description for keyword in entertainment_keywords)
    )
    if is_entertainment:
        return False

    # Skip if explicitly marked as GST taxable (probably just zero GST amount)
    if 'gst on' in gst_rate_name and 'free' not in gst_rate_name:
        return False

    # Skip OVERSEAS DIGITAL SERVICES - GST Free is CORRECT for these
    # Per ATO: When overseas suppliers don't charge GST, code as GST Free
    # Reverse charge may apply but results in net zero for most businesses
    # Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/rules-for-specific-transactions/international-transactions/reverse-charge-gst-on-offshore-goods-and-services-purchases
    overseas_saas_keywords = [
        # Major US SaaS providers
        'adobe', 'creative cloud', 'photoshop', 'illustrator', 'acrobat',
        'slack', 'zoom', 'dropbox', 'canva', 'grammarly', 'notion',
        'asana', 'trello', 'monday.com', 'basecamp', 'clickup',
        'github', 'gitlab', 'bitbucket', 'atlassian', 'jira', 'confluence',
        'salesforce', 'hubspot', 'mailchimp', 'mailerlite', 'sendgrid',
        'twilio', 'intercom', 'zendesk', 'freshdesk', 'helpscout',
        'figma', 'miro', 'invision', 'sketch',
        'shopify', 'wix', 'squarespace', 'webflow',
        'openai', 'chatgpt', 'anthropic', 'claude',
        'aws', 'amazon web services', 'azure', 'google cloud', 'gcp',
        'digitalocean', 'heroku', 'vercel', 'netlify', 'cloudflare',
        'airtable', 'zapier', 'make.com', 'ifttt',
        'loom', 'calendly', 'docusign', 'pandadoc',
        'semrush', 'ahrefs', 'moz', 'hotjar', 'mixpanel', 'amplitude',
        # Advertising platforms (billed from overseas)
        'google ads', 'google adwords', 'facebook ads', 'meta ads', 'instagram ads',
        'linkedin ads', 'twitter ads', 'tiktok ads', 'pinterest ads',
        'bing ads', 'microsoft ads',
        # Streaming & media (often billed from overseas)
        'netflix', 'spotify', 'apple music', 'youtube premium', 'disney+',
        'hulu', 'hbo', 'amazon prime', 'audible',
        # Microsoft products (often billed from Ireland/Singapore)
        'microsoft 365', 'office 365', 'microsoft office', 'onedrive',
        'linkedin premium', 'linkedin learning',
        # Other overseas services
        'godaddy', 'namecheap', 'hover', 'google domains',
        '99designs', 'fiverr', 'upwork', 'toptal',
    ]

    # Check for overseas location indicators in description
    overseas_indicators = [
        ' usa', ' us ', 'united states', 'ireland', 'singapore',
        'netherlands', 'luxembourg', 'california', 'seattle', 'san francisco',
        'delaware', 'new york', 'dublin',
    ]

    is_overseas_saas = any(keyword in description for keyword in overseas_saas_keywords)
    has_overseas_indicator = any(indicator in description for indicator in overseas_indicators)

    # If it's a known overseas SaaS or has overseas indicator in software/subscription account
    if is_overseas_saas:
        return False  # GST Free is correct for overseas digital services

    if has_overseas_indicator and ('software' in account or 'subscription' in account):
        return False  # Overseas subscription - GST Free is correct

    # Items that should typically have GST (10%) in Australia per ATO rules
    # Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/when-to-charge-gst-and-when-not-to/taxable-sales
    gst_applicable_keywords = [
        # Office supplies & stationery (taxable goods)
        'toner', 'cartridge', 'ink', 'paper', 'stationery', 'officeworks',
        'staples', 'pens', 'folders', 'envelopes', 'printer',
        # Software & subscriptions from Australian suppliers (digital products - taxable)
        'xero', 'myob', 'quickbooks',  # Australian accounting software
        'software license', 'ebook', 'e-book', 'digital download',
        'app subscription', 'saas', 'cloud software',
        # Utilities & communications (taxable services)
        'phone', 'mobile', 'telstra', 'optus', 'vodafone', 'internet',
        'nbn', 'broadband', 'electricity', 'gas', 'power', 'energy australia',
        'origin energy', 'agl',
        # Transport & travel (domestic - taxable)
        'taxi', 'uber', 'didi', 'ola', 'parking', 'car park',
        'fuel', 'petrol', 'diesel', 'caltex', 'bp ', 'shell ', '7-eleven',
        # Professional services (taxable)
        'accounting fee', 'bookkeeping', 'audit fee', 'legal fee', 'solicitor',
        'consulting fee', 'professional fee', 'advisory fee',
        # General business expenses (taxable)
        'cleaning', 'repairs', 'maintenance', 'freight', 'courier',
        'advertising', 'marketing', 'printing', 'auspost', 'australia post',
        'commercial rent', 'office rent', 'equipment hire', 'tool hire',
        # Real estate agent fees (taxable per ATO)
        'real estate fee', 'agent fee', 'agent commission', 'selling fee',
        'property sale fee', 'conveyancing',
        # Business assets (taxable when sold/disposed)
        'asset sale', 'equipment sale', 'vehicle sale', 'trade in',
        # Catering for work purposes (taxable - not entertainment)
        'catering', 'morning tea', 'working lunch', 'staff meeting',
        # Prepared/processed food (taxable)
        'restaurant', 'cafe', 'takeaway', 'fast food', 'mcdonald', 'kfc',
        'subway', 'hungry jack', 'pizza', 'uber eats', 'doordash', 'menulog',
        'biscuits', 'soft drink', 'chips', 'chocolate', 'confectionery',
        # Credit card fees/surcharges (taxable - NOT input-taxed like bank fees)
        'cc fee', 'credit card fee', 'surcharge', 'credit card surcharge',
        'card fee', 'merchant fee', 'eftpos fee', 'payment processing',
    ]

    # Items that are legitimately GST-FREE per ATO rules
    # Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/when-to-charge-gst-and-when-not-to/gst-free-sales
    gst_free_keywords = [
        # Exports (GST-free if exported within 60 days)
        'export', 'overseas delivery', 'international shipping',
        # Basic food (GST-free) - fresh, unprocessed
        'fresh fruit', 'fresh vegetable', 'fresh meat', 'fresh fish',
        'bread', 'milk', 'eggs', 'cheese', 'butter', 'flour', 'rice', 'pasta',
        'canned food', 'frozen vegetable', 'fresh produce',
        # Health & Medical services (GST-free)
        'doctor', 'gp visit', 'medical', 'hospital', 'dentist', 'dental',
        'optometrist', 'physiotherapy', 'chiropractor', 'psychologist',
        'prescription', 'pharmacy', 'chemist', 'pathology', 'x-ray', 'scan',
        'ambulance', 'health service', 'medical service',
        # Medical aids & appliances (GST-free)
        'hearing aid', 'wheelchair', 'walking frame', 'crutches',
        'blood glucose', 'insulin', 'medical equipment',
        # Education (GST-free for accredited courses)
        'tuition fee', 'course fee', 'university', 'tafe', 'school fee',
        'education', 'training course', 'accredited course',
        # Childcare (GST-free)
        'childcare', 'daycare', 'child care', 'kindergarten', 'early learning',
        # Water & sewerage (GST-free)
        'water rates', 'sewerage', 'water bill',
        # Certain religious/charitable
        'donation', 'charity', 'church',
    ]

    # Items that are INPUT-TAXED per ATO rules (no GST charged, no GST credit)
    # Source: https://www.ato.gov.au/business/gst/when-to-charge-gst-(and-when-not-to)/input-taxed-sales/
    # These should NOT be flagged as missing GST - they're correctly GST-free
    # NOTE: Credit card fees, merchant fees, eftpos fees are TAXABLE (have GST) - not input-taxed
    input_taxed_keywords = [
        # Bank fees only (NOT credit card fees which are taxable)
        'bank fee', 'bank charge', 'account fee', 'account keeping fee',
        'overdraft fee', 'dishonour fee', 'maintenance fee', 'bank maintenance',
        'account maintenance',
        # Interest (input-taxed)
        'interest', 'loan fee', 'loan interest',
        # Residential property (input-taxed)
        'residential rent', 'house rent', 'apartment rent', 'unit rent',
        # Insurance & superannuation (input-taxed financial services)
        'life insurance', 'income protection', 'superannuation', 'super contribution',
        # Shares & investments (input-taxed)
        'brokerage', 'share purchase', 'dividend', 'investment fee',
    ]

    # Check categories
    has_gst_keyword = any(keyword in description for keyword in gst_applicable_keywords)
    is_gst_free_item = any(keyword in description for keyword in gst_free_keywords)
    is_input_taxed = any(keyword in description for keyword in input_taxed_keywords)

    # Flag if it looks like it should have GST but doesn't
    # Don't flag GST-free or input-taxed items
    if has_gst_keyword and not is_gst_free_item and not is_input_taxed:
        # Check if GST rate name indicates GST Free
        is_gst_free_coded = (
            'free' in gst_rate_name or
            'exempt' in gst_rate_name or
            'n-t' in gst_rate_name or
            'no tax' in gst_rate_name or
            'bas excluded' in gst_rate_name or
            'input' in gst_rate_name or  # Input taxed
            gst_rate_name == ''
        )
        if is_gst_free_coded and gross_amount > 0:
            return True

    return False


def check_gst_calculation(transaction):
    """Check if GST calculation is correct"""
    gst_rate_name = transaction.get('gst_rate_name', '').lower()
    gst_amount = abs(transaction.get('gst', 0))
    net_amount = abs(transaction.get('net', 0))

    if net_amount == 0:
        return True

    if 'gst' in gst_rate_name and 'free' not in gst_rate_name and 'exempt' not in gst_rate_name:
        expected_gst = round(net_amount * 0.10, 2)
        return abs(gst_amount - expected_gst) <= 0.02

    return True


def check_drawings_loan_error(transaction):
    """
    Check if expense is coded to Drawings or Loan account incorrectly.
    Drawings = owner's personal withdrawals (not business expense) - should be BAS Excluded
    Loan accounts should ALWAYS be BAS Excluded.
    """
    account = transaction.get('account', '').lower()
    account_code = str(transaction.get('account_code', ''))
    gst_rate_name = transaction.get('gst_rate_name', '').lower()
    gst_amount = abs(transaction.get('gst', 0))

    # Check if coded to Drawings account
    drawings_accounts = ['drawings', 'draw', 'owner draw']
    is_drawings = any(keyword in account for keyword in drawings_accounts)

    # Check if coded to Loan account - these should ALWAYS be BAS Excluded
    loan_accounts = ['shareholder loan', 'director loan', 'related party loan',
                     'loan - director', 'loan - shareholder', 'loan to', 'loan from',
                     'loan account', 'loan payable', 'loan receivable']
    is_loan = any(keyword in account for keyword in loan_accounts)

    # Also check account codes commonly used for drawings (800-899 range often)
    try:
        code_num = int(account_code)
        if 800 <= code_num < 900:  # Common range for equity/drawings
            is_drawings = True
    except:
        pass

    # Check if it's correctly coded as BAS Excluded
    is_bas_excluded = 'bas excluded' in gst_rate_name or 'out of scope' in gst_rate_name

    # Loan accounts should ALWAYS be BAS Excluded - flag if not
    if is_loan and not is_bas_excluded:
        return True

    # Drawings with GST claimed is an error (should be BAS Excluded)
    if is_drawings:
        has_gst_coding = (
            gst_amount > 0 or
            ('gst on' in gst_rate_name and 'free' not in gst_rate_name)
        )
        if has_gst_coding:
            return True

    return False


def check_asset_capitalization(transaction):
    """
    Check if high-value assets should be capitalized instead of expensed.
    Per ATO instant asset write-off rules (2024-25): threshold is $20,000 for small business.
    Assets >= $20,000 should be capitalized and depreciated.
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gross_amount = abs(transaction.get('gross', 0))

    # Asset keywords
    asset_keywords = ['laptop', 'computer', 'macbook', 'imac', 'pc ', 'desktop',
                      'server', 'printer', 'scanner', 'monitor', 'phone', 'iphone',
                      'vehicle', 'car ', 'truck', 'van ', 'ute ', 'machinery',
                      'equipment', 'furniture', 'desk', 'chair', 'fit out', 'fitout',
                      'air conditioner', 'hvac', 'signage', 'tools']

    is_asset_purchase = any(keyword in description for keyword in asset_keywords)

    # Check if it's expensed (not capitalized)
    is_expensed = (
        'expense' in account or
        'repairs' in account or
        'maintenance' in account or
        'consumable' in account or
        'cost of' in account
    )
    is_asset_account = (
        'asset' in account or
        'plant' in account or
        'equipment' in account or
        'motor vehicle' in account or
        'computer' in account
    )

    # ATO instant asset write-off threshold is $20,000 (GST exclusive) for small business
    # Flag if asset purchase over $20,000 is expensed instead of capitalized
    threshold = 20000

    if is_asset_purchase and gross_amount >= threshold and is_expensed and not is_asset_account:
        return True

    return False


def check_computer_equipment_expense(transaction):
    """
    Check if computer equipment is incorrectly coded to an expense account.
    Per ATO depreciation rules:
    - Items <= $300: Can be immediately expensed (low-value threshold for individuals)
    - Items > $300: Should be capitalized as assets (even if instantly written off)

    For proper accounting and asset tracking, computer equipment over $300 should be
    recorded through an asset account (Computer Equipment at Cost), not directly to
    expense accounts like Office Expenses or Computer Expenses.

    Source: https://atotaxrates.info/tax-deductions/ato-depreciation/depreciation-of-computers/
    ATO effective life: Computers 4 years, Laptops/tablets 2 years
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gross_amount = abs(transaction.get('gross', 0))

    # Computer equipment keywords
    computer_keywords = [
        'laptop', 'macbook', 'notebook', 'chromebook',
        'computer', 'desktop', 'imac', 'pc ',
        'tablet', 'ipad', 'surface',
        'monitor', 'display', 'screen',
        'printer', 'scanner', 'copier',
        'server', 'nas ', 'network storage',
        'keyboard', 'mouse', 'webcam',
    ]

    is_computer_equipment = any(keyword in description for keyword in computer_keywords)

    if not is_computer_equipment:
        return False

    # Check if coded to expense account (incorrect)
    is_expensed = (
        'expense' in account or
        'office supplies' in account or
        'consumable' in account or
        'sundry' in account or
        'general' in account or
        'miscellaneous' in account
    )

    # Check if correctly coded to asset account
    is_asset_account = (
        'asset' in account or
        'computer equipment' in account or
        'office equipment' in account or
        'plant' in account or
        'at cost' in account
    )

    # Low-value threshold: $300 (items under $300 can be expensed)
    # Items over $300 should be capitalized for proper asset tracking
    threshold = 300

    if is_computer_equipment and gross_amount > threshold and is_expensed and not is_asset_account:
        return True

    return False


def check_interest_gst_error(transaction):
    """
    Check if interest income/expense has incorrect GST coding.
    Interest is INPUT-TAXED per ATO rules.
    Valid GST codes: GST Free Income/Expenses OR Input Taxed
    Invalid: GST on Income/Expenses, BAS Excluded
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_rate_name = transaction.get('gst_rate_name', '').lower()
    gst_amount = abs(transaction.get('gst', 0))

    # Interest keywords
    interest_keywords = ['interest income', 'interest received', 'interest earned',
                         'interest expense', 'interest paid', 'interest charge',
                         'loan interest', 'bank interest', 'term deposit interest']
    is_interest_account = 'interest' in account

    is_interest = any(keyword in description for keyword in interest_keywords) or is_interest_account

    if not is_interest:
        return False

    # Valid GST codes for interest: GST Free or Input Taxed
    is_gst_free = 'free' in gst_rate_name or 'exempt' in gst_rate_name
    is_input_taxed = 'input' in gst_rate_name and 'taxed' in gst_rate_name

    # Flag if GST is claimed (should not have GST)
    has_gst_claimed = (
        gst_amount > 0 or
        ('gst on' in gst_rate_name and 'free' not in gst_rate_name)
    )
    if has_gst_claimed:
        return True

    # Flag if BAS Excluded (should be GST Free or Input Taxed, not BAS Excluded)
    is_bas_excluded = 'bas excluded' in gst_rate_name or 'out of scope' in gst_rate_name
    if is_bas_excluded:
        return True

    # GST Free or Input Taxed is acceptable
    return False


def check_motor_vehicle_gst_limit(transaction):
    """
    Check if motor vehicle purchase exceeds the ATO car limit.
    Per ATO rules (2025-26):
    - Car limit: $69,674 (GST inclusive)
    - Maximum GST credit claimable: $6,334 (1/11 of car limit)
    - Cannot claim GST on Luxury Car Tax (LCT)
    - Amount over car limit cannot have GST claimed

    Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/your-industry/motor-vehicle-and-transport/gst-and-motor-vehicles/purchasing-a-motor-vehicle
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_amount = abs(transaction.get('gst', 0))
    gross_amount = abs(transaction.get('gross', 0))

    # Car purchase keywords
    car_keywords = ['car ', 'vehicle', 'motor vehicle', 'toyota', 'ford', 'holden',
                    'mazda', 'hyundai', 'kia', 'nissan', 'honda', 'bmw', 'mercedes',
                    'audi', 'volkswagen', 'subaru', 'mitsubishi', 'lexus', 'tesla',
                    'ute ', 'suv ', 'sedan', 'hatchback', 'wagon']

    is_motor_vehicle_account = 'motor vehicle' in account or 'vehicle' in account

    is_car_purchase = (
        any(keyword in description for keyword in car_keywords) or
        is_motor_vehicle_account
    )

    if not is_car_purchase:
        return False

    # ATO car limit (2025-26): $69,674
    car_limit = 69674

    # Flag if gross amount exceeds the car limit (GST on excess is not claimable)
    if gross_amount > car_limit and gst_amount > 0:
        return True

    return False


def check_overseas_subscription_gst(transaction):
    """
    Check overseas/international subscriptions and imported digital products/services.

    UPDATED: This check is now DISABLED.

    Per ATO rules, BOTH scenarios are valid for overseas digital services:
    1. WITH GST - Overseas supplier charges GST (Netflix tax) OR business applies reverse charge
       = Valid, GST credit CAN be claimed
    2. WITHOUT GST - Coded as GST Free
       = Valid, reverse charge may apply but nets to zero for most businesses

    Since both scenarios are acceptable, we no longer flag overseas subscriptions.

    Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/rules-for-specific-transactions/international-transactions/reverse-charge-gst-on-offshore-goods-and-services-purchases
    """
    # DISABLED - both GST and GST-Free are valid for overseas digital services
    return False


def check_government_charges_gst(transaction):
    """
    Check if GST is incorrectly claimed on government charges.
    Per ATO rules, NO GST applies to:
    - Stamp duty
    - Council rates
    - Land tax
    - ASIC fees
    - Motor vehicle registration
    - Water rates
    - Government fines/penalties

    Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/your-industry/financial-services-and-insurance/gst-and-insurance
    """
    description = transaction.get('description', '').lower()
    gst_amount = abs(transaction.get('gst', 0))

    # Government charges keywords - NO GST on these
    # Note: 'fine' and 'penalty' removed - handled by fines_penalties_gst check
    govt_charge_keywords = [
        'stamp duty', 'council rates', 'land tax', 'asic', 'rego ',
        'registration fee', 'motor vehicle registration', 'water rates',
        'government fee', 'govt fee', 'infringement',
        'court fee', 'filing fee', 'lodgement fee', 'license fee', 'licence fee',
        'payroll tax', 'workers comp levy', 'epa levy',
    ]

    is_govt_charge = any(keyword in description for keyword in govt_charge_keywords)

    if not is_govt_charge:
        return False

    # Flag if GST is claimed on government charges
    if gst_amount > 0:
        return True

    return False


def check_client_entertainment_gst(transaction):
    """
    Check if GST is incorrectly claimed on client entertainment.
    Per ATO rules:
    - Client entertainment: NO GST credit claimable
    - Staff entertainment: GST claimable only if FBT is paid
    - Minor benefits exemption (<$300): NO GST credit

    Source: https://www.ato.gov.au/businesses-and-organisations/hiring-and-paying-your-workers/fringe-benefits-tax/types-of-fringe-benefits/entertainment-related-fringe-benefits/common-entertainment-scenarios-for-business
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_amount = abs(transaction.get('gst', 0))

    # Client entertainment keywords
    client_entertainment_keywords = [
        'client lunch', 'client dinner', 'client meal', 'client entertainment',
        'business lunch', 'business dinner', 'business meal',
        'client gift', 'corporate hospitality', 'client function',
        'networking event', 'golf day', 'sporting event',
    ]

    # Check if it's explicitly client entertainment
    is_client_entertainment = any(keyword in description for keyword in client_entertainment_keywords)

    # Check if it's in an entertainment account with client context
    is_entertainment_account = 'entertainment' in account
    has_client_context = 'client' in description or 'customer' in description

    if is_client_entertainment or (is_entertainment_account and has_client_context):
        # Flag if GST is claimed on client entertainment
        if gst_amount > 0:
            return True

    return False


def check_staff_entertainment_gst(transaction):
    """
    Check if GST is incorrectly claimed on staff entertainment.
    Per ATO rules:
    - Entertainment expenses are NOT GST claimable UNLESS FBT is paid
    - If FBT exempt (minor benefits <$300, on-premises meals), NO GST credit
    - Christmas parties on premises = FBT exempt = NO GST credit
    - Staff functions/events = NO GST credit (unless FBT paid)

    EXCEPTIONS that ARE GST claimable:
    - Light refreshments (morning/afternoon tea, biscuits)
    - Staff amenities (tea/coffee facilities on premises)

    Source: https://www.ato.gov.au/businesses-and-organisations/hiring-and-paying-your-workers/fringe-benefits-tax/types-of-fringe-benefits/entertainment-related-fringe-benefits
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_amount = abs(transaction.get('gst', 0))

    # Check if it's in an Entertainment account
    is_entertainment_account = 'entertainment' in account

    # Staff entertainment keywords (not client-related)
    staff_entertainment_keywords = [
        'christmas party', 'xmas party', 'staff party', 'end of year party',
        'staff lunch', 'staff dinner', 'staff function', 'team lunch', 'team dinner',
        'team building', 'staff event', 'office party', 'farewell', 'farewell lunch',
        'farewell dinner', 'welcome lunch', 'welcome dinner', 'celebration',
        'staff celebration', 'company event', 'end of year function',
        'eofy party', 'eofy function', 'staff drinks', 'friday drinks',
    ]

    # Check for restaurant/cafe in Entertainment account (meal entertainment)
    restaurant_keywords = [
        'restaurant', 'cafe', 'bistro', 'diner', 'eatery', 'pub ', 'bar ',
        'tavern', 'hotel ', 'brewery', 'pizzeria', 'steakhouse', 'sushi',
        'thai ', 'chinese ', 'italian ', 'indian ', 'mexican ', 'japanese ',
    ]

    # Light refreshments - NOT entertainment, GST claimable
    light_refreshment_keywords = [
        'morning tea', 'afternoon tea', 'tea and coffee', 'coffee and tea',
        'biscuits', 'fruit', 'sandwiches', 'light lunch', 'working lunch',
    ]

    # Check if it's light refreshments (excluded from entertainment rules)
    is_light_refreshment = any(keyword in description for keyword in light_refreshment_keywords)
    if is_light_refreshment:
        return False  # Light refreshments are NOT entertainment - GST claimable

    is_staff_entertainment = any(keyword in description for keyword in staff_entertainment_keywords)
    is_restaurant_expense = any(keyword in description for keyword in restaurant_keywords)

    # Flag if:
    # 1. It's explicitly staff entertainment with GST claimed
    # 2. It's a restaurant/cafe expense in Entertainment account with GST claimed
    if gst_amount > 0:
        if is_staff_entertainment:
            return True
        if is_entertainment_account and is_restaurant_expense:
            return True

    return False


def check_residential_premises_gst(transaction):
    """
    Check if GST is incorrectly claimed on residential property expenses.

    RESIDENTIAL PROPERTY (INPUT-TAXED) - NO GST credits:
    Per ATO rules, residential rent is INPUT-TAXED meaning:
    - No GST is charged on residential rent
    - NO GST credits can be claimed on related expenses
    - Property management fees, agent commissions, repairs, maintenance,
      advertising, insurance - all NO GST credit

    COMMERCIAL PROPERTY (TAXABLE) - GST credits allowed:
    Per ATO rules, commercial property lease/rent is TAXABLE:
    - GST IS charged on commercial rent
    - GST credits CAN be claimed on related expenses
    - Office, warehouse, retail, factory, industrial - GST claimable

    Exceptions (GST can apply on residential):
    - New residential premises (first sale by developer)
    - Commercial residential premises (hotels, motels, serviced apartments)
    - Short-term accommodation (Airbnb-style where commercial)

    Sources:
    - https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/when-to-charge-gst-and-when-not-to/input-taxed-sales/residential-premises
    - https://www.ato.gov.au/businesses-and-organisations/assets-and-property/property/property-used-in-running-a-business/leasing-and-renting-commercial-premises
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_amount = abs(transaction.get('gst', 0))

    # Residential property expense keywords
    residential_property_keywords = [
        'rental property', 'investment property', 'residential property',
        'rental expense', 'property expense', 'landlord',
        'tenant', 'tenancy', 'lease residential', 'residential lease',
        'unit ', 'apartment ', 'house ', 'townhouse', 'villa ',
    ]

    # Property management/agent keywords
    property_management_keywords = [
        'property management', 'property manager', 'real estate agent',
        'letting fee', 'letting agent', 'rental agent', 'rental management',
        'property agent', 'agent fee', 'agent commission', 'management fee',
        're agent', 'estate agent',
    ]

    # Residential repairs/maintenance keywords
    # NOTE: Do NOT include "strata fees", "body corporate", "owners corporation" here
    # as these can be for EITHER residential OR commercial properties.
    # Only flag as residential if there's explicit evidence (e.g., "residential strata")
    residential_repairs_keywords = [
        'rental repairs', 'rental maintenance', 'property repairs',
        'property maintenance', 'tenant repairs', 'landlord insurance',
        'rental insurance', 'building insurance residential',
        'residential strata', 'residential body corporate',
    ]

    # Advertising for tenants
    tenant_advertising_keywords = [
        'tenant advertising', 'rental advertising', 'property advertising',
        'realestate.com', 'domain.com', 'rental listing', 'tenant finder',
    ]

    # Account names suggesting residential property
    is_rental_property_account = (
        'rental' in account or
        'investment property' in account or
        'residential' in account
    )

    # Check if expense is related to residential property
    is_residential_property_expense = (
        any(keyword in description for keyword in residential_property_keywords) or
        any(keyword in description for keyword in property_management_keywords) or
        any(keyword in description for keyword in residential_repairs_keywords) or
        any(keyword in description for keyword in tenant_advertising_keywords) or
        is_rental_property_account
    )

    if not is_residential_property_expense:
        return False

    # Exclude commercial property - GST IS claimable on commercial premises
    # Per ATO: Commercial property lease/rent is TAXABLE (GST applies, credits allowed)
    # Source: https://www.ato.gov.au/businesses-and-organisations/assets-and-property/property/property-used-in-running-a-business/leasing-and-renting-commercial-premises
    commercial_property_keywords = [
        'commercial property', 'commercial premises', 'commercial lease',
        'office lease', 'office rent', 'warehouse', 'retail shop', 'shop rent',
        'factory', 'industrial', 'business premises', 'commercial building',
        'office space', 'retail premises', 'medical centre', 'shopping centre',
    ]

    is_commercial_property = any(keyword in description for keyword in commercial_property_keywords)
    if is_commercial_property:
        return False  # Commercial property - GST IS claimable

    # Exclude commercial residential (hotels, motels, serviced apartments, Airbnb)
    # Per ATO: Commercial residential premises are TAXABLE
    commercial_residential_keywords = [
        'hotel', 'motel', 'serviced apartment', 'airbnb', 'short term',
        'short-term', 'holiday rental', 'vacation rental', 'hostel',
        'boarding house', 'student accommodation',
    ]

    is_commercial_residential = any(keyword in description for keyword in commercial_residential_keywords)
    if is_commercial_residential:
        return False  # Commercial residential - GST may apply

    # Flag if GST is claimed on residential property expense (should be input-taxed)
    if gst_amount > 0:
        return True

    return False


def infer_business_context(all_transactions):
    """
    Analyze all transactions to infer the client's business type/industry.
    This helps validate that sales transactions make sense for the business.

    Returns a dict with:
    - industry: The inferred industry type
    - likely_income_sources: List of expected income types for this business
    - confidence: How confident we are in the inference
    """
    if not all_transactions:
        return {'industry': 'unknown', 'likely_income_sources': [], 'confidence': 0}

    # Collect all descriptions and accounts
    all_text = ' '.join([
        (t.get('description', '') + ' ' + t.get('account', '')).lower()
        for t in all_transactions
    ])

    # Industry detection patterns based on expenses and accounts
    industry_patterns = {
        'it_services': {
            'keywords': ['software', 'microsoft', 'adobe', 'it support', 'tech', 'computer',
                        'server', 'network', 'cloud', 'aws', 'azure', 'hosting', 'domain',
                        'website', 'development', 'programming', 'database', 'cybersecurity'],
            'income_sources': ['training', 'support', 'consulting', 'implementation', 'project',
                              'development', 'maintenance', 'software', 'services', 'hourly',
                              'rate as agreed', 'fixed fee', 'retainer', 'technical']
        },
        'professional_services': {
            'keywords': ['consulting', 'advisory', 'professional', 'project management',
                        'strategy', 'business', 'management', 'coaching', 'mentoring'],
            'income_sources': ['consulting', 'advisory', 'coaching', 'mentoring', 'training',
                              'project', 'engagement', 'retainer', 'hourly', 'fee']
        },
        'management_consulting': {
            'keywords': ['consulting', 'consultant', 'advisory', 'strategy', 'transformation',
                        'change management', 'business analysis', 'process improvement',
                        'stakeholder', 'workshop', 'facilitation', 'governance', 'compliance',
                        'risk management', 'due diligence', 'merger', 'acquisition', 'restructure',
                        'operating model', 'target operating', 'capability', 'benchmark',
                        'kpi', 'dashboard', 'report', 'presentation', 'client meeting',
                        'big 4', 'deloitte', 'pwc', 'kpmg', 'ey', 'mckinsey', 'bcg', 'bain'],
            'income_sources': ['consulting', 'advisory', 'strategy', 'engagement', 'project',
                              'retainer', 'monthly fee', 'fixed fee', 'hourly', 'per hour',
                              'rate as agreed', 'professional fee', 'management fee',
                              'workshop', 'facilitation', 'training', 'coaching', 'review',
                              'assessment', 'analysis', 'recommendations', 'implementation']
        },
        'trades': {
            'keywords': ['bunnings', 'hardware', 'tools', 'materials', 'building', 'construction',
                        'plumbing', 'electrical', 'carpentry', 'painting', 'renovation'],
            'income_sources': ['labour', 'materials', 'installation', 'repair', 'maintenance',
                              'quote', 'job', 'project', 'contract', 'site work']
        },
        'automotive': {
            'keywords': ['mechanic', 'workshop', 'car parts', 'auto parts', 'spare parts',
                        'repco', 'supercheap', 'burson', 'autobarn', 'bursons',
                        'oil', 'filter', 'brake', 'tyre', 'tire', 'battery', 'engine',
                        'transmission', 'gearbox', 'exhaust', 'suspension', 'radiator',
                        'car service', 'vehicle service', 'roadworthy', 'rwc', 'pink slip',
                        'rego check', 'log book', 'smash repair', 'panel beater', 'spray paint',
                        'detailing', 'car wash', 'windscreen', 'wiper', 'alternator', 'starter motor'],
            'income_sources': ['repair', 'service', 'labour', 'parts', 'car service', 'vehicle service',
                              'mechanic', 'roadworthy', 'rwc', 'pink slip', 'inspection',
                              'log book service', 'major service', 'minor service', 'brake service',
                              'tyre', 'wheel alignment', 'smash repair', 'panel', 'paint',
                              'detailing', 'diagnostic', 'quote', 'job']
        },
        'health_medical': {
            'keywords': ['medical supplies', 'clinic', 'patient', 'health', 'therapy',
                        'physio', 'chiro', 'dental', 'psychology', 'ndis', 'medicare'],
            'income_sources': ['consultation', 'treatment', 'therapy', 'session', 'appointment',
                              'ndis', 'medicare', 'patient', 'health service']
        },
        'retail': {
            'keywords': ['inventory', 'stock', 'merchandise', 'cost of goods', 'wholesale',
                        'supplier', 'products', 'retail'],
            'income_sources': ['sales', 'product', 'merchandise', 'goods', 'order']
        },
        'hospitality': {
            'keywords': ['food', 'beverage', 'restaurant', 'cafe', 'catering', 'kitchen',
                        'ingredients', 'menu', 'bar', 'pub', 'bistro', 'takeaway', 'coffee',
                        'chef', 'waitstaff', 'front of house', 'food cost', 'liquor',
                        'pos system', 'uber eats', 'deliveroo', 'menulog', 'doordash',
                        'foodworks', 'bidfood', 'pfd', 'countrywide', 'bakery', 'patisserie'],
            'income_sources': ['catering', 'function', 'event', 'meal', 'food', 'beverage',
                              'dine in', 'takeaway', 'delivery', 'coffee', 'drinks', 'bar',
                              'breakfast', 'lunch', 'dinner', 'functions', 'private event',
                              'sales', 'daily takings', 'cash sales', 'card sales']
        },
        'agricultural': {
            'keywords': ['farm', 'livestock', 'cattle', 'sheep', 'crop', 'harvest', 'tractor',
                        'fertiliser', 'fertilizer', 'seed', 'feed', 'hay', 'grain', 'wool',
                        'dairy', 'elders', 'landmark', 'agri', 'rural', 'paddock', 'irrigation',
                        'stockfeed', 'veterinary', 'vet', 'shearing', 'fencing', 'silo'],
            'income_sources': ['livestock', 'cattle', 'sheep', 'wool', 'grain', 'crop', 'harvest',
                              'produce', 'dairy', 'milk', 'sale yards', 'auction', 'agistment',
                              'farm produce', 'hay', 'fodder', 'breeding', 'stud']
        },
        'education_training': {
            'keywords': ['training', 'course', 'curriculum', 'education', 'student', 'workshop',
                        'seminar', 'learning', 'materials', 'instructor'],
            'income_sources': ['training', 'course', 'workshop', 'seminar', 'tuition',
                              'instruction', 'session', 'program', 'per hour', 'rate as agreed']
        }
    }

    # Score each industry
    industry_scores = {}
    for industry, patterns in industry_patterns.items():
        score = sum(1 for kw in patterns['keywords'] if kw in all_text)
        industry_scores[industry] = score

    # Get the best match
    if not industry_scores or max(industry_scores.values()) == 0:
        return {'industry': 'general_business', 'likely_income_sources': [
            'services', 'fee', 'invoice', 'hourly', 'project', 'consulting'
        ], 'confidence': 0.3}

    best_industry = max(industry_scores, key=industry_scores.get)
    confidence = min(1.0, industry_scores[best_industry] / 10)  # Cap at 1.0

    return {
        'industry': best_industry,
        'likely_income_sources': industry_patterns[best_industry]['income_sources'],
        'confidence': confidence
    }


# Global variable to store inferred business context (set during review)
_business_context = None


def set_business_context(all_transactions):
    """Set the business context based on all transactions."""
    global _business_context
    _business_context = infer_business_context(all_transactions)
    return _business_context


def get_business_context():
    """Get the current business context."""
    global _business_context
    if _business_context is None:
        return {'industry': 'unknown', 'likely_income_sources': [], 'confidence': 0}
    return _business_context


# Global variable to store detected allocation patterns from deep scan
_allocation_patterns = {}


def round_to_common_split(account_percentages):
    """
    Round two-way split percentages to common business allocation ratios.

    Common splits: 90/10, 75/25, 70/30, 67/33, 60/40, 50/50
    Note: 80/20 removed as it often misrepresents actual 70/30 splits due to price differences.
    This makes the displayed percentages more meaningful for typical business allocations.
    """
    if len(account_percentages) != 2:
        return account_percentages

    # Common split ratios - ordered to prefer lower ratios when equidistant
    # 80/20 and 75/25 omitted - often caused by price variations in 70/30 splits
    common_splits = [0.70, 0.67, 0.60, 0.50, 0.90]

    # Get the two accounts and their percentages
    accounts = list(account_percentages.keys())
    percentages = list(account_percentages.values())

    # Identify higher and lower percentage
    if percentages[0] >= percentages[1]:
        high_idx, low_idx = 0, 1
    else:
        high_idx, low_idx = 1, 0

    high_pct = percentages[high_idx]

    # Find closest common split
    closest_split = min(common_splits, key=lambda x: abs(x - high_pct))

    # Round if within 12% of a common split (generous tolerance for price variations)
    if abs(closest_split - high_pct) <= 0.12:
        return {
            accounts[high_idx]: closest_split,
            accounts[low_idx]: 1.0 - closest_split
        }

    return account_percentages


def detect_allocation_patterns(all_transactions):
    """
    Analyze 12 months of transaction history to detect allocation patterns.

    Detects patterns like:
    - Telstra expenses split 70% to Telephone, 30% to Drawings
    - Fuel split between Motor Vehicle and Cost of Sales
    - Any vendor/description that consistently goes to multiple accounts

    Returns a dict of patterns:
    {
        'telstra': {
            'accounts': {'telephone': 0.7, 'drawings': 0.3},
            'count': 50,
            'is_split_allocation': True
        },
        ...
    }
    """
    if not all_transactions:
        return {}

    # Group transactions by vendor/description keywords
    vendor_accounts = {}  # vendor_key -> {account: total_amount}
    vendor_counts = {}    # vendor_key -> count

    # Common vendor keywords to track
    vendor_keywords = [
        'telstra', 'optus', 'vodafone', 'tpg', 'iinet', 'aussie broadband',
        'origin', 'agl', 'energy australia', 'alinta',
        'ampol', 'bp ', 'caltex', 'shell', '7-eleven', 'united petroleum',
        'officeworks', 'bunnings', 'jb hi-fi', 'harvey norman',
        'coles', 'woolworths', 'aldi', 'iga',
        'commbank', 'westpac', 'anz ', 'nab ', 'st george',
        'paypal', 'stripe', 'square',
        'uber', 'didi', 'ola',
        'qantas', 'virgin', 'jetstar', 'rex ',
        'aws', 'azure', 'google cloud', 'microsoft', 'adobe', 'xero',
    ]

    for t in all_transactions:
        # Check contact, narration, reference, AND description for vendor matching
        description = (t.get('description', '') or '').lower()
        contact = (t.get('contact', '') or '').lower()
        reference = (t.get('reference', '') or '').lower()
        narration = (t.get('narration', '') or '').lower()
        search_text = f"{contact} {narration} {reference} {description}"

        account = (t.get('account', '') or '').lower()
        amount = abs(t.get('gross', 0) or t.get('amount', 0) or 0)

        if not account or amount == 0:
            continue

        # Find matching vendor keyword
        for vendor in vendor_keywords:
            if vendor in search_text:
                if vendor not in vendor_accounts:
                    vendor_accounts[vendor] = {}
                    vendor_counts[vendor] = 0

                if account not in vendor_accounts[vendor]:
                    vendor_accounts[vendor][account] = 0

                vendor_accounts[vendor][account] += amount
                vendor_counts[vendor] += 1
                break  # Only match first vendor

    # Analyze patterns
    patterns = {}
    for vendor, accounts in vendor_accounts.items():
        if vendor_counts[vendor] < 5:  # Need at least 5 transactions to establish pattern
            continue

        total = sum(accounts.values())
        if total == 0:
            continue

        # Calculate percentage allocation per account
        account_percentages = {acct: amt / total for acct, amt in accounts.items()}

        # Check if this is a split allocation (multiple accounts with significant %)
        significant_accounts = {acct: pct for acct, pct in account_percentages.items() if pct >= 0.1}  # 10%+

        # Round to common business split ratios (70/30, 60/40, 50/50, 80/20, 90/10)
        if len(significant_accounts) == 2:
            significant_accounts = round_to_common_split(significant_accounts)

        is_split = len(significant_accounts) > 1

        patterns[vendor] = {
            'accounts': significant_accounts,
            'count': vendor_counts[vendor],
            'total_amount': total,
            'is_split_allocation': is_split
        }

    return patterns


def set_allocation_patterns(patterns):
    """Set the detected allocation patterns from deep scan."""
    global _allocation_patterns
    _allocation_patterns = patterns
    return _allocation_patterns


def get_allocation_patterns():
    """Get the detected allocation patterns."""
    global _allocation_patterns
    return _allocation_patterns


def is_correctly_coded_telstra(transaction):
    """
    Check if a Telstra transaction is correctly coded and should NOT be flagged.

    Correct codings:
    - "telstra business" in Telephone/Internet account with GST → CORRECT
    - "telstra personal" in Drawings/Loan account with BAS Excluded → CORRECT

    Returns True if correctly coded (should skip flagging), False otherwise.
    """
    description = (transaction.get('description', '') or '').lower()
    account = (transaction.get('account', '') or '').lower()
    contact = (transaction.get('contact', '') or '').lower()
    narration = (transaction.get('narration', '') or '').lower()
    search_text = f"{description} {contact} {narration}"

    # Check if this is a Telstra transaction
    if 'telstra' not in search_text:
        return False

    # Business expense accounts where Telstra business is correct
    business_accounts = ['telephone', 'internet', 'phone', 'communication']

    # Personal/drawings accounts where Telstra personal is correct
    personal_accounts = ['drawing', 'drawings', 'loan', 'private', 'personal']

    # Check for business portion correctly coded
    is_business_desc = 'business' in description or 'business' in narration
    is_business_account = any(acct in account for acct in business_accounts)

    if is_business_desc and is_business_account:
        return True  # Telstra business in Telephone - CORRECT

    # Check for personal portion correctly coded
    is_personal_desc = 'personal' in description or 'personal' in narration or 'private' in description
    is_personal_account = any(acct in account for acct in personal_accounts)

    if is_personal_desc and is_personal_account:
        return True  # Telstra personal in Drawings - CORRECT

    # Also allow Telstra (without business/personal suffix) in Telephone
    # as this is the common coding for 100% business use
    if is_business_account and not is_personal_desc:
        return True  # Telstra in Telephone without "personal" - assume business

    return False


def is_known_allocation_pattern(transaction):
    """
    Check if a transaction matches a known NON-SPLIT allocation pattern.

    Only returns True for vendors that ALWAYS go to the same account.
    If a vendor has a SPLIT pattern (e.g., Telstra 70/30), we DON'T skip -
    we want to flag it so the user checks if the split was done correctly.

    Returns:
    - True if this matches a known SINGLE-account pattern (don't flag)
    - False if this is a split pattern or no patterns detected
    """
    patterns = get_allocation_patterns()
    if not patterns:
        return False

    # Check contact, narration, reference, and description for vendor matching
    description = (transaction.get('description', '') or '').lower()
    contact = (transaction.get('contact', '') or '').lower()
    reference = (transaction.get('reference', '') or '').lower()
    narration = (transaction.get('narration', '') or '').lower()
    search_text = f"{contact} {narration} {reference} {description}"
    account = (transaction.get('account', '') or '').lower()

    for vendor, pattern in patterns.items():
        if vendor in search_text:
            # Only skip flagging if this vendor does NOT have a split allocation
            # If it's a split pattern, we want to flag it for review
            if not pattern.get('is_split_allocation', False):
                # Single account allocation - check if current account matches
                if account in pattern['accounts'] or any(acct in account for acct in pattern['accounts']):
                    return True
                for known_acct in pattern['accounts']:
                    if known_acct in account or account in known_acct:
                        return True

    return False


def check_split_allocation_pattern(transaction):
    """
    Check if a transaction should have been split but wasn't.

    If Telstra is known to be split 70/30 between Telephone and Drawings,
    then a Telstra expense that's 100% to Telephone should be FLAGGED
    because it might be missing the Drawings portion.

    Returns:
    - dict with 'should_flag', 'vendor', 'expected_split' if pattern mismatch found
    - None if no issue
    """
    patterns = get_allocation_patterns()
    if not patterns:
        return None

    # Check contact, narration, reference, and description for vendor matching
    description = (transaction.get('description', '') or '').lower()
    contact = (transaction.get('contact', '') or '').lower()
    reference = (transaction.get('reference', '') or '').lower()
    narration = (transaction.get('narration', '') or '').lower()
    search_text = f"{contact} {narration} {reference} {description}"
    account = (transaction.get('account', '') or '').lower()

    for vendor, pattern in patterns.items():
        if vendor in search_text:
            # Check if this vendor has a split allocation pattern
            if pattern.get('is_split_allocation', False):
                # This vendor normally has split allocations
                # Only flag if the description/account combination seems WRONG

                # Check if description indicates personal vs business
                is_personal_desc = 'personal' in description or 'private' in description
                is_business_desc = 'business' in description

                # Check if account is personal (drawings/loan) vs business expense
                is_personal_account = any(x in account for x in ['drawing', 'loan', 'private'])
                is_business_account = any(x in account for x in ['telephone', 'internet', 'expense', 'motor', 'fuel', 'office'])

                # Flag MISMATCH: personal description but business account, or vice versa
                should_flag = False
                if is_personal_desc and is_business_account:
                    should_flag = True  # "telstra personal" in Telephone = WRONG
                elif is_business_desc and is_personal_account:
                    should_flag = True  # "telstra business" in Drawings = WRONG

                if should_flag:
                    account_names = ', '.join(pattern['accounts'].keys())
                    return {
                        'should_flag': True,
                        'vendor': vendor,
                        'expected_split': account_names,
                        'transaction_count': pattern.get('count', 0)
                    }

    return None


def check_sales_gst_error(transaction):
    """
    Check if Sales has incorrect GST coding.
    Sales can ONLY be GST on Income or GST Free Income - NEVER BAS Excluded.

    Uses business context to validate that sales transactions make sense
    for the inferred industry type.

    Per ATO rules, most sales are GST inclusive (10%) EXCEPT (GST-FREE categories):
    - Medical/health services (doctors, physiotherapy, chiropractic, psychology, etc.)
    - Education courses (accredited pre-school, primary, secondary, tertiary)
    - Childcare services (daycare, early learning, after school care)
    - Exports (goods exported within 60 days, services for overseas consumption)
    - Basic food (raw meats, fruits, vegetables, bread, cooking ingredients)
    - Water, sewerage, drainage services
    - Religious services
    - Certain charity activities
    - Cars for people with disability
    - Community care for elderly/disabled
    - Farmland sales

    IMPORTANT: Commercial/professional training (IT, software, business training) is NOT GST-free.
    Only accredited educational courses (schools, universities, TAFE) are GST-free.

    Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/when-to-charge-gst-and-when-not-to/gst-free-sales
    """
    account = transaction.get('account', '').lower()
    description = transaction.get('description', '').lower()
    gst_rate_name = transaction.get('gst_rate_name', '').lower()
    gst_amount = abs(transaction.get('gst', 0))

    # Check if it's a Sales account (NOT Other Income)
    is_sales_account = 'sales' in account and 'other' not in account

    if not is_sales_account:
        return False

    # Sales should NEVER be BAS Excluded
    is_bas_excluded = 'bas excluded' in gst_rate_name or 'out of scope' in gst_rate_name
    if is_bas_excluded:
        return True

    # Get business context to understand what income sources are expected
    business_context = get_business_context()
    likely_income_sources = business_context.get('likely_income_sources', [])

    # Check if this sale matches the expected income sources for this business
    is_expected_income = any(source in description for source in likely_income_sources)

    # Commercial/professional service revenue - these should have GST (not GST-free)
    # Don't flag these as errors - they are correctly coded with GST on Income
    commercial_service_keywords = [
        # IT/Software services
        'training', 'microsoft', 'ms office', 'office 365', 'excel', 'word', 'powerpoint',
        'software', 'it support', 'tech support', 'technical support', 'computer',
        'website', 'web development', 'app development', 'programming', 'coding',
        'network', 'server', 'cloud', 'database', 'cybersecurity', 'data',
        # Professional services
        'consulting', 'consultancy', 'advisory', 'advice', 'strategy',
        'project management', 'business', 'management', 'coaching', 'mentoring',
        # Service delivery terms
        'hourly', 'per hour', 'rate as agreed', 'as agreed', 'fixed fee',
        'retainer', 'monthly fee', 'service fee', 'support', 'maintenance',
        'implementation', 'setup', 'installation', 'configuration', 'migration',
        # General service terms
        'invoice', 'services', 'professional', 'billable', 'engagement',
    ]

    # Check if this is commercial service revenue (should have GST, not GST-free)
    is_commercial_service = any(keyword in description for keyword in commercial_service_keywords)

    # If it has GST (GST on Income), and it's either:
    # 1. A commercial service, OR
    # 2. An expected income source for this business type
    # Then it's CORRECT - don't flag
    has_gst = gst_amount > 0 and 'free' not in gst_rate_name and 'exempt' not in gst_rate_name
    if has_gst and (is_commercial_service or is_expected_income):
        return False  # Correctly coded with GST - matches business type

    # GST-free sales categories per ATO rules
    # These are the ONLY valid reasons for sales to be GST Free
    # NOTE: Only ACCREDITED education is GST-free, not commercial training
    gst_free_sales_keywords = [
        # Medical/Health services (s38-7 and s38-10 GST Act)
        'medical', 'doctor', 'gp ', 'specialist', 'hospital', 'surgery',
        'physiotherapy', 'physio', 'chiropractic', 'chiropractor',
        'psychology', 'psychologist', 'counselling', 'counseling',
        'acupuncture', 'naturopath', 'naturopathy', 'osteopath', 'osteopathy',
        'podiatry', 'podiatrist', 'chiropody', 'dental', 'dentist',
        'optometry', 'optometrist', 'audiology', 'audiologist',
        'occupational therapy', 'speech therapy', 'speech pathology',
        'nursing', 'dietitian', 'dietary', 'pharmacy', 'pharmacist',
        'herbal medicine', 'chinese medicine', 'social work',
        'massage therapy', 'remedial massage', 'myotherapy',
        'therapy', 'therapeutic', 'treatment', 'consultation',
        'health service', 'allied health', 'ndis', 'medicare',
        'aged care', 'disability', 'community care',
        # ACCREDITED Education only (s38-85 GST Act) - NOT commercial training
        'tuition', 'course fee', 'accredited course', 'accredited education',
        'university', 'tafe', 'school fee', 'vocational education',
        'pre-school', 'preschool', 'primary school', 'secondary school',
        'curriculum', 'rto course', 'registered training organisation',
        # Childcare (s38-145 GST Act)
        'childcare', 'child care', 'daycare', 'day care', 'kindergarten',
        'early learning', 'after school care', 'vacation care',
        'before school care', 'oshc', 'family day care',
        # Exports (s38-185 GST Act)
        'export', 'overseas sale', 'international sale',
        # Basic food (Div 38-A GST Act)
        'food sale', 'produce sale', 'fruit sale', 'vegetable sale',
        'meat sale', 'bread sale', 'farm produce', 'agricultural',
        # Water/sewerage (s38-350 GST Act)
        'water supply', 'sewerage', 'drainage',
        # Religious services (s38-220 GST Act)
        'religious service', 'church service', 'worship',
        # Charity (Div 38-G GST Act)
        'charity', 'charitable', 'donation', 'fundraising',
        # Farmland (s38-475 GST Act)
        'farmland', 'farm sale',
        # Insurance settlements (GST-free if insurer notified of GST status)
        'insurance settlement', 'insurance payout', 'insurance claim',
        'insurance recovery', 'claim settlement', 'insurance proceeds',
        # Grants (GST-free if no supply in return)
        'government grant', 'grant income', 'subsidy', 'jobkeeper',
        'cash flow boost', 'stimulus payment', 'covid support',
        'business support grant', 'wage subsidy', 'apprentice subsidy',
    ]

    # Check if GST Free is coded
    is_gst_free = (
        'free' in gst_rate_name or
        'exempt' in gst_rate_name or
        gst_amount == 0
    )

    # If GST Free, check if it's a valid GST-free category
    if is_gst_free:
        # Commercial services coded as GST-free is an ERROR (should have GST)
        if is_commercial_service:
            return True  # Error: commercial service should have GST

        is_valid_gst_free = any(keyword in description for keyword in gst_free_sales_keywords)
        if not is_valid_gst_free:
            # GST Free sales without valid reason - flag for review
            return True

    return False


def check_export_gst_error(transaction):
    """
    Check if export sales/income has incorrect GST coding.
    Exports should be GST-FREE (no GST charged, but CAN claim input credits).

    Per ATO rules for GST-free exports:
    - Goods must be exported within 60 days of payment/invoice
    - Services must be used/enjoyed outside Australia OR
      made to non-resident not in Australia at time of supply
    - Documentation required: shipping docs, customs declaration, proof of delivery

    IMPORTANT: If GST is charged on exports, this is INCORRECT.
    Export sales should be coded as 'GST Free' or 'Export' tax code.

    Source: https://www.ato.gov.au/businesses-and-organisations/international-tax-for-business/australians-doing-business-overseas/exports-and-gst
    GSTR 2002/6 - When supplies of goods are GST-free exports
    """
    account = transaction.get('account', '').lower()
    description = transaction.get('description', '').lower()
    gst_rate_name = transaction.get('gst_rate_name', '').lower()
    gst_amount = abs(transaction.get('gst', 0))

    # Check if it's an income/sales account
    is_income_account = (
        'sales' in account or
        'income' in account or
        'revenue' in account or
        'export' in account or
        'overseas' in account or
        'international' in account or
        'foreign' in account or
        'trading' in account or
        'turnover' in account
    )

    # If account check fails but description clearly indicates export sale, still proceed
    if not is_income_account:
        # Check if description strongly indicates it's an export sale
        if 'export' in description and ('sale' in description or 'invoice' in description):
            is_income_account = True  # Allow it to proceed

    if not is_income_account:
        return False

    # Export-related keywords
    export_keywords = [
        # Direct export terms
        'export', 'exported', 'exporting',
        # International sales
        'overseas sale', 'overseas customer', 'overseas client',
        'international sale', 'international customer', 'international client',
        'foreign sale', 'foreign customer', 'foreign client',
        # Shipping overseas
        'shipped overseas', 'shipping overseas', 'international shipping',
        'overseas delivery', 'international delivery', 'freight overseas',
        'sea freight export', 'air freight export',
        # Services to overseas
        'overseas service', 'international service', 'offshore service',
        'consulting overseas', 'services overseas',
        # Specific countries/regions (common export destinations)
        'usa sale', 'us sale', 'uk sale', 'nz sale', 'new zealand',
        ' nz', 'to nz', 'nz -', 'nz-',  # NZ abbreviations
        'singapore', 'hong kong', 'japan', 'china', 'europe',
        'asia pacific', 'apac',
        # Currency indicators for sales
        'usd invoice', 'gbp invoice', 'eur invoice', 'nzd invoice',
    ]

    is_export_sale = any(keyword in description for keyword in export_keywords)

    if not is_export_sale:
        return False

    # Flag if GST is charged on export (should be GST-free)
    # GST on Income means 10% GST was charged - incorrect for exports
    has_gst = gst_amount > 0 or 'gst on income' in gst_rate_name

    if has_gst:
        return True

    return False


def check_other_income_error(transaction):
    """
    Check if Other Income/Sundry Income has incorrect GST coding.

    Per ATO rules on income GST treatment:

    TAXABLE INCOME (GST on Income - 10%):
    - Business income from sales and services
    - Commission income
    - Rental income (commercial property)
    - Reimbursements received (with GST component)
    - Insurance payouts (for taxable supplies)
    - Asset sale proceeds (business assets)

    GST-FREE INCOME (GST Free Income):
    - Export income
    - Medical/health services
    - Educational services
    - Childcare services
    - Some food sales

    INPUT-TAXED INCOME (no GST, no credits):
    - Interest income (financial supply)
    - Dividend income (financial supply)
    - Residential rental income

    BAS EXCLUDED (not reported on BAS):
    - Private/personal income (not from business)
    - Gifts received (not in course of business)
    - Loan proceeds (not income)
    - Capital contributions

    COMMON ERROR: Coding business income as BAS Excluded when it should
    be GST on Income or GST Free Income.

    Note: Interest/dividends are handled separately (input-taxed).
    Note: Regular Sales are handled by check_sales_gst_error.

    Source: ATO GST and BAS reporting rules
    """
    account = transaction.get('account', '').lower()
    description = transaction.get('description', '').lower()
    gst_rate_name = transaction.get('gst_rate_name', '').lower()
    gross = abs(transaction.get('gross', 0))

    # Don't process Interest/Dividend accounts - handled separately (input-taxed)
    if 'interest' in account or 'dividend' in account:
        return False

    # Don't process regular Sales accounts - handled by check_sales_gst_error
    if 'sales' in account and 'other' not in account:
        return False

    # Other/Sundry/Miscellaneous income account names
    other_income_accounts = [
        'other income', 'sundry income', 'miscellaneous income',
        'misc income', 'other revenue', 'sundry revenue',
        'ancillary income', 'incidental income',
    ]

    # Income types that should generally be TAXABLE (GST on Income)
    taxable_income_types = [
        'commission', 'rebate', 'refund', 'reimbursement',
        'insurance payout', 'insurance proceeds', 'insurance claim',
        'hire income', 'rental income', 'lease income',
        'service income', 'consulting income', 'contract income',
        'management fee', 'admin fee', 'late fee', 'cancellation fee',
        'scrap sale', 'waste sale', 'by-product sale',
    ]

    is_other_income_account = any(acc in account for acc in other_income_accounts)
    is_taxable_income_type = any(inc in description for inc in taxable_income_types)

    if not is_other_income_account and not is_taxable_income_type:
        return False

    # Check for incorrect GST coding
    is_bas_excluded = (
        'bas excluded' in gst_rate_name or
        'out of scope' in gst_rate_name or
        'n-t' in gst_rate_name
    )

    # Flag if Other Income is coded as BAS Excluded (should be taxable or GST-free)
    if (is_other_income_account or is_taxable_income_type) and is_bas_excluded and gross > 0:
        return True

    return False


def check_insurance_gst_error(transaction):
    """
    Check if insurance has incorrect GST coding.

    Per ATO rules:
    - General insurance (property, motor vehicle, public liability, etc.):
      GST IS claimable on premiums
    - Life insurance, income protection, disability insurance:
      INPUT-TAXED - NO GST credit claimable

    Insurance settlements/payouts:
    - No GST if insurer was notified of GST status before claim
    - May have GST component if insurer wasn't notified

    Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/when-to-charge-gst-and-when-not-to/insurance-settlements
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_amount = abs(transaction.get('gst', 0))

    # Input-taxed insurance types - NO GST credit claimable
    input_taxed_insurance_keywords = [
        'life insurance', 'life cover', 'life policy',
        'income protection', 'income protection insurance',
        'disability insurance', 'tpd', 'total permanent disability',
        'trauma insurance', 'critical illness',
    ]

    # Check if it's an input-taxed insurance type
    is_input_taxed_insurance = any(keyword in description for keyword in input_taxed_insurance_keywords)

    # Check if it's in an insurance account
    is_insurance_account = 'insurance' in account

    # Flag if GST is claimed on input-taxed insurance
    if is_input_taxed_insurance and gst_amount > 0:
        return True

    return False


def check_life_insurance_personal(transaction):
    """
    Check if life/income protection insurance is incorrectly coded as a business expense.

    Per ATO rules:
    - Life insurance, TPD, trauma insurance are NOT deductible business expenses
    - Income protection MAY be deductible on the OWNER'S personal tax return, not business
    - If paid through the business, should be coded to Owner Drawings (personal expense)
    - Exception: Employer-paid insurance for EMPLOYEES may be deductible

    The owner can claim income protection on their personal tax return.
    Life insurance is generally not deductible at all.

    Source: https://www.ato.gov.au/individuals-and-families/your-tax-return/instructions-to-complete-your-tax-return/mytax-instructions/2025/deductions/other-deductions/other-deductions
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()

    # Life/personal insurance types - NOT business deductible
    personal_insurance_keywords = [
        'life insurance', 'life cover', 'life policy', 'life premium',
        'income protection', 'income protection insurance', 'ip insurance',
        'disability insurance', 'tpd', 'total permanent disability',
        'trauma insurance', 'critical illness', 'critical care',
        'death cover', 'death benefit',
    ]

    # Check if it's personal insurance
    is_personal_insurance = any(keyword in description for keyword in personal_insurance_keywords)

    # Skip if already in Drawings/Loan account (correctly coded)
    drawings_accounts = ['drawing', 'drawings', 'loan', 'director loan', 'shareholder', 'owner']
    is_drawings_account = any(keyword in account for keyword in drawings_accounts)
    if is_drawings_account:
        return False

    # Skip if description mentions "employee" or "staff" (may be valid business expense)
    employee_keywords = ['employee', 'staff', 'worker', 'team member']
    is_employee_insurance = any(keyword in description for keyword in employee_keywords)
    if is_employee_insurance:
        return False

    # Flag if personal insurance is in a business expense account
    if is_personal_insurance:
        return True

    return False


def check_wages_gst_error(transaction):
    """
    Check if GST is incorrectly claimed on wages/salaries.

    Per ATO rules:
    - Wages and salaries are NOT a supply - they have NO GST
    - You cannot claim GST credits on wages paid to staff
    - Wages should be coded as BAS Excluded or Out of Scope

    Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/claiming-gst-credits/when-you-can-claim-a-gst-credit
    """
    description = str(transaction.get('description', '') or '').lower()
    account = str(transaction.get('account', '') or '').lower()
    try:
        gst_amount = abs(float(transaction.get('gst', 0) or 0))
    except (ValueError, TypeError):
        gst_amount = 0

    # Wage/salary keywords
    wage_keywords = [
        'wage', 'salary', 'salaries', 'payroll', 'pay run',
        'gross pay', 'net pay', 'staff pay', 'employee pay',
        'superannuation', 'super guarantee', 'sgc',
        'payg withholding', 'payg', 'annual leave', 'sick leave',
        'long service leave', 'leave loading', 'allowance',
        'bonus', 'commission paid', 'overtime',
    ]

    # Wage account names
    is_wage_account = (
        'wage' in account or
        'salary' in account or
        'payroll' in account
    )

    is_wage_expense = (
        any(keyword in description for keyword in wage_keywords) or
        is_wage_account
    )

    # Flag if GST is claimed on wages (incorrect - wages have no GST)
    if is_wage_expense and gst_amount > 0:
        return True

    return False


def check_allowance_gst_error(transaction):
    """
    Check if GST is incorrectly claimed on employee allowances.

    Per ATO special rules for GST credits:
    - ALLOWANCES (notional/fixed amounts) - NO GST credit claimable
      Examples: cents-per-km, travel allowance, meal allowance, tool allowance
    - REIMBURSEMENTS (actual expenses with tax invoice) - GST credit IS claimable

    Key distinction:
    - Allowance: Employer pays employee a set amount (e.g., $0.85/km)
      → NOT a taxable supply, no GST applies
    - Reimbursement: Employee pays vendor, employer reimburses actual cost
      → GST credit available if employer has tax invoice

    Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/claiming-gst-credits/special-rules-for-specific-gst-credit-claims
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_amount = abs(transaction.get('gst', 0))

    # Specific allowance keywords (these are NOT reimbursements)
    allowance_keywords = [
        # Per-km allowances
        'cents per km', 'cents per kilometre', 'c/km', 'per km',
        'kilometre allowance', 'kilometer allowance', 'km allowance',
        'mileage allowance', 'car allowance', 'motor vehicle allowance',
        # Travel allowances (fixed amounts, not actual expenses)
        'travel allowance', 'travelling allowance', 'accommodation allowance',
        'overnight allowance', 'meal allowance', 'incidental allowance',
        'living away from home', 'lafha', 'per diem',
        # Tool/uniform allowances
        'tool allowance', 'uniform allowance', 'clothing allowance',
        'laundry allowance', 'dry cleaning allowance',
        # Phone/internet allowances
        'phone allowance', 'mobile allowance', 'internet allowance',
        'home office allowance', 'working from home allowance',
        # Other allowances
        'first aid allowance', 'on call allowance', 'shift allowance',
        'site allowance', 'height allowance', 'danger allowance',
    ]

    # Account names that suggest allowances
    is_allowance_account = 'allowance' in account

    is_allowance_payment = (
        any(keyword in description for keyword in allowance_keywords) or
        is_allowance_account
    )

    # Flag if GST is claimed on allowance (incorrect - allowances have no GST)
    # Allowances are payments to employees, not purchases from suppliers
    if is_allowance_payment and gst_amount > 0:
        return True

    return False


def check_personal_expense_in_business_account(transaction):
    """
    Check if a personal expense is incorrectly coded to a business expense account.

    Common pattern: Telstra bills with "personal" portion should go to:
    - Owner A Drawings (or similar) with BAS Excluded
    NOT to business expense accounts like Telephone & Internet with GST.

    This catches errors like:
    - "telstra personal" coded to Telephone & Internet (WRONG)
    - "personal use" coded to Motor Vehicle Expenses (WRONG)
    - Any description containing "personal" in a business expense account

    Personal expenses should be in Drawings/Loan accounts with no GST claimed.
    """
    description = (transaction.get('description', '') or '').lower()
    account = (transaction.get('account', '') or '').lower()

    # Keywords indicating personal expense
    personal_keywords = [
        'personal', 'private', 'private use', 'personal use',
        'owner personal', 'directors personal', 'shareholder personal'
    ]

    # Check if description indicates personal expense
    is_personal = any(keyword in description for keyword in personal_keywords)

    if not is_personal:
        return False

    # Accounts where personal expenses SHOULD go (these are OK)
    personal_accounts = [
        'drawing', 'drawings', 'owner a drawing', 'owner b drawing',
        'loan', 'director loan', 'shareholder loan',
        'private', 'personal'
    ]

    # Check if it's in an appropriate personal account
    is_in_personal_account = any(acct in account for acct in personal_accounts)

    if is_in_personal_account:
        return False  # Correctly coded to personal account

    # Personal expense in a business account - this is an error
    # Business expense accounts include most things except drawings/loans
    business_expense_indicators = [
        'expense', 'telephone', 'internet', 'motor vehicle', 'fuel',
        'office', 'supplies', 'equipment', 'utilities', 'rent',
        'advertising', 'travel', 'entertainment', 'subscriptions'
    ]

    is_business_account = any(indicator in account for indicator in business_expense_indicators)

    if is_business_account:
        return True  # ERROR: Personal expense in business account

    return False


def check_reimbursement_gst(transaction):
    """
    Check if employee reimbursement transactions have valid GST credit claims.

    Per ATO GST and employee reimbursements (TR 1999/10):
    - REIMBURSEMENTS CAN claim GST credits IF:
      1. The employer holds a tax invoice (or could have obtained one)
      2. The expense was work-related
      3. The employer is registered for GST

    TAX INVOICE REQUIREMENTS:
    - Amounts > $82.50 (incl GST): MUST have tax invoice in employer's name
    - Amounts <= $82.50: Simplified record-keeping, but still need evidence
    - The invoice must show the supplier's ABN

    COMMON ISSUES:
    - Claiming GST on expenses without tax invoice
    - Employee provides receipt but not tax invoice
    - Tax invoice in employee's name (not employer's)
    - Petty cash reimbursements without proper documentation

    This check is ADVISORY - flags reimbursements over $82.50 to verify tax invoice exists.

    Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/rules-for-specific-transactions/gst-and-employee-reimbursements
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_amount = abs(transaction.get('gst', 0))
    gross = abs(transaction.get('gross', 0))

    # Skip travel accounts - employee reimbursements for travel are normal business practice
    # Employees commonly pay for travel expenses first and get reimbursed later
    travel_account_keywords = [
        'travel', 'travel national', 'travel - national', 'domestic travel',
        'travel international', 'travel - international',
        'accommodation', 'airfare', 'flights', 'transport',
        'motor vehicle', 'fuel', 'parking', 'tolls', 'uber', 'taxi'
    ]
    if any(keyword in account for keyword in travel_account_keywords):
        return False

    # Reimbursement keywords (actual expenses paid by employee and reimbursed)
    reimbursement_keywords = [
        # Direct reimbursement terms
        'reimbursement', 'reimburse', 'reimbursed', 'repay', 'repaid',
        'expense claim', 'expense reimbursement', 'staff expense',
        'employee expense', 'petty cash', 'out of pocket',
        # Common reimbursed expense types
        'parking reimbursement', 'fuel reimbursement', 'travel reimbursement',
        'airfare reimbursement', 'accommodation reimbursement',
        'meal reimbursement', 'client entertainment reimbursement',
        'phone reimbursement', 'internet reimbursement',
        'uniform purchase', 'tool purchase', 'equipment purchase',
        # Employee paid expenses
        'paid by employee', 'employee paid', 'staff paid',
        'personal card', 'own funds',
    ]

    # Account names that suggest reimbursements
    is_reimbursement_account = (
        'reimbursement' in account or
        'reimburse' in account or
        'expense claim' in account or
        'petty cash' in account or
        'employee expense' in account
    )

    is_reimbursement = (
        any(keyword in description for keyword in reimbursement_keywords) or
        is_reimbursement_account
    )

    # Only flag significant reimbursements (> $82.50) where GST is claimed
    # $82.50 is the ATO threshold for requiring a full tax invoice
    # Below this, simplified record-keeping applies
    tax_invoice_threshold = 82.50

    if is_reimbursement and gst_amount > 0 and gross > tax_invoice_threshold:
        return True

    return False


def check_voucher_gst(transaction):
    """
    Check if gift cards and vouchers have correct GST treatment.

    Per ATO GSTR 2003/5 - GST and Vouchers:

    FACE VALUE VOUCHERS (gift cards, store credits):
    - NO GST at time of SALE (if sold at or below face value)
    - GST applies only at REDEMPTION based on goods/services purchased
    - If unredeemed/expired: 1/11th GST adjustment required
    - Examples: Retail gift cards, Visa/Mastercard gift cards, store credits

    NON-FACE VALUE VOUCHERS (specific goods/services):
    - GST applies at time of SALE
    - No additional GST at redemption
    - Examples: Spa voucher for specific treatment, restaurant meal voucher

    COMMON ERRORS:
    1. Charging GST on face value voucher SALES (incorrect - defer until redemption)
    2. NOT charging GST on non-face value voucher SALES (incorrect - GST applies at sale)
    3. Forgetting 1/11th adjustment on unredeemed face value vouchers

    GST CREDIT TIMING:
    - Face value vouchers: Claim credit when REDEEMED
    - Non-face value vouchers: Claim credit when PURCHASED

    Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/rules-for-specific-transactions/gst-and-vouchers
    GSTR 2003/5
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_amount = abs(transaction.get('gst', 0))
    gross = abs(transaction.get('gross', 0))

    # Face value voucher keywords (broad redemption - NO GST at sale)
    face_value_keywords = [
        # Gift cards
        'gift card', 'giftcard', 'gift voucher', 'store credit',
        'gift certificate', 'shopping voucher', 'retail voucher',
        # Prepaid cards
        'prepaid card', 'prepaid voucher', 'visa gift', 'mastercard gift',
        'eftpos gift', 'amex gift',
        # General vouchers sold
        'voucher sale', 'sold voucher', 'voucher issued',
        'gift card sale', 'gift card sold',
    ]

    # Non-face value voucher keywords (specific goods/services - GST at sale)
    non_face_value_keywords = [
        # Specific service vouchers
        'spa voucher', 'massage voucher', 'facial voucher',
        'treatment voucher', 'service voucher',
        'meal voucher', 'dinner voucher', 'restaurant voucher',
        'experience voucher', 'activity voucher',
        'class voucher', 'lesson voucher', 'course voucher',
        # Package deals
        'package voucher', 'bundle voucher',
    ]

    # Account names that suggest voucher transactions
    is_voucher_account = (
        'voucher' in account or
        'gift card' in account or
        'giftcard' in account or
        'gift certificate' in account
    )

    is_face_value_voucher = any(keyword in description for keyword in face_value_keywords)
    is_non_face_value_voucher = any(keyword in description for keyword in non_face_value_keywords)

    # Check if it's an income/sales transaction (voucher being SOLD)
    is_income = gross > 0 and (
        'income' in account or
        'sales' in account or
        'revenue' in account or
        'liability' in account  # Gift cards often go to liability account
    )

    # Flag 1: Face value voucher SALE with GST (incorrect - should be no GST at sale)
    if is_face_value_voucher and is_income and gst_amount > 0:
        return 'face_value_with_gst'

    # Flag 2: Check for voucher income that might need review
    # This is advisory - to remind about face value vs non-face value distinction
    if is_voucher_account and is_income and gst_amount > 0:
        # Could be face value voucher incorrectly charging GST
        return 'voucher_gst_review'

    return False


def check_general_expenses(transaction):
    """
    Check if transaction is coded to General/Sundry/Miscellaneous expenses.

    AUDIT RISK AND BEST PRACTICE ISSUES:

    1. ATO AUDIT RED FLAG:
       - Tax authorities view large sundry/general accounts as lacking transparency
       - Indicates potential poor record-keeping practices
       - Makes substantiation of deductions more difficult

    2. GST TREATMENT RISK:
       - General expense accounts often contain mixed GST items
       - May include GST-free, input-taxed, or non-deductible items
       - Harder to verify correct GST treatment

    3. BEST PRACTICE THRESHOLD:
       - If sundry/general expenses exceed 5% of total expenses, triggers concern
       - Indicates failure to properly categorize costs
       - Poor internal controls

    4. SUBSTANTIATION ISSUES:
       - ATO requires written evidence for claims over $300
       - General categories make it harder to prove business purpose
       - Higher risk of disallowed deductions in audit

    RECOMMENDATION:
    Recode transactions to specific expense accounts:
    - Motor Vehicle Expenses
    - Office Supplies
    - Travel Expenses
    - Communication/Telephone
    - Professional Fees
    - etc.

    Source: ATO substantiation requirements, audit best practices
    """
    account = transaction.get('account', '').lower()
    account_code = str(transaction.get('account_code', '')).lower()

    # General/sundry/miscellaneous expense account indicators
    general_expense_indicators = [
        # Direct matches
        'general expense', 'general expenses',
        'sundry expense', 'sundry expenses',
        'miscellaneous expense', 'miscellaneous expenses',
        'misc expense', 'misc expenses',
        'other expense', 'other expenses',
        'sundries', 'miscellaneous',
        # Common variations
        'general operating', 'general business',
        'unclassified expense', 'uncategorized expense',
        'unallocated expense', 'suspense expense',
    ]

    is_general_expense = any(indicator in account for indicator in general_expense_indicators)

    # Also check for account codes that might indicate general expenses
    # Common patterns: accounts ending in 99, 00, or named "other"
    if 'other' in account and 'expense' in account:
        is_general_expense = True

    if is_general_expense:
        return True

    return False


def check_travel_gst(transaction):
    """
    Check if travel expenses have correct GST treatment.

    Per ATO GST and international travel rules:

    DOMESTIC TRAVEL (within Australia) - TAXABLE:
    - Domestic flights: GST applies (10%) - CAN claim GST credits
    - Accommodation (hotels, motels): GST applies - CAN claim GST credits
    - Car hire: GST applies - CAN claim GST credits
    - Taxis, rideshare, Uber: GST applies - CAN claim GST credits
    - Meals during travel: GST applies - CAN claim GST credits
    - Parking: GST applies - CAN claim GST credits

    INTERNATIONAL TRAVEL - GST-FREE:
    - International flights: GST-FREE - NO GST credits available
    - Domestic legs of international journey: GST-FREE (if booked together)
    - Overseas accommodation: NO Australian GST - NO credits
    - Overseas car hire: NO Australian GST - NO credits
    - Overseas meals/tours: NO Australian GST - NO credits

    COMMON ERRORS:
    1. Claiming GST credits on international flights (incorrect - GST-free)
    2. Claiming GST credits on overseas expenses (no Australian GST)
    3. NOT claiming GST on domestic travel (missing credits)
    4. Treating domestic flights as GST-free when not part of international journey

    Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/your-industry/travel-and-tourism/gst-and-international-travel
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_amount = abs(transaction.get('gst', 0))
    gst_rate_name = transaction.get('gst_rate_name', '').lower()
    gross = abs(transaction.get('gross', 0))

    # International travel keywords (should be GST-FREE, no GST credits)
    international_keywords = [
        # International flights
        'international flight', 'overseas flight', 'international airfare',
        'flight to usa', 'flight to uk', 'flight to europe', 'flight to asia',
        'flight to nz', 'flight to new zealand', 'flight to singapore',
        'flight to bali', 'flight to fiji', 'flight to japan', 'flight to china',
        'los angeles', 'london', 'new york', 'singapore', 'hong kong', 'tokyo',
        'auckland', 'wellington', 'christchurch', 'denpasar', 'bali',
        # Overseas accommodation
        'overseas hotel', 'overseas accommodation', 'international hotel',
        'hotel london', 'hotel singapore', 'hotel usa', 'hotel nz',
        'overseas accom', 'foreign hotel', 'international accom',
        # Overseas expenses
        'overseas expense', 'international expense', 'foreign expense',
        'overseas car hire', 'international car rental',
        'overseas meal', 'overseas tour', 'international tour',
    ]

    # Domestic travel keywords (should be TAXABLE with GST)
    domestic_keywords = [
        # Domestic flights
        'domestic flight', 'qantas', 'virgin australia', 'jetstar', 'rex airlines',
        'flight to sydney', 'flight to melbourne', 'flight to brisbane',
        'flight to perth', 'flight to adelaide', 'flight to hobart',
        'flight to darwin', 'flight to cairns', 'flight to gold coast',
        'sydney', 'melbourne', 'brisbane', 'perth', 'adelaide', 'hobart',
        'darwin', 'cairns', 'canberra', 'gold coast', 'newcastle',
        # Domestic accommodation
        'hotel sydney', 'hotel melbourne', 'hotel brisbane', 'motel',
        'accommodation sydney', 'accommodation melbourne', 'airbnb',
        # Domestic transport
        'taxi', 'uber', 'didi', 'ola', 'rideshare', 'car hire', 'car rental',
        'hertz', 'avis', 'budget rent', 'thrifty', 'europcar',
        # Travel account indicators
        'travel expense', 'business travel', 'travel cost',
    ]

    # Check if it's a travel-related account
    is_travel_account = (
        'travel' in account or
        'airfare' in account or
        'flight' in account or
        'accommodation' in account or
        'hotel' in account
    )

    is_international = any(keyword in description for keyword in international_keywords)
    is_domestic = any(keyword in description for keyword in domestic_keywords)

    # Check GST coding
    has_gst = gst_amount > 0
    is_gst_free = 'free' in gst_rate_name or gst_amount == 0

    # Flag 1: International travel WITH GST claimed (incorrect - should be GST-free)
    if is_international and has_gst:
        return 'international_with_gst'

    # Flag 2: Domestic travel WITHOUT GST (incorrect - should have GST)
    # Only flag if it's clearly domestic and coded as GST-free
    if is_domestic and not is_international and is_gst_free and gross > 50:
        # Check if it might be part of international journey
        if 'international' not in description and 'overseas' not in description:
            return 'domestic_no_gst'

    # Flag 3: Travel expense that needs review (ambiguous)
    if is_travel_account and gross > 500:
        if has_gst and is_international:
            return 'international_with_gst'

    return False


def check_payment_processor_fees(transaction):
    """
    Check if payment processor fees have correct GST treatment.

    Per ATO rules on financial supplies and merchant fees:

    PAYPAL FEES - GST EXEMPT (Input-Taxed Financial Supply):
    - PayPal transaction fees are GST EXEMPT
    - NO GST credit can be claimed
    - Exception: Virtual Terminal, PayPal Payments Pro (these include GST)
    - Source: PayPal Australia PDS - "our fees are GST exempt"

    EBAY FEES - GST INCLUDED (unless ABN exemption):
    - eBay charges GST on seller fees (final value, insertion, subscriptions)
    - If GST-registered with ABN tax exemption: fees charged NET of GST
    - If no ABN exemption: GST included, CAN claim credits
    - Source: eBay Australia Tax Policy

    STRIPE FEES - GST INCLUDED (10%):
    - All Stripe fees include GST at 10%
    - GST credits CAN be claimed
    - Source: Stripe Australia pricing

    BANK MERCHANT FEES - GST INCLUDED:
    - Bank/EFTPOS merchant fees generally include GST
    - GST credits CAN be claimed

    COMMON ERRORS:
    1. Claiming GST on PayPal fees (incorrect - PayPal is GST exempt)
    2. Not claiming GST on Stripe/eBay fees (missing credits)
    3. Treating all merchant fees the same (they have different GST treatment)

    Source: ATO financial supplies rules, PayPal PDS, eBay Tax Policy
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_amount = abs(transaction.get('gst', 0))
    gross = abs(transaction.get('gross', 0))

    # PayPal fee keywords (GST EXEMPT - no GST credit)
    paypal_keywords = [
        'paypal fee', 'paypal fees', 'paypal charge', 'paypal charges',
        'paypal transaction', 'paypal merchant', 'paypal cost',
        'pp fee', 'pp fees', 'paypal commission',
    ]

    # Stripe fee keywords (GST INCLUDED - can claim)
    stripe_keywords = [
        'stripe fee', 'stripe fees', 'stripe charge', 'stripe charges',
        'stripe transaction', 'stripe merchant', 'stripe cost',
        'stripe commission', 'stripe processing',
    ]

    # eBay fee keywords (GST INCLUDED unless ABN exempt)
    ebay_keywords = [
        'ebay fee', 'ebay fees', 'ebay charge', 'ebay charges',
        'ebay final value', 'ebay insertion', 'ebay store',
        'ebay seller fee', 'ebay commission', 'ebay listing',
    ]

    # Generic merchant fee keywords (usually GST INCLUDED)
    merchant_keywords = [
        'merchant fee', 'merchant fees', 'eftpos fee', 'eftpos fees',
        'card processing', 'card fee', 'credit card fee',
        'transaction fee', 'processing fee', 'payment processing',
        'square fee', 'square fees', 'afterpay fee', 'afterpay fees',
        'zippay fee', 'tyro fee', 'tyro fees',
    ]

    is_paypal = any(keyword in description for keyword in paypal_keywords)
    is_stripe = any(keyword in description for keyword in stripe_keywords)
    is_ebay = any(keyword in description for keyword in ebay_keywords)
    is_merchant = any(keyword in description for keyword in merchant_keywords)

    # Check account names too
    if 'paypal' in account:
        is_paypal = True
    if 'stripe' in account:
        is_stripe = True
    if 'ebay' in account:
        is_ebay = True
    if 'merchant' in account or 'eftpos' in account:
        is_merchant = True

    # Flag 1: PayPal fees WITH GST claimed (incorrect - PayPal is GST exempt)
    if is_paypal and gst_amount > 0:
        return 'paypal_with_gst'

    # Flag 2: Stripe fees WITHOUT GST (incorrect - Stripe includes GST)
    if is_stripe and gst_amount == 0 and gross > 0:
        return 'stripe_no_gst'

    # Flag 3: eBay fees - advisory to check ABN exemption status
    if is_ebay and gst_amount == 0 and gross > 10:
        return 'ebay_check_gst'

    # Flag 4: Bank merchant fees WITHOUT GST (likely incorrect)
    if is_merchant and not is_paypal and gst_amount == 0 and gross > 5:
        return 'merchant_no_gst'

    return False


def check_fines_penalties_gst(transaction):
    """
    Check if GST is incorrectly claimed on fines, penalties, and government charges.

    Per ATO Simpler BAS bookkeeping guide:
    - Council fines - non-reportable GST
    - ATO late payment penalties - non-reportable GST
    - ATO General Interest Charge (GIC) - non-reportable GST (but tax deductible)
    - Superannuation late payment charges - non-reportable GST
    - Traffic/parking fines - non-reportable GST
    - Court fines - non-reportable GST

    These items have NO GST regardless of whether they are tax deductible.

    Source: https://www.ato.gov.au/businesses-and-organisations/preparing-lodging-and-paying/business-activity-statements-bas/goods-and-services-tax-gst/simpler-bas-gst-bookkeeping-guide
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_amount = abs(transaction.get('gst', 0))

    # Fines and penalties keywords
    fines_penalties_keywords = [
        # Government fines
        'fine', 'penalty', 'infringement', 'breach',
        'traffic fine', 'parking fine', 'speeding fine',
        'council fine', 'epa fine', 'environmental fine',
        'court fine', 'court penalty',
        # ATO charges
        'ato penalty', 'ato fine', 'late lodgement penalty',
        'late payment penalty', 'failure to lodge',
        'general interest charge', 'gic', 'shortfall interest',
        # Superannuation penalties
        'super guarantee charge', 'sgc penalty', 'super late fee',
        'superannuation penalty', 'super interest charge',
        # Other penalties
        'late fee', 'overdue fee', 'default fee',
        'dishonour fee', 'returned payment fee',
    ]

    # Account names that suggest fines/penalties
    is_fines_account = (
        'fine' in account or
        'penalty' in account or
        'penalt' in account
    )

    is_fine_or_penalty = (
        any(keyword in description for keyword in fines_penalties_keywords) or
        is_fines_account
    )

    # Flag if GST is claimed on fines/penalties (incorrect - no GST applies)
    if is_fine_or_penalty and gst_amount > 0:
        return True

    return False


def check_donations_gst(transaction):
    """
    Check if GST is incorrectly claimed on donations.

    Per ATO Simpler BAS bookkeeping guide:
    - ALL donations are non-reportable GST regardless of DGR status
    - Donations to DGRs are tax deductible but still no GST
    - Donations to non-DGRs are not tax deductible and no GST
    - Sponsorship is DIFFERENT - it's taxable (you get advertising in return)

    A donation is a voluntary gift with no expectation of return.
    If you receive something in return, it may be sponsorship (GST applies).

    Source: https://www.ato.gov.au/businesses-and-organisations/preparing-lodging-and-paying/business-activity-statements-bas/goods-and-services-tax-gst/simpler-bas-gst-bookkeeping-guide
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_amount = abs(transaction.get('gst', 0))

    # Donation keywords
    donation_keywords = [
        'donation', 'donate', 'donated',
        'charitable contribution', 'charity',
        'gift to', 'gifted to',
        'philanthropic', 'benevolent',
        'appeal', 'fundraiser', 'fundraising',
        'red cross', 'salvation army', 'cancer council',
        'heart foundation', 'smith family', 'oxfam',
        'world vision', 'unicef', 'msf', 'doctors without borders',
    ]

    # Account names that suggest donations
    is_donation_account = (
        'donation' in account or
        'charit' in account
    )

    is_donation = (
        any(keyword in description for keyword in donation_keywords) or
        is_donation_account
    )

    # Exclude if it looks like sponsorship (which IS taxable)
    sponsorship_indicators = ['sponsor', 'advertising', 'promotion', 'naming rights']
    is_sponsorship = any(indicator in description for indicator in sponsorship_indicators)

    # Flag if GST is claimed on donation (incorrect - no GST applies)
    if is_donation and not is_sponsorship and gst_amount > 0:
        return True

    return False


def check_property_gst_withholding(transaction):
    """
    Check if a property purchase may be subject to GST withholding at settlement.

    Per ATO GST at settlement rules (effective 1 July 2018):
    - Buyers of NEW residential premises or POTENTIAL RESIDENTIAL LAND must
      withhold GST and pay directly to ATO at settlement
    - Withholding amount: 1/11th of price (or 7% if margin scheme applies)
    - Seller must notify buyer of withholding obligation before settlement
    - Buyer must lodge Form 1 and Form 2 with ATO

    NEW RESIDENTIAL PREMISES:
    - Property that has not been previously sold as residential
    - Property not rented continuously for 5+ years
    - Includes newly constructed homes, apartments, townhouses

    POTENTIAL RESIDENTIAL LAND:
    - Land that can be used for residential purposes
    - Land in a subdivision plan zoned for residential
    - Does not contain buildings in commercial use

    EXCLUSIONS:
    - Substantial renovations (not new build)
    - Commercial residential (hotels, boarding houses)
    - Previously sold residential premises
    - Land with commercial buildings

    Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/your-industry/property/gst-at-settlement
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gross = abs(transaction.get('gross', 0))
    gst_amount = abs(transaction.get('gst', 0))

    # Property purchase keywords that may trigger GST withholding
    property_purchase_keywords = [
        # New residential premises
        'new home', 'new house', 'new apartment', 'new unit', 'new townhouse',
        'new villa', 'new duplex', 'off the plan', 'off-the-plan',
        'house and land', 'house & land', 'land and house',
        'newly built', 'new construction', 'new build',
        # Property settlement
        'property settlement', 'settlement', 'conveyancing',
        'property purchase', 'land purchase', 'land acquisition',
        # Subdivision/development
        'subdivision', 'vacant land', 'residential land', 'building block',
        'lot purchase', 'land lot', 'development site',
        # Developer purchases
        'developer', 'property developer', 'real estate development',
    ]

    # Account names that suggest property/asset purchases
    is_property_account = (
        'property' in account or
        'land' in account or
        'building' in account or
        'real estate' in account or
        'asset' in account
    )

    is_property_purchase = (
        any(keyword in description for keyword in property_purchase_keywords) or
        is_property_account
    )

    # Only flag significant property purchases (typically > $100,000)
    # with GST that may be subject to withholding rules
    if is_property_purchase and gross > 100000 and gst_amount > 0:
        return True

    return False


def check_livestock_gst(transaction):
    """
    Check if livestock and game sales have correct GST treatment.

    Per ATO GST on livestock and game sales:
    - LIVE ANIMALS (livestock/game) → TAXABLE (GST applies)
    - MEAT FOR HUMAN CONSUMPTION → GST-FREE (once inspected/passed)

    Key rules:
    - Sales of livestock to processors are TAXABLE
    - Sales of livestock to other producers are TAXABLE (if GST-registered)
    - Auction sales of livestock are TAXABLE (if GST-registered)
    - Meat becomes GST-free only after inspection and passing for human consumption

    The GST definition of food does NOT include live animals
    (except crustaceans and molluscs).

    COMMON ERROR: Coding livestock sales as GST-free when they should be taxable.

    Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/your-industry/gst-on-livestock-and-game-sales
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_rate_name = transaction.get('gst_rate_name', '').lower()
    gross = abs(transaction.get('gross', 0))
    gst_amount = abs(transaction.get('gst', 0))

    # Livestock and game keywords
    livestock_keywords = [
        # Cattle
        'cattle', 'cow', 'cows', 'bull', 'bulls', 'calf', 'calves',
        'heifer', 'heifers', 'steer', 'steers', 'beef cattle',
        # Sheep
        'sheep', 'lamb', 'lambs', 'ewe', 'ewes', 'ram', 'rams',
        'wether', 'wethers', 'mutton',
        # Pigs
        'pig', 'pigs', 'piglet', 'piglets', 'swine', 'sow', 'boar', 'pork',
        # Poultry
        'chicken', 'chickens', 'poultry', 'hen', 'hens', 'rooster',
        'turkey', 'turkeys', 'duck', 'ducks', 'goose', 'geese',
        # Other livestock
        'goat', 'goats', 'alpaca', 'alpacas', 'llama', 'llamas',
        'horse', 'horses', 'donkey', 'donkeys',
        # Game
        'deer', 'venison', 'kangaroo', 'wallaby', 'emu', 'ostrich',
        'rabbit', 'rabbits', 'wild boar', 'game meat', 'game animal',
        # General terms
        'livestock', 'stock sale', 'animal sale', 'saleyard',
        'auction sale', 'head of cattle', 'head of sheep',
    ]

    # Account names that suggest livestock
    is_livestock_account = (
        'livestock' in account or
        'cattle' in account or
        'sheep' in account or
        'stock sale' in account
    )

    is_livestock_sale = (
        any(keyword in description for keyword in livestock_keywords) or
        is_livestock_account
    )

    if not is_livestock_sale:
        return False

    # Check if it's an income/sales transaction (not a purchase)
    is_income = gross > 0 and (
        'income' in account or
        'sales' in account or
        'revenue' in account
    )

    # Flag if livestock sale is coded as GST-free when it should be taxable
    # Live animal sales should have GST (unless hobby farmer)
    is_gst_free = (
        'free' in gst_rate_name or
        'exempt' in gst_rate_name or
        gst_amount == 0
    )

    if is_livestock_sale and is_income and is_gst_free:
        return True

    return False


def check_asset_disposal_gst(transaction):
    """
    Check if disposal of capital assets has correct GST treatment.

    Per ATO GST and disposal of capital assets:
    - Business asset disposals are generally TAXABLE (GST applies)
    - Must report sale proceeds at G1 (total sales) on BAS
    - Must remit GST (1/11th of sale price) at 1A

    COMMON ERROR: Coding asset sales as BAS Excluded or GST-free
    when they should be GST on Income (taxable).

    Assets that are TAXABLE when sold:
    - Office equipment, computers, furniture
    - Motor vehicles (business use)
    - Machinery, plant and equipment
    - Tools and trade equipment
    - Trade-ins of business assets

    EXCEPTIONS (GST may NOT apply):
    - Non-business/private assets (no GST credit claimed on purchase)
    - GST-free going concern (entire business sale)
    - Farmland (used for farming 5+ years, buyer continues farming)
    - Residential premises (input-taxed)

    Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/rules-for-specific-transactions/business-asset-transactions/gst-and-the-disposal-of-capital-assets
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_rate_name = transaction.get('gst_rate_name', '').lower()
    gross = abs(transaction.get('gross', 0))
    gst_amount = abs(transaction.get('gst', 0))

    # Asset disposal keywords
    asset_disposal_keywords = [
        # Sale/disposal terms
        'sale of asset', 'asset sale', 'asset disposal', 'disposed asset',
        'sold asset', 'disposal of', 'trade in', 'trade-in', 'traded in',
        # Equipment
        'sold equipment', 'equipment sale', 'sale of equipment',
        'sold machinery', 'machinery sale', 'plant sale',
        'sold computer', 'computer sale', 'sold laptop',
        'sold furniture', 'furniture sale', 'office furniture sale',
        # Vehicles
        'sold vehicle', 'vehicle sale', 'car sale', 'sold car',
        'truck sale', 'sold truck', 'ute sale', 'sold ute',
        'van sale', 'sold van', 'motor vehicle sale',
        # Other assets
        'sold tools', 'tool sale', 'equipment disposal',
        'fixed asset sale', 'capital asset sale',
    ]

    # Account names that suggest asset disposals
    is_asset_disposal_account = (
        'asset sale' in account or
        'disposal' in account or
        'gain on sale' in account or
        'loss on sale' in account or
        'profit on sale' in account or
        'sale of asset' in account or
        'proceeds' in account
    )

    is_asset_disposal = (
        any(keyword in description for keyword in asset_disposal_keywords) or
        is_asset_disposal_account
    )

    if not is_asset_disposal:
        return False

    # Check if it's an income transaction
    is_income = gross > 0

    # Check if incorrectly coded as BAS Excluded or GST-free
    is_bas_excluded_or_gst_free = (
        'excluded' in gst_rate_name or
        'bas excluded' in gst_rate_name or
        'free' in gst_rate_name or
        'n-t' in gst_rate_name or
        gst_amount == 0
    )

    # Exclude if it looks like farmland or going concern (legitimate GST-free)
    exemption_indicators = [
        'going concern', 'farmland', 'farm sale', 'rural property',
        'private sale', 'private asset', 'personal use',
    ]
    is_exempt = any(indicator in description for indicator in exemption_indicators)

    # Flag if asset disposal is coded without GST when it should be taxable
    if is_asset_disposal and is_income and is_bas_excluded_or_gst_free and not is_exempt:
        return True

    return False


def check_grants_sponsorship_gst(transaction):
    """
    Check if grants and sponsorship income have correct GST coding.

    Per ATO rules:
    GRANTS (typically GST-FREE):
    - No GST if you only meet eligibility criteria (no supply in return)
    - GST applies ONLY if you make a supply in return (binding legal obligation)
    - Government income support payments - typically NOT subject to GST

    SPONSORSHIP (TAXABLE - GST applies):
    - GST applies because you provide something in return
      (advertising, signage, naming rights, promotion)
    - Sponsors can claim GST credits

    Source: https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/when-to-charge-gst-and-when-not-to/grants-and-sponsorship
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gst_amount = abs(transaction.get('gst', 0))
    gst_rate_name = transaction.get('gst_rate_name', '').lower()
    gross_amount = abs(transaction.get('gross', 0))

    # Check if it's income (positive gross or income account)
    is_income_account = (
        'income' in account or
        'revenue' in account or
        'sales' in account
    )

    # Grant keywords - typically GST-free (unless supply in return)
    grant_keywords = [
        'grant', 'government grant', 'funding', 'subsidy',
        'jobkeeper', 'cash flow boost', 'stimulus', 'covid support',
        'business support', 'apprentice subsidy', 'wage subsidy',
        'export grant', 'r&d grant', 'innovation grant',
    ]

    # Sponsorship keywords - TAXABLE (GST applies)
    sponsorship_keywords = [
        'sponsorship', 'sponsor', 'naming rights', 'advertising rights',
        'promotional rights', 'event sponsor', 'corporate sponsor',
    ]

    is_grant_income = any(keyword in description for keyword in grant_keywords)
    is_sponsorship_income = any(keyword in description for keyword in sponsorship_keywords)

    # Check GST coding
    is_gst_free = 'free' in gst_rate_name or gst_amount == 0
    has_gst = gst_amount > 0 or ('gst on' in gst_rate_name and 'free' not in gst_rate_name)

    # Flag 1: Sponsorship income without GST (should be taxable)
    if is_sponsorship_income and is_gst_free and gross_amount > 0:
        return 'sponsorship_no_gst'

    # Flag 2: Grant income with GST (usually should be GST-free, flag for review)
    # Note: Some grants ARE taxable if supply is made, so this is advisory
    if is_grant_income and has_gst and gross_amount > 0:
        return 'grant_with_gst'

    return False


def check_borrowing_expenses_error(transaction):
    """
    Check if borrowing expenses over $100 are incorrectly fully expensed.
    Per ATO rules (ITAA 1997 s25.25):
    - Borrowing expenses <= $100: Fully deductible in year incurred
    - Borrowing expenses > $100: Must be spread over 5 years (or loan term if shorter)

    Borrowing expenses include:
    - Loan establishment fees
    - Title search fees
    - Stamp duty on mortgage
    - Mortgage broker fees
    - Valuation fees for loan approval
    - Lender's mortgage insurance (LMI)

    Note: Most borrowing expenses are GST Free (financial supplies are input-taxed)

    Source: https://www.ato.gov.au/individuals-and-families/investments-and-assets/property-and-land/residential-rental-properties/rental-expenses/borrowing-expenses
    """
    description = transaction.get('description', '').lower()
    account = transaction.get('account', '').lower()
    gross_amount = abs(transaction.get('gross', 0))

    # Borrowing expense keywords
    borrowing_keywords = [
        'loan establishment', 'establishment fee', 'loan fee', 'application fee',
        'mortgage fee', 'mortgage broker', 'broker fee', 'brokerage',
        'title search', 'valuation fee', 'valuation for loan', 'bank valuation',
        'lmi', 'lender mortgage insurance', 'mortgage insurance',
        'loan documentation', 'mortgage documentation', 'mortgage stamp duty',
        'loan stamp duty', 'discharge fee', 'settlement fee',
        'borrowing cost', 'borrowing expense', 'loan cost',
    ]

    # Account names that suggest borrowing expenses
    borrowing_accounts = [
        'borrowing', 'loan cost', 'loan expense', 'financing cost',
        'bank charges', 'finance charge',
    ]

    is_borrowing_expense = (
        any(keyword in description for keyword in borrowing_keywords) or
        any(keyword in account for keyword in borrowing_accounts)
    )

    if not is_borrowing_expense:
        return False

    # ATO threshold: $100
    # If borrowing expenses > $100, they should be capitalized and spread over 5 years
    threshold = 100

    # Check if it's coded to an expense account (not a prepayment/asset account)
    is_expensed = (
        'expense' in account or
        'cost' in account or
        'fee' in account or
        'charge' in account
    )
    is_capitalized = (
        'prepaid' in account or
        'prepayment' in account or
        'borrowing cost asset' in account or
        'deferred' in account
    )

    # Flag if borrowing expense > $100 is fully expensed instead of capitalized
    if is_borrowing_expense and gross_amount > threshold and is_expensed and not is_capitalized:
        return True

    return False


# =============================================================================
# ATO RULING SEARCH WITH CACHING
# =============================================================================

# Cache for ATO rulings - persists for the session
_ato_ruling_cache = {}

# Mapping of issue types to ATO search queries and fallback info
ATO_RULING_QUERIES = {
    'export_gst_error': {
        'query': 'GST exports GST-free Division 38 site:ato.gov.au',
        'fallback': {
            'ruling': 'GSTR 2002/6',
            'title': 'When is a supply of goods or services GST-free under Division 38?',
            'summary': 'Exports are GST-free under Division 38 of the GST Act. No GST should be charged on exported goods/services.',
            'url': 'https://www.ato.gov.au/law/view/document?DocID=GST/GSTR20026/NAT/ATO/00001'
        }
    },
    'entertainment': {
        'query': 'entertainment expenses GST FBT TR 97/17 site:ato.gov.au',
        'fallback': {
            'ruling': 'TR 97/17 & GSTR 2001/6',
            'title': 'Entertainment expenses - income tax and GST treatment',
            'summary': 'Entertainment is non-deductible and GST credits cannot be claimed unless FBT is paid on the benefit.',
            'url': 'https://www.ato.gov.au/law/view/document?DocID=TXR/TR9717/NAT/ATO/00001'
        }
    },
    'wages_gst_error': {
        'query': 'wages salaries GST BAS excluded site:ato.gov.au',
        'fallback': {
            'ruling': 'GST Act - Division 9',
            'title': 'Wages and salaries are not taxable supplies',
            'summary': 'Wages/salaries/super are outside the GST system (BAS Excluded). They are not taxable supplies and should have $0 GST.',
            'url': 'https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/what-is-gst'
        }
    },
    'government_charges_gst': {
        'query': 'council rates stamp duty GST Division 81 government charges site:ato.gov.au',
        'fallback': {
            'ruling': 'GST Act - Division 81',
            'title': 'Payments to government agencies - taxes, fees and charges',
            'summary': 'Council rates, stamp duty, land tax are government TAXES under Division 81 - NOT subject to GST. These are not taxable supplies. GST should be $0.',
            'url': 'https://www.ato.gov.au/businesses-and-organisations/corporate-tax-measures-and-assurance/government-entities/gst-for-government/payments-to-government-agencies-under-division-81'
        }
    },
    'grants_sponsorship_gst': {
        'query': 'grants GST-free GSTR 2012/2 site:ato.gov.au',
        'fallback': {
            'ruling': 'GSTR 2012/2',
            'title': 'Goods and services tax: government grants',
            'summary': 'Grants are typically GST-free unless there is a binding obligation to provide specific goods/services in return.',
            'url': 'https://www.ato.gov.au/law/view/document?DocID=GST/GSTR20122/NAT/ATO/00001'
        }
    },
    'residential_premises_gst': {
        'query': 'residential property input taxed GST site:ato.gov.au',
        'fallback': {
            'ruling': 'GST Act - Division 40',
            'title': 'Residential premises are input taxed',
            'summary': 'Residential rent and related expenses are input-taxed. No GST is charged and no GST credits can be claimed.',
            'url': 'https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/your-industry/property/gst-and-residential-property'
        }
    },
    'input_taxed_gst_error': {
        'query': 'input taxed financial supplies GST site:ato.gov.au',
        'fallback': {
            'ruling': 'GST Act - Division 40',
            'title': 'Input taxed financial supplies',
            'summary': 'Financial supplies (interest, bank fees, residential rent) are input-taxed. No GST credits can be claimed.',
            'url': 'https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/your-industry/financial-services-and-insurance/gst-and-financial-supplies'
        }
    },
    'insurance_gst_error': {
        'query': 'life insurance income protection input taxed GST site:ato.gov.au',
        'fallback': {
            'ruling': 'GST Act - Division 40',
            'title': 'Life insurance and income protection',
            'summary': 'Life insurance and income protection insurance are input-taxed financial supplies. No GST credits can be claimed.',
            'url': 'https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/your-industry/financial-services-and-insurance/gst-and-insurance'
        }
    },
    'travel_gst': {
        'query': 'international travel GST-free domestic taxable site:ato.gov.au',
        'fallback': {
            'ruling': 'GST Act - Division 38',
            'title': 'Travel and GST treatment',
            'summary': 'International travel is GST-free (no GST credits). Domestic travel within Australia is taxable (GST credits can be claimed).',
            'url': 'https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/your-industry/travel-and-tourism'
        }
    },
    'travel_gst_international': {
        'query': 'international flights GST-free Division 38 site:ato.gov.au',
        'fallback': {
            'ruling': 'GST Act - Division 38',
            'title': 'International travel is GST-free',
            'summary': 'International flights and travel are GST-free. No GST should be charged and no GST credits can be claimed.',
            'url': 'https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/your-industry/travel-and-tourism/gst-and-international-travel'
        }
    },
    'travel_gst_domestic': {
        'query': 'domestic travel GST taxable site:ato.gov.au',
        'fallback': {
            'ruling': 'GST Act - Section 9-5',
            'title': 'Domestic travel is taxable',
            'summary': 'Domestic travel within Australia is taxable. GST (10%) applies and GST credits can be claimed.',
            'url': 'https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/your-industry/travel-and-tourism'
        }
    },
    'fines_penalties_gst': {
        'query': 'fines penalties BAS excluded GST site:ato.gov.au',
        'fallback': {
            'ruling': 'GST Act - Section 9-10',
            'title': 'Fines and penalties are not taxable supplies',
            'summary': 'Fines and penalties are not consideration for a supply and are BAS Excluded. No GST applies.',
            'url': 'https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/what-is-gst'
        }
    },
    'asset_capitalization_error': {
        'query': 'instant asset write-off threshold $20000 site:ato.gov.au',
        'fallback': {
            'ruling': 'Division 328 ITAA 1997',
            'title': 'Instant asset write-off',
            'summary': 'Assets over the instant asset write-off threshold should be capitalized and depreciated, not expensed immediately.',
            'url': 'https://www.ato.gov.au/businesses-and-organisations/income-deductions-and-concessions/depreciation-and-capital-allowances/simpler-depreciation-for-small-business/instant-asset-write-off'
        }
    },
    'donations_gst': {
        'query': 'donations gifts GST-free site:ato.gov.au',
        'fallback': {
            'ruling': 'GST Act - Section 9-15',
            'title': 'Donations and gifts - GST treatment',
            'summary': 'Donations (gifts) are not consideration for a supply. No GST applies. Use GST Free for expenses.',
            'url': 'https://www.ato.gov.au/businesses-and-organisations/not-for-profit-organisations/gst/gst-for-charities-and-gift-deductible-entities'
        }
    },
    'paypal_fees': {
        'query': 'PayPal fees GST financial supply input taxed site:ato.gov.au',
        'fallback': {
            'ruling': 'GST Act - Division 40',
            'title': 'PayPal fees are input-taxed financial supplies',
            'summary': 'PayPal (Singapore) fees are GST-exempt financial supplies. No GST is charged and no GST credits can be claimed. Code as Input Taxed.',
            'url': 'https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/your-industry/financial-services-and-insurance/gst-and-financial-supplies'
        }
    },
    'merchant_fees': {
        'query': 'merchant fees EFTPOS credit card GST taxable site:ato.gov.au',
        'fallback': {
            'ruling': 'ATO Financial Services Q&A',
            'title': 'Merchant fees are taxable supplies',
            'summary': 'Bank merchant fees, EFTPOS fees, Stripe fees include GST. GST credits CAN be claimed on these fees.',
            'url': 'https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/your-industry/financial-services-and-insurance'
        }
    },
    'life_insurance_personal': {
        'query': 'life insurance income protection business deduction owner site:ato.gov.au',
        'fallback': {
            'ruling': 'ATO - Insurance premiums deductions',
            'title': 'Life insurance is not a business deduction',
            'summary': 'Life/income protection insurance for the business owner is NOT a deductible business expense. Recode to Owner Drawings. Owner may claim income protection on personal tax return.',
            'url': 'https://www.ato.gov.au/individuals-and-families/income-deductions-offsets-and-records/deductions-you-can-claim/other-deductions/income-protection-insurance'
        }
    },
    'body_corporate_fees': {
        'query': 'body corporate strata fees GST commercial residential site:ato.gov.au',
        'fallback': {
            'ruling': 'GST Act - Division 40',
            'title': 'Body corporate fees GST treatment',
            'summary': 'Body corporate/strata fees for commercial property include GST (credits claimable). Residential strata fees are input-taxed (no GST credits).',
            'url': 'https://www.ato.gov.au/businesses-and-organisations/gst-excise-and-indirect-taxes/gst/in-detail/your-industry/property'
        }
    }
}


def search_ato_ruling(issue_type):
    """
    Search ATO website for relevant ruling information.
    Returns cached result if available, otherwise performs web search.

    Args:
        issue_type: The type of GST issue (e.g., 'export_gst_error', 'entertainment')

    Returns:
        dict with 'ruling', 'title', 'summary', 'url' keys
    """
    global _ato_ruling_cache

    # Check cache first
    if issue_type in _ato_ruling_cache:
        return _ato_ruling_cache[issue_type]

    # Get query config for this issue type
    config = ATO_RULING_QUERIES.get(issue_type)
    if not config:
        return None

    try:
        # Perform web search for ATO ruling
        search_query = config['query']

        # Use requests to search (simplified approach)
        # In production, you might use a proper search API
        search_url = f"https://www.google.com/search?q={requests.utils.quote(search_query)}"

        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }

        response = requests.get(search_url, headers=headers, timeout=5)

        if response.status_code == 200:
            # Try to extract ATO URLs from search results
            ato_urls = re.findall(r'https?://(?:www\.)?ato\.gov\.au[^\s"\'<>]+', response.text)

            if ato_urls:
                # Found ATO URL, try to fetch and extract info
                ato_url = ato_urls[0].rstrip('.')

                # Fetch the ATO page
                ato_response = requests.get(ato_url, headers=headers, timeout=5)

                if ato_response.status_code == 200:
                    # Extract title from page
                    title_match = re.search(r'<title>([^<]+)</title>', ato_response.text, re.IGNORECASE)
                    title = title_match.group(1) if title_match else config['fallback']['title']

                    # Extract ruling number if present
                    ruling_match = re.search(r'(GSTR\s*\d{4}/\d+|TR\s*\d{2,4}/\d+|TD\s*\d{4}/\d+)', ato_response.text)
                    ruling = ruling_match.group(1) if ruling_match else config['fallback']['ruling']

                    result = {
                        'ruling': ruling,
                        'title': title.strip()[:100],  # Limit title length
                        'summary': config['fallback']['summary'],  # Use curated summary
                        'url': ato_url,
                        'source': 'live'
                    }

                    _ato_ruling_cache[issue_type] = result
                    return result

        # Fall through to fallback

    except Exception as e:
        print(f"ATO ruling search error for {issue_type}: {e}")

    # Use fallback if search fails
    fallback = config['fallback'].copy()
    fallback['source'] = 'fallback'
    _ato_ruling_cache[issue_type] = fallback
    return fallback


def get_ato_ruling_comment(issue_type, base_comment):
    """
    Enhance a base comment with ATO ruling reference.

    Args:
        issue_type: The type of GST issue
        base_comment: The base comment to enhance

    Returns:
        Enhanced comment with ATO ruling reference
    """
    ruling_info = search_ato_ruling(issue_type)

    if ruling_info:
        return f"{base_comment} Per {ruling_info['ruling']}: {ruling_info['summary']}"

    return base_comment


def clear_ato_ruling_cache():
    """Clear the ATO ruling cache (call at start of new review session)"""
    global _ato_ruling_cache
    _ato_ruling_cache = {}


def generate_ato_comment(issue_type, transaction=None):
    """
    Generate a comment with ATO ruling reference for a specific issue type.

    Args:
        issue_type: The type of GST issue
        transaction: Optional transaction dict for context

    Returns:
        Formatted comment string with ATO ruling reference
    """
    ruling_info = search_ato_ruling(issue_type)

    if not ruling_info:
        return None

    # Base comments for each issue type
    base_comments = {
        'export_gst_error': 'Export sale with GST charged - exports should be GST-FREE.',
        'entertainment': 'Entertainment expense - NO GST credit claimable.',
        'client_entertainment_gst': 'Client entertainment - NO GST credit claimable.',
        'staff_entertainment_gst': 'Staff entertainment - NO GST credit unless FBT paid.',
        'wages_gst_error': 'Wages/salaries/super incorrectly coded with GST - should be BAS Excluded.',
        'government_charges_gst': 'Government charge (council rates, stamp duty, land tax) - NO GST. These are government taxes under Division 81, not taxable supplies.',
        'grants_sponsorship_gst': 'Grant income GST treatment requires review.',
        'residential_premises_gst': 'Residential property expense - Input Taxed (no GST credit).',
        'input_taxed_gst_error': 'Financial supply - Input Taxed (no GST credit claimable).',
        'insurance_gst_error': 'Life/income protection insurance - NOT a deductible business expense. Recode to Owner Drawings. No GST credit (input-taxed).',
        'travel_gst_international': 'International travel - GST-FREE (no GST credits).',
        'travel_gst_domestic': 'Domestic travel - TAXABLE (should include GST).',
        'fines_penalties_gst': 'Fine/penalty - BAS Excluded (no GST).',
        'asset_capitalization_error': 'Asset over threshold - should be capitalized.',
        'donations_gst': 'Donation - NO GST applies.',
        'paypal_fees': 'PayPal fees - NO GST. PayPal (Singapore) does not charge GST. Recode to Input Taxed.',
        'merchant_fees': 'Merchant/Stripe/eBay fees - GST INCLUDED. Credits can be claimed. Recode to GST on Expenses.',
        'life_insurance_personal': 'Life/income protection insurance - NOT a deductible business expense. Recode to Owner Drawings.',
        'body_corporate_fees': 'Body corporate/strata fees - commercial property GST claimable, residential input-taxed.'
    }

    base = base_comments.get(issue_type, 'GST treatment requires review.')

    # Format: Base comment | Per [Ruling]: [Summary]
    return f"{base} Per {ruling_info['ruling']}: {ruling_info['summary']}"


def review_with_ai(transaction):
    """Review transaction with DeepSeek AI based on ATO GST rules"""
    if not DEEPSEEK_API_KEY:
        # Return basic review without AI
        has_issues = (
            transaction.get('account_coding_suspicious') or
            transaction.get('alcohol_gst_error') or
            transaction.get('input_taxed_gst_error') or
            transaction.get('missing_gst_error') or
            not transaction.get('gst_calculation_correct', True) or
            transaction.get('drawings_loan_error') or
            transaction.get('personal_in_business_account') or
            transaction.get('asset_capitalization_error') or
            transaction.get('computer_equipment_expense') or
            transaction.get('residential_premises_gst') or
            transaction.get('insurance_gst_error') or
            transaction.get('grants_sponsorship_gst') or
            transaction.get('wages_gst_error') or
            transaction.get('allowance_gst_error') or
            transaction.get('reimbursement_gst') or
            transaction.get('voucher_gst') or
            transaction.get('general_expenses') or
            transaction.get('travel_gst') or
            transaction.get('payment_processor_fees') or
            transaction.get('fines_penalties_gst') or
            transaction.get('donations_gst') or
            transaction.get('property_gst_withholding') or
            transaction.get('livestock_gst') or
            transaction.get('asset_disposal_gst') or
            transaction.get('export_gst_error') or
            transaction.get('interest_gst_error') or
            transaction.get('other_income_error') or
            transaction.get('sales_gst_error')
        )
        return {
            'has_issues': has_issues,
            'severity': 'high' if has_issues else 'low',
            'comments': 'Manual review required - AI not configured' if has_issues else '',
            'issues': []
        }

    prompt = f"""Review this Australian transaction for BAS compliance per ATO rules:

CRITICAL RULES - NEVER VIOLATE THESE:
1. NEVER mention "duplicate" transactions. You are reviewing ONE transaction in isolation. You have NO knowledge of other transactions.
2. Monthly subscriptions (Xero, software, utilities) with same amount each month are NORMAL recurring charges.

Date: {transaction.get('date')}
Account: {transaction.get('account_code')} - {transaction.get('account')}
Description: {transaction.get('description')}
Gross: ${transaction.get('gross', 0):,.2f}
GST: ${transaction.get('gst', 0):,.2f}
Net: ${transaction.get('net', 0):,.2f}
GST Rate: {transaction.get('gst_rate_name')}

IMPORTANT GST CALCULATION RULE:
For GST-inclusive amounts in Australia, GST = Gross ÷ 11 (NOT Gross × 10%).
The GST amount has been VERIFIED as mathematically correct. Do NOT mention GST calculation at all - it is correct. Only mention account coding or other issues.

Pre-checks (VERIFIED - do not contradict these):
- Account coding suspicious: {transaction.get('account_coding_suspicious', False)}
- Entertainment alcohol GST error: {transaction.get('alcohol_gst_error', False)}
- Input-taxed GST error (GST claimed on financial supply): {transaction.get('input_taxed_gst_error', False)}
- Missing GST error (should have GST but coded GST Free): {transaction.get('missing_gst_error', False)}
- GST calculation: {'VERIFIED CORRECT - the GST math (Gross÷11) is accurate, do NOT say it is incorrect or suspicious' if transaction.get('gst_calculation_correct', True) else 'INCORRECT - GST amount does not match expected calculation'}
- Drawings/Loan error (should be BAS Excluded): {transaction.get('drawings_loan_error', False)}
- Asset capitalization error (over $20k threshold): {transaction.get('asset_capitalization_error', False)}
- Computer equipment coded to expense (should be asset): {transaction.get('computer_equipment_expense', False)}
- Residential property expense GST error (input-taxed, no GST credit): {transaction.get('residential_premises_gst', False)}
- Life/income protection insurance GST error (input-taxed): {transaction.get('insurance_gst_error', False)}
- Grants/sponsorship GST error: {transaction.get('grants_sponsorship_gst', False)}
- Wages GST error (wages have no GST): {transaction.get('wages_gst_error', False)}
- Allowance GST error (allowances have no GST): {transaction.get('allowance_gst_error', False)}
- Reimbursement GST check (verify tax invoice for >$82.50): {transaction.get('reimbursement_gst', False)}
- Voucher/gift card GST (face value=no GST at sale, non-face value=GST at sale): {transaction.get('voucher_gst', False)}
- General/Sundry Expenses (audit risk, recode to specific category): {transaction.get('general_expenses', False)}
- Travel GST error (international=GST-free, domestic=taxable): {transaction.get('travel_gst', False)}
- Payment processor fees (PayPal=no GST, Stripe/eBay/bank=GST included): {transaction.get('payment_processor_fees', False)}
- Fines/penalties GST error (non-reportable): {transaction.get('fines_penalties_gst', False)}
- Donations GST error (non-reportable): {transaction.get('donations_gst', False)}
- Property GST withholding (new residential/land): {transaction.get('property_gst_withholding', False)}
- Livestock GST error (live animals are taxable): {transaction.get('livestock_gst', False)}
- Asset disposal GST error (business assets are taxable): {transaction.get('asset_disposal_gst', False)}
- Interest GST error (should be GST Free): {transaction.get('interest_gst_error', False)}
- Other income GST error: {transaction.get('other_income_error', False)}
- Sales GST error (invalid GST Free or BAS Excluded): {transaction.get('sales_gst_error', False)}
- Export GST error (exports should be GST-free): {transaction.get('export_gst_error', False)}

ATO GST Rules to check:
1. GST-FREE (no GST, but CAN claim input credits): Basic food, health/medical, education, childcare, exports
2. INPUT-TAXED (no GST, CANNOT claim input credits): Bank ACCOUNT fees (monthly fees, overdraft fees), interest income/expense, residential rent, life insurance
3. BAS EXCLUDED (outside the GST system - not reportable on BAS): Wages, salaries, superannuation, PAYG withholding, allowances - these are NOT "input-taxed", they are simply outside the GST system entirely. Should have $0 GST.
4. TAXABLE (10% GST applies): Office supplies, utilities, parking, fuel, professional services, commercial rent
5. MERCHANT FEES (TAXABLE - GST INCLUDED): Bank merchant fees, EFTPOS fees, credit card processing fees, merchant facility fees - these are NOT input-taxed! Businesses CAN claim GST credits on merchant fees. Source: ATO Financial Services Q&A.
6. OVERSEAS DIGITAL SERVICES (GST FREE IS CORRECT): Adobe, Slack, Zoom, Canva, Google Ads, Facebook Ads, Microsoft 365, AWS, Dropbox, and other overseas SaaS - when billed from USA/Ireland/Singapore WITHOUT GST, code as GST Free. Do NOT flag these as missing GST. Reverse charge may apply but results in net zero for most businesses.
7. ENTERTAINMENT: Non-deductible, NO GST credits (includes alcohol in social context)
8. RESIDENTIAL PROPERTY: Input-taxed - NO GST credit on property management, repairs, maintenance, agent fees. IMPORTANT: Body corporate fees, strata fees, and owners corporation fees can be for EITHER residential OR commercial properties - do NOT assume they are residential unless the transaction explicitly says "residential". Commercial strata fees DO have GST credits claimable.
9. Software subscriptions should be coded to Subscriptions, NOT Consulting
10. Parking should be coded to Motor Vehicle, NOT Legal Expenses
11. Office supplies (toner, cartridges) MUST include GST (10%)
12. GRANTS (per GSTR 2012/2): "Other Income" is a VALID account for grants - do NOT flag as wrong account. Grants are typically GST-FREE unless there's a binding obligation to provide specific services/goods in return. Only flag GST treatment if GST is charged on a grant that appears to have no supply obligation.
13. GOVERNMENT CHARGES (NO GST - not a taxable supply): Council rates, stamp duty, land tax, water rates, motor vehicle registration, ASIC fees, court fees, licence fees - these are government levies, NOT taxable supplies. They should have $0 GST. IMPORTANT: Council rates have NO GST regardless of whether the property is residential OR commercial - do NOT mention "residential property" for council rates. Government charges are simply not taxable supplies.

If issues found, respond with specific problems and ATO rule reference. If OK, respond "OK - Transaction appears correct"
"""

    try:
        response = requests.post(
            'https://api.deepseek.com/v1/chat/completions',
            headers={
                'Authorization': f'Bearer {DEEPSEEK_API_KEY}',
                'Content-Type': 'application/json'
            },
            json={
                'model': 'deepseek-chat',
                'messages': [
                    {'role': 'system', 'content': 'You are an Australian tax accountant reviewing BAS transactions. Be critical and flag issues.'},
                    {'role': 'user', 'content': prompt}
                ],
                'temperature': 0.3,
                'max_tokens': 500
            },
            timeout=30
        )

        if response.status_code == 200:
            result = response.json()
            ai_response = result['choices'][0]['message']['content']

            response_lower = ai_response.lower()
            has_issues = not ('ok -' in response_lower or 'appears correct' in response_lower or 'no issues' in response_lower)

            if has_issues:
                if any(word in response_lower for word in ['critical', 'must', 'incorrect', 'error']):
                    severity = 'high'
                elif any(word in response_lower for word in ['should', 'review', 'check', 'unusual']):
                    severity = 'medium'
                else:
                    severity = 'low'
            else:
                severity = 'low'

            # Filter out OK comments - only show issues
            filtered_comment = ''
            if has_issues:
                filtered_comment = ai_response
            else:
                # Check if there's useful info beyond just "OK"
                # Remove OK prefix and check if there's substance
                cleaned = ai_response.strip()
                for prefix in ['OK -', 'OK:', 'OK.', 'OK,', 'OK ']:
                    if cleaned.upper().startswith(prefix.upper()):
                        cleaned = cleaned[len(prefix):].strip()
                        break
                # If there's no substantial content after removing OK, leave empty
                if len(cleaned) < 20 or 'appears correct' in cleaned.lower() or 'correctly coded' in cleaned.lower():
                    filtered_comment = ''
                else:
                    filtered_comment = cleaned

            return {
                'has_issues': has_issues,
                'severity': severity,
                'comments': filtered_comment,
                'issues': []
            }
    except Exception as e:
        print(f"AI review error: {e}")

    # Fallback to rule-based review
    has_issues = (
        transaction.get('account_coding_suspicious') or
        transaction.get('alcohol_gst_error') or
        transaction.get('input_taxed_gst_error') or
        transaction.get('missing_gst_error') or
        not transaction.get('gst_calculation_correct', True)
    )
    return {
        'has_issues': has_issues,
        'severity': 'high' if has_issues else 'low',
        'comments': 'AI review unavailable - rule-based check performed',
        'issues': []
    }


def process_single_batch(batch_data):
    """Process a single batch of transactions - used for parallel processing"""
    batch_index, batch, api_key = batch_data

    # Build batch prompt
    batch_prompt = """Review these Australian transactions for BAS compliance per ATO rules.
For EACH transaction, provide a brief assessment.

CRITICAL RULES:
1. NEVER flag transactions as "duplicates" or mention "double claim". Recurring charges are NORMAL.
2. GST calculations have been verified correct. Do NOT mention GST amounts or calculations - only mention account coding or other issues.

ATO GST Rules:
1. GST-FREE (no GST, CAN claim input credits): Basic food, health/medical, education, exports
2. INPUT-TAXED (no GST, CANNOT claim credits): Bank fees, interest, residential rent, life insurance
3. TAXABLE (10% GST): Office supplies, utilities, parking, fuel, professional services
4. ENTERTAINMENT: Non-deductible, NO GST credits
5. RESIDENTIAL PROPERTY: Input-taxed - NO GST credit. Body corporate/strata fees can be residential OR commercial - do NOT assume residential.
6. GRANTS (GSTR 2012/2): "Other Income" is VALID for grants - do NOT flag account. Grants typically GST-FREE unless binding supply obligation exists.

TRANSACTIONS TO REVIEW:
"""
    for idx, t in enumerate(batch, 1):
        flags = []
        if t.get('account_coding_suspicious'): flags.append('suspicious coding')
        if t.get('overseas_subscription_gst'): flags.append('overseas subscription')
        if t.get('reimbursement_gst'): flags.append('reimbursement >$82.50')
        if t.get('general_expenses'): flags.append('general expenses')
        if t.get('travel_gst'): flags.append('travel GST')
        if t.get('payment_processor_fees'): flags.append('payment processor fees')
        if t.get('missing_gst_error'): flags.append('missing GST')

        batch_prompt += f"""
---
Transaction {idx}:
Account: {t.get('account_code')} - {t.get('account')}
Description: {t.get('description', '')[:100]}
Gross: ${t.get('gross', 0):,.2f} | GST: ${t.get('gst', 0):,.2f} | Rate: {t.get('gst_rate_name')}
Flags: {', '.join(flags) if flags else 'None'}
"""

    batch_prompt += """
---
For each transaction, respond in this format:
Transaction 1: [OK or ISSUE: brief description]
Transaction 2: [OK or ISSUE: brief description]
...etc
"""

    try:
        response = requests.post(
            'https://api.deepseek.com/v1/chat/completions',
            headers={
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json'
            },
            json={
                'model': 'deepseek-chat',
                'messages': [
                    {'role': 'system', 'content': 'You are an Australian tax accountant reviewing BAS transactions. Be concise. For each transaction, respond with OK or ISSUE followed by a brief explanation. NEVER mention duplicates - recurring charges are normal.'},
                    {'role': 'user', 'content': batch_prompt}
                ],
                'temperature': 0.3,
                'max_tokens': 1000
            },
            timeout=60
        )

        if response.status_code == 200:
            result = response.json()
            ai_response = result['choices'][0]['message']['content']

            # Parse batch response
            lines = ai_response.strip().split('\n')
            batch_results = []
            current_result = None

            for line in lines:
                line = line.strip()
                if not line:
                    continue

                # Check if this is a new transaction line
                for i in range(1, len(batch) + 1):
                    if line.lower().startswith(f'transaction {i}:') or line.startswith(f'{i}:') or line.startswith(f'{i}.'):
                        if current_result:
                            batch_results.append(current_result)

                        response_text = line.split(':', 1)[-1].strip() if ':' in line else line
                        response_lower = response_text.lower()
                        has_issues = not ('ok' in response_lower[:10] and 'issue' not in response_lower[:20])

                        if has_issues:
                            if any(word in response_lower for word in ['critical', 'must', 'incorrect', 'error']):
                                severity = 'high'
                            elif any(word in response_lower for word in ['should', 'review', 'check']):
                                severity = 'medium'
                            else:
                                severity = 'low'
                        else:
                            severity = 'low'

                        # Filter out OK comments - only show issues
                        filtered_comment = ''
                        if has_issues:
                            filtered_comment = response_text
                        else:
                            # Remove OK prefix and check if there's substance
                            cleaned = response_text.strip()
                            for prefix in ['OK -', 'OK:', 'OK.', 'OK,', 'OK ', 'ok -', 'ok:', 'ok.', 'ok,', 'ok ']:
                                if cleaned.startswith(prefix):
                                    cleaned = cleaned[len(prefix):].strip()
                                    break
                            # If there's no substantial content, leave empty
                            if len(cleaned) < 20 or 'appears correct' in cleaned.lower() or 'correctly coded' in cleaned.lower():
                                filtered_comment = ''
                            else:
                                filtered_comment = cleaned

                        current_result = {
                            'has_issues': has_issues,
                            'severity': severity,
                            'comments': filtered_comment,
                            'issues': []
                        }
                        break

            if current_result:
                batch_results.append(current_result)

            # Fill in any missing results
            while len(batch_results) < len(batch):
                batch_results.append({
                    'has_issues': True,
                    'severity': 'medium',
                    'comments': 'AI batch response incomplete - manual review required',
                    'issues': []
                })

            return (batch_index, batch_results[:len(batch)])
        else:
            # API error - fall back to basic results
            fallback = []
            for t in batch:
                has_issues = t.get('account_coding_suspicious') or t.get('missing_gst_error')
                fallback.append({
                    'has_issues': has_issues,
                    'severity': 'high' if has_issues else 'low',
                    'comments': 'AI batch review failed - rule-based check performed',
                    'issues': []
                })
            return (batch_index, fallback)

    except Exception as e:
        print(f"Batch {batch_index} AI review error: {e}")
        # Fall back to basic results for this batch
        fallback = []
        for t in batch:
            has_issues = t.get('account_coding_suspicious') or t.get('missing_gst_error')
            fallback.append({
                'has_issues': has_issues,
                'severity': 'high' if has_issues else 'low',
                'comments': 'AI batch review error - rule-based check performed',
                'issues': []
            })
        return (batch_index, fallback)


def review_batch_with_ai(transactions, batch_size=5, max_workers=4):
    """Review multiple transactions in parallel batches with DeepSeek AI for faster processing"""
    if not DEEPSEEK_API_KEY or not transactions:
        # Return basic review without AI for all transactions
        results = []
        for transaction in transactions:
            has_issues = (
                transaction.get('account_coding_suspicious') or
                transaction.get('alcohol_gst_error') or
                transaction.get('input_taxed_gst_error') or
                transaction.get('missing_gst_error') or
                not transaction.get('gst_calculation_correct', True)
            )
            results.append({
                'has_issues': has_issues,
                'severity': 'high' if has_issues else 'low',
                'comments': 'Manual review required - AI not configured' if has_issues else '',
                'issues': []
            })
        return results

    # Prepare batches for parallel processing
    batches = []
    for batch_start in range(0, len(transactions), batch_size):
        batch = transactions[batch_start:batch_start + batch_size]
        batch_index = batch_start // batch_size
        batches.append((batch_index, batch, DEEPSEEK_API_KEY))

    # Process batches in parallel (4 concurrent API calls)
    batch_results = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_single_batch, batch_data): batch_data[0] for batch_data in batches}

        for future in as_completed(futures):
            try:
                batch_index, results = future.result()
                batch_results[batch_index] = results
            except Exception as e:
                print(f"Parallel batch error: {e}")
                batch_index = futures[future]
                batch_results[batch_index] = [{'has_issues': True, 'severity': 'high', 'comments': 'Parallel processing error', 'issues': []}] * batch_size

    # Reassemble results in order
    all_results = []
    for i in range(len(batches)):
        if i in batch_results:
            all_results.extend(batch_results[i])
        else:
            # Fallback for missing batches
            batch_size_actual = len(batches[i][1])
            all_results.extend([{'has_issues': True, 'severity': 'high', 'comments': 'Batch missing', 'issues': []}] * batch_size_actual)

    # Trim to exact transaction count
    return all_results[:len(transactions)]


@app.route('/download-report', methods=['GET', 'POST'])
@login_required
def download_report():
    """Download review results as Excel with formatting"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    results = None

    # Try POST data first (more reliable), then session
    if request.method == 'POST':
        results = request.json
    elif 'review_results' in session:
        results = session['review_results']

    if not results:
        return jsonify({'error': 'No review results found'}), 404

    try:
        # Create Excel file
        output = BytesIO()
        company_name = results.get('tenant_name', results.get('company_name', 'Unknown'))

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = [
                ['BAS REVIEW REPORT', ''],
                ['', ''],
                ['Company', company_name],
                ['Review Date', results.get('review_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))],
                ['', ''],
                ['Total Transactions', results.get('total_transactions', len(results.get('transactions', [])))],
                ['Flagged Items', results.get('flagged_count', len(results.get('flagged_items', [])))]
            ]
            summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

            # Format summary sheet
            ws_summary = writer.sheets['Summary']
            ws_summary['A1'].font = Font(bold=True, size=16)
            ws_summary.column_dimensions['A'].width = 20
            ws_summary.column_dimensions['B'].width = 40

            # Flagged items sheet
            flagged_items = results.get('flagged_items', [])
            if flagged_items:
                # Format comments - replace pipe with newlines
                flagged_data = []
                for item in flagged_items:
                    comments = item.get('comments', '')
                    # Split by pipe and join with newlines
                    formatted_comments = comments.replace(' | ', '\n• ')
                    if formatted_comments and not formatted_comments.startswith('•'):
                        formatted_comments = '• ' + formatted_comments

                    flagged_data.append({
                        'Row': item.get('row_number', ''),
                        'Severity': (item.get('severity', 'medium')).upper(),
                        'Date': item.get('date', ''),
                        'Account Code': item.get('account_code', ''),
                        'Account Name': item.get('account', ''),
                        'Description': item.get('description', ''),
                        'Gross': item.get('gross', 0),
                        'GST': item.get('gst', 0),
                        'Net': item.get('net', 0),
                        'GST Code': item.get('gst_rate_name', ''),
                        'Comments': formatted_comments
                    })

                flagged_df = pd.DataFrame(flagged_data)
                flagged_df.to_excel(writer, sheet_name='Flagged Items', index=False)

                # Format flagged items sheet
                ws = writer.sheets['Flagged Items']

                # Column widths
                col_widths = {
                    'A': 8,   # Row
                    'B': 10,  # Severity
                    'C': 12,  # Date
                    'D': 12,  # Account Code
                    'E': 25,  # Account Name
                    'F': 40,  # Description
                    'G': 12,  # Gross
                    'H': 12,  # GST
                    'I': 12,  # Net
                    'J': 15,  # GST Code
                    'K': 50,  # Comments
                }
                for col, width in col_widths.items():
                    ws.column_dimensions[col].width = width

                # Header formatting
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for col in range(1, 12):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

                # Severity colors and cell formatting
                severity_colors = {
                    'HIGH': PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'),
                    'MEDIUM': PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'),
                    'LOW': PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'),
                }

                for row in range(2, len(flagged_items) + 2):
                    severity = ws.cell(row=row, column=2).value
                    if severity in severity_colors:
                        ws.cell(row=row, column=2).fill = severity_colors[severity]

                    # Format currency columns
                    for col in [7, 8, 9]:  # Gross, GST, Net
                        cell = ws.cell(row=row, column=col)
                        cell.number_format = '$#,##0.00'
                        cell.alignment = Alignment(horizontal='right')

                    # Wrap text for comments
                    ws.cell(row=row, column=11).alignment = Alignment(wrap_text=True, vertical='top')

                    # Wrap text for description
                    ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical='top')

                    # Calculate row height based on content
                    comments_cell = ws.cell(row=row, column=11).value or ''
                    desc_cell = ws.cell(row=row, column=6).value or ''

                    # Count lines in comments and description
                    comment_lines = comments_cell.count('\n') + 1
                    desc_lines = max(1, len(str(desc_cell)) // 50)  # Approximate lines based on length

                    # Calculate height: ~15 points per line, minimum 30
                    max_lines = max(comment_lines, desc_lines)
                    row_height = max(30, max_lines * 15)
                    ws.row_dimensions[row].height = row_height

                    # Add borders to all cells
                    for col in range(1, 12):
                        ws.cell(row=row, column=col).border = thin_border

                # Create Correcting Journals sheet
                journal_data = []
                for item in flagged_items:
                    journal = item.get('correcting_journal', {})
                    entries = journal.get('entries', [])
                    if entries:
                        # Add a header row for this correction
                        journal_data.append({
                            'Original Row': item.get('row_number', ''),
                            'Date': item.get('date', ''),
                            'Narration': journal.get('narration', ''),
                            'Account Code': '',
                            'Account Name': '',
                            'Debit': '',
                            'Credit': '',
                            'Tax Code': '',
                            'Line Description': ''
                        })
                        # Add each journal line
                        for entry in entries:
                            journal_data.append({
                                'Original Row': '',
                                'Date': '',
                                'Narration': '',
                                'Account Code': entry.get('account_code', ''),
                                'Account Name': entry.get('account_name', ''),
                                'Debit': entry.get('debit', 0) if entry.get('debit', 0) > 0 else '',
                                'Credit': entry.get('credit', 0) if entry.get('credit', 0) > 0 else '',
                                'Tax Code': entry.get('tax_code', ''),
                                'Line Description': entry.get('description', '')
                            })
                        # Add blank row between journals
                        journal_data.append({col: '' for col in ['Original Row', 'Date', 'Narration', 'Account Code', 'Account Name', 'Debit', 'Credit', 'Tax Code', 'Line Description']})

                if journal_data:
                    journal_df = pd.DataFrame(journal_data)
                    journal_df.to_excel(writer, sheet_name='Correcting Journals', index=False)

                    # Format journal sheet
                    ws_journal = writer.sheets['Correcting Journals']

                    # Column widths
                    journal_col_widths = {'A': 10, 'B': 12, 'C': 40, 'D': 12, 'E': 25, 'F': 12, 'G': 12, 'H': 15, 'I': 35}
                    for col, width in journal_col_widths.items():
                        ws_journal.column_dimensions[col].width = width

                    # Header formatting
                    for col in range(1, 10):
                        cell = ws_journal.cell(row=1, column=col)
                        cell.fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.alignment = Alignment(horizontal='center')

                    # Format currency columns and highlight narration rows
                    for row in range(2, len(journal_data) + 2):
                        # If this is a narration row (has original row number)
                        if ws_journal.cell(row=row, column=1).value:
                            for col in range(1, 10):
                                ws_journal.cell(row=row, column=col).fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
                                ws_journal.cell(row=row, column=col).font = Font(bold=True)

                        # Format debit/credit columns
                        for col in [6, 7]:
                            cell = ws_journal.cell(row=row, column=col)
                            if cell.value and cell.value != '':
                                cell.number_format = '$#,##0.00'
                                cell.alignment = Alignment(horizontal='right')

            else:
                # Create empty flagged items sheet with message
                empty_df = pd.DataFrame([{'Message': 'No issues found - all transactions appear correct'}])
                empty_df.to_excel(writer, sheet_name='Flagged Items', index=False)
                ws = writer.sheets['Flagged Items']
                ws['A1'].font = Font(bold=True, color='28A745')
                ws.column_dimensions['A'].width = 60

        output.seek(0)

        company_name_clean = company_name.replace(' ', '_').replace('/', '_')
        filename = f"BAS_Review_{company_name_clean}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
