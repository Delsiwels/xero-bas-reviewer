"""
Output generator for BAS review results
"""
import os
from typing import Dict, Any, List
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class OutputGenerator:
    """Generate formatted output reports for BAS review"""

    def __init__(self, review_results: Dict[str, Any], original_file: str):
        self.results = review_results
        self.original_file = original_file

    def generate_excel_report(self) -> str:
        """
        Generate Excel report with flagged items

        Returns:
            Path to the generated report file
        """
        # Create output filename
        base_name = os.path.splitext(os.path.basename(self.original_file))[0]
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f"{base_name}_REVIEW_{timestamp}.xlsx"

        # Create DataFrames for different sheets
        summary_df = self._create_summary_dataframe()
        flagged_df = self._create_flagged_items_dataframe()

        # Write to Excel with multiple sheets
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Summary sheet
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

            # Flagged items sheet
            if not flagged_df.empty:
                flagged_df.to_excel(writer, sheet_name='Flagged Items', index=False)
            else:
                # Create empty sheet with message
                no_issues_df = pd.DataFrame([['No issues found - all transactions appear correct']])
                no_issues_df.to_excel(writer, sheet_name='Flagged Items', index=False, header=False)

        # Apply formatting
        self._apply_formatting(output_file)

        return output_file

    def _create_summary_dataframe(self) -> pd.DataFrame:
        """Create summary DataFrame"""
        summary_data = []

        # Review metadata
        summary_data.append(['REVIEW SUMMARY', ''])
        summary_data.append(['Review Date', self.results.get('review_date', '')])
        summary_data.append(['Company', self.results['metadata'].get('company_name', '')])
        summary_data.append(['Period', self.results['metadata'].get('period', '')])
        summary_data.append(['', ''])

        # Transaction counts
        summary_data.append(['TRANSACTION SUMMARY', ''])
        summary_data.append(['Total Transactions Reviewed', self.results.get('total_reviewed', 0)])
        summary_data.append(['Transactions Flagged', self.results.get('flagged_count', 0)])
        summary_data.append(['', ''])

        # Financial summary
        summary = self.results.get('summary', {})
        summary_data.append(['FINANCIAL SUMMARY', ''])
        summary_data.append(['Total Sales', f"${summary.get('total_sales', 0):,.2f}"])
        summary_data.append(['Total Purchases', f"${summary.get('total_purchases', 0):,.2f}"])
        summary_data.append(['GST Collected', f"${summary.get('total_gst_collected', 0):,.2f}"])
        summary_data.append(['GST Paid', f"${summary.get('total_gst_paid', 0):,.2f}"])
        summary_data.append(['Net GST', f"${summary.get('net_gst', 0):,.2f}"])
        summary_data.append(['', ''])

        # Severity breakdown
        if self.results.get('flagged_count', 0) > 0:
            summary_data.append(['ISSUES BY SEVERITY', ''])
            severity_counts = self._count_by_severity()
            for severity, count in severity_counts.items():
                if count > 0:
                    summary_data.append([severity.upper(), count])

        return pd.DataFrame(summary_data, columns=['Metric', 'Value'])

    def _create_flagged_items_dataframe(self) -> pd.DataFrame:
        """Create DataFrame for flagged items"""
        if not self.results.get('flagged_items'):
            return pd.DataFrame()

        flagged_data = []

        for item in self.results['flagged_items']:
            # Get amounts - handle both old and new field names
            gross_amount = item.get('amount', item.get('gross', 0))
            gst_amount = item.get('gst_amount', item.get('gst', 0))
            net_amount = item.get('net_amount', item.get('net', 0))

            flagged_data.append({
                'Row': item.get('row_number', ''),
                'Severity': item.get('severity', '').upper(),
                'Date': item.get('date', ''),
                'Account': item.get('account', ''),
                'Description': item.get('description', ''),
                'Gross (Inc-GST)': f"${gross_amount:,.2f}",
                'GST Amount': f"${gst_amount:,.2f}",
                'Net Amount': f"${net_amount:,.2f}",
                'GST Code': item.get('gst_code', item.get('gst_rate_name', '')),
                'Issues': ', '.join(item.get('issues', [])),
                'AI Comments': item.get('comments', '')
            })

        return pd.DataFrame(flagged_data)

    def _count_by_severity(self) -> Dict[str, int]:
        """Count flagged items by severity"""
        counts = {'high': 0, 'medium': 0, 'low': 0, 'info': 0}

        for item in self.results.get('flagged_items', []):
            severity = item.get('severity', 'info')
            counts[severity] = counts.get(severity, 0) + 1

        return counts

    def _apply_formatting(self, filename: str):
        """Apply formatting to the Excel file"""
        wb = load_workbook(filename)

        # Format Summary sheet
        if 'Summary' in wb.sheetnames:
            ws = wb['Summary']
            self._format_summary_sheet(ws)

        # Format Flagged Items sheet
        if 'Flagged Items' in wb.sheetnames:
            ws = wb['Flagged Items']
            self._format_flagged_items_sheet(ws)

        wb.save(filename)

    def _format_summary_sheet(self, ws):
        """Apply formatting to summary sheet"""
        # Header style
        header_font = Font(bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        # Section header style
        section_font = Font(bold=True, size=11)
        section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        # Apply formatting to section headers
        section_rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            cell_value = str(row[0].value).upper()
            if any(header in cell_value for header in ['SUMMARY', 'FINANCIAL', 'ISSUES', 'TRANSACTION']):
                section_rows.append(row[0].row)

        for row_num in section_rows:
            cell = ws.cell(row=row_num, column=1)
            cell.font = section_font
            cell.fill = section_fill

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _format_flagged_items_sheet(self, ws):
        """Apply formatting to flagged items sheet"""
        if ws.max_row < 2:
            return

        # Header style
        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Apply header formatting
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Severity colors
        severity_colors = {
            'HIGH': 'FFCCCC',    # Light red
            'MEDIUM': 'FFFF99',  # Light yellow
            'LOW': 'CCFFCC',     # Light green
            'INFO': 'E6F3FF'     # Light blue
        }

        # Apply row coloring based on severity
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            severity_cell = row[1]  # Severity column
            severity = str(severity_cell.value).upper()

            if severity in severity_colors:
                fill = PatternFill(start_color=severity_colors[severity],
                                   end_color=severity_colors[severity],
                                   fill_type='solid')

                for cell in row:
                    cell.fill = fill

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Special handling for comments column (make it wider)
            if column[0].value == 'AI Comments':
                adjusted_width = 60
            else:
                adjusted_width = min(max_length + 2, 40)

            ws.column_dimensions[column_letter].width = adjusted_width

        # Wrap text in AI Comments column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            comments_cell = row[-1]  # Last column (AI Comments)
            comments_cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Freeze header row
        ws.freeze_panes = 'A2'

    def generate_text_report(self) -> str:
        """
        Generate a text-based report (for console output or text file)

        Returns:
            Formatted text report
        """
        lines = []

        lines.append("=" * 80)
        lines.append("BAS REVIEW REPORT")
        lines.append("=" * 80)
        lines.append(f"Review Date: {self.results.get('review_date', '')}")
        lines.append(f"Company: {self.results['metadata'].get('company_name', '')}")
        lines.append(f"Period: {self.results['metadata'].get('period', '')}")
        lines.append("")

        # Summary
        summary = self.results.get('summary', {})
        lines.append("FINANCIAL SUMMARY")
        lines.append("-" * 80)
        lines.append(f"Total Sales:        ${summary.get('total_sales', 0):>15,.2f}")
        lines.append(f"Total Purchases:    ${summary.get('total_purchases', 0):>15,.2f}")
        lines.append(f"GST Collected:      ${summary.get('total_gst_collected', 0):>15,.2f}")
        lines.append(f"GST Paid:           ${summary.get('total_gst_paid', 0):>15,.2f}")
        lines.append(f"Net GST:            ${summary.get('net_gst', 0):>15,.2f}")
        lines.append("")

        # Flagged items
        lines.append("FLAGGED ITEMS")
        lines.append("-" * 80)
        lines.append(f"Total Reviewed: {self.results.get('total_reviewed', 0)}")
        lines.append(f"Items Flagged:  {self.results.get('flagged_count', 0)}")
        lines.append("")

        if self.results.get('flagged_items'):
            for item in self.results['flagged_items']:
                severity_icon = self._get_severity_icon(item.get('severity', 'info'))
                lines.append(f"{severity_icon} Row {item.get('row_number', '')}: {item.get('description', '')}")
                lines.append(f"   Date: {item.get('date', '')} | Amount: ${item.get('amount', 0):,.2f}")
                lines.append(f"   Issues: {', '.join(item.get('issues', []))}")
                lines.append(f"   Comments: {item.get('comments', '')}")
                lines.append("")

        lines.append("=" * 80)

        return '\n'.join(lines)

    def _get_severity_icon(self, severity: str) -> str:
        """Get icon for severity level"""
        icons = {
            'high': '🔴',
            'medium': '🟡',
            'low': '🟢',
            'info': 'ℹ️'
        }
        return icons.get(severity, '❓')
